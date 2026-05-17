"""
epicenter_map_attributes.py
---------------------------
Заповнює prom_param_name для атрибутів Epicenter на основі ручного маппінгу.
Читає epicenter_mappings.xlsx, записує prom_param_name, підсвічує червоним обов'язкові поля.

Запуск:
    python scripts/epicenter_map_attributes.py

Алгоритм:
1. Читає аркуш "Маппінг": для кожного рядка з prom_category_id + epicenter_category_id.
2. Знаходить відповідні рядки в аркуші "Сети атрибутів" за умовою set_code == epicenter_category_id.
3. Для кожного isRequired-атрибута заповнює prom_param_name за пріоритетом:

   Порядок заповнення prom_param_name
   -------------------------------------------------------
   Ступінь 0 (пропуск):  комірка вже заповнена — не перезаписуємо.
   Ступінь 1 (hard):     назва є в аркуші attr_true_mappings → беремо звідти.
   Ступінь 2 (пусто):    не знайшло — комірка порожньою, підсвічується
                          червоним якщо isRequired=TRUE.

4. Зберігає книгу.
"""

from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

# ──────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────
XLSX_PATH = Path(__file__).parents[1] / "data" / "markets" / "epicenter_mappings.xlsx"

SHEET_MAPPING       = "Маппінг"
SHEET_ATTRS         = "Сети атрибутів"
SHEET_HARD_MAPPINGS = "attr_true_mappings"

# Columns – Маппінг
MAP_COL_PROM_CAT = "prom_category_id"
MAP_COL_EPI_CAT  = "epicenter_category_id"

# Columns – Сети атрибутів
ATTR_COL_SET_CODE    = "set_code"
ATTR_COL_ATTR_NAME   = "attr_name_uk"
ATTR_COL_PROM_PARAM  = "prom_param_name"
ATTR_COL_IS_REQUIRED = "isRequired"

RED_FILL = PatternFill("solid", start_color="FF9999", end_color="FF9999")
NO_FILL  = PatternFill(fill_type=None)


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def get_header_map(ws) -> dict[str, int]:
    """Returns {column_name: 1-based column index} from first row."""
    return {
        str(cell.value).strip(): idx + 1
        for idx, cell in enumerate(ws[1])
        if cell.value is not None
    }


def load_hard_mappings(wb: openpyxl.Workbook) -> dict[str, str]:
    """
    Лист attr_true_mappings.
    Обов'язкові колонки: attr_name_uk | prom_param_name
    Повертає: attr_name_uk → prom_param_name
    """
    if SHEET_HARD_MAPPINGS not in wb.sheetnames:
        print(f"  [warn] Sheet '{SHEET_HARD_MAPPINGS}' not found — hard mappings disabled.")
        return {}

    ws = wb[SHEET_HARD_MAPPINGS]
    headers = [
        str(c.value).strip() if c.value else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    if "attr_name_uk" not in headers or "prom_param_name" not in headers:
        print(
            f"  [warn] Sheet '{SHEET_HARD_MAPPINGS}' missing required columns "
            f"— hard mappings disabled."
        )
        return {}

    name_col  = headers.index("attr_name_uk")
    param_col = headers.index("prom_param_name")

    result = {
        row[name_col].value.strip(): row[param_col].value.strip()
        for row in ws.iter_rows(min_row=2)
        if row[name_col].value and row[param_col].value
    }
    print(f"  Loaded {len(result)} hard mappings from '{SHEET_HARD_MAPPINGS}'.")
    return result


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

def main() -> None:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"File not found: {XLSX_PATH}")

    # ── 1. Open workbook ──────────────────────
    wb = openpyxl.load_workbook(XLSX_PATH)

    hard_mappings = load_hard_mappings(wb)

    for sheet in (SHEET_MAPPING, SHEET_ATTRS):
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")

    ws_map  = wb[SHEET_MAPPING]
    ws_attr = wb[SHEET_ATTRS]

    map_hdrs  = get_header_map(ws_map)
    attr_hdrs = get_header_map(ws_attr)

    for col, sheet_name, hdrs in [
        (MAP_COL_PROM_CAT,     SHEET_MAPPING, map_hdrs),
        (MAP_COL_EPI_CAT,      SHEET_MAPPING, map_hdrs),
        (ATTR_COL_SET_CODE,    SHEET_ATTRS,   attr_hdrs),
        (ATTR_COL_ATTR_NAME,   SHEET_ATTRS,   attr_hdrs),
        (ATTR_COL_PROM_PARAM,  SHEET_ATTRS,   attr_hdrs),
        (ATTR_COL_IS_REQUIRED, SHEET_ATTRS,   attr_hdrs),
    ]:
        if col not in hdrs:
            raise ValueError(
                f"Column '{col}' not found in sheet '{sheet_name}'.\n"
                f"Found: {list(hdrs.keys())}"
            )

    col_prom_cat    = map_hdrs[MAP_COL_PROM_CAT]
    col_epi_cat     = map_hdrs[MAP_COL_EPI_CAT]
    col_set_code    = attr_hdrs[ATTR_COL_SET_CODE]
    col_attr_name   = attr_hdrs[ATTR_COL_ATTR_NAME]
    col_prom_param  = attr_hdrs[ATTR_COL_PROM_PARAM]
    col_is_required = attr_hdrs[ATTR_COL_IS_REQUIRED]

    # ── 2. Build index: set_code → [(row_idx, attr_name_uk, is_required)] ──
    attr_index: dict[str, list[tuple[int, str, bool]]] = defaultdict(list)
    for row_idx in range(2, ws_attr.max_row + 1):
        set_code    = str(ws_attr.cell(row_idx, col_set_code).value or "").strip()
        attr_name   = str(ws_attr.cell(row_idx, col_attr_name).value or "").strip()
        is_required = (
            str(ws_attr.cell(row_idx, col_is_required).value or "").strip().upper() == "TRUE"
        )
        if set_code and attr_name:
            attr_index[set_code].append((row_idx, attr_name, is_required))

    print(
        f"Loaded {sum(len(v) for v in attr_index.values())} attribute rows "
        f"across {len(attr_index)} set_codes.\n"
    )

    # ── 3. Fill prom_param_name: hard mapping only ────────────────────────────
    total_written       = 0
    total_skipped       = 0
    processed_epi_cats: set[str] = set()

    for map_row in range(2, ws_map.max_row + 1):
        prom_cat = str(ws_map.cell(map_row, col_prom_cat).value or "").strip()
        epi_cat  = str(ws_map.cell(map_row, col_epi_cat).value or "").strip()

        if not prom_cat or not epi_cat:
            continue

        epi_attrs = attr_index.get(epi_cat, [])
        if not epi_attrs:
            print(f"  [row {map_row}] epi_cat={epi_cat} — not found in Сети атрибутів, skipping")
            continue

        required_attrs = [(r, n, req) for r, n, req in epi_attrs if req]

        print(
            f"[row {map_row}] prom_cat={prom_cat} → epi_cat={epi_cat} "
            f"| {len(epi_attrs)} epi attrs ({len(required_attrs)} required)"
        )

        if not required_attrs:
            continue

        processed_epi_cats.add(epi_cat)

        for attr_row, attr_name, _ in required_attrs:
            cell = ws_attr.cell(attr_row, col_prom_param)

            # Ступінь 0: вже заповнено — не перезаписуємо
            if str(cell.value or "").strip():
                total_skipped += 1
                continue

            # Ступінь 1: hard mapping
            match = hard_mappings.get(attr_name)
            if match:
                cell.value = match
                print(f"    ✓ [hard] '{attr_name}' → '{match}'")
                total_written += 1
            else:
                # Ступінь 2: не знайшло — залишаємо порожнім
                total_skipped += 1

    # ── 4. Red highlight: isRequired + empty + matched category ──────────────
    red_count = 0
    for set_code, attrs in attr_index.items():
        if set_code not in processed_epi_cats:
            continue
        for attr_row, attr_name, is_required in attrs:
            if not is_required:
                continue
            prom_cell   = ws_attr.cell(attr_row, col_prom_param)
            prom_filled = bool(str(prom_cell.value or "").strip())
            if not prom_filled:
                prom_cell.fill = RED_FILL
                red_count += 1
            else:
                prom_cell.fill = NO_FILL  # знімаємо червоний якщо вже заповнили

    wb.save(XLSX_PATH)
    print(f"\nDone. Written: {total_written} | No match / skipped: {total_skipped} | Red: {red_count}")
    print(f"Saved → {XLSX_PATH}")


if __name__ == "__main__":
    main()
