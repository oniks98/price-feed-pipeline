# -*- coding: utf-8 -*-
"""Експорт товарів із Prom.ua до шаблону товарного фіда Вчасно.

Запуск:

    python scripts/products_export_vchasno.py

Книга призначення є шаблоном. Рядок заголовків зберігається, а всі наявні
рядки товарів замінюються, тому повторні запуски ідемпотентні.
"""

from __future__ import annotations

import logging
import tempfile
import time
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Налаштування
# ---------------------------------------------------------------------------

SOURCE_XLSX = Path(r"C:\FullStack\PriceFeedPipeline\data\markets\products_prom.xlsx")
SOURCE_SHEET_NAME = "Export Products Sheet"

DEST_XLSX = Path(r"C:\FullStack\PriceFeedPipeline\data\markets\products_vchasno.xlsx")
DEST_SHEET_NAME = "Всі товари"

# Колонка призначення -> колонка джерела. Ціна є фіксованою константою.
FIELD_MAP: dict[str, str] = {
    "Назва товару": "Назва_позиції_укр",
    "Артикул товару": "Код_товару",
    "Категорія товарів": "Назва_групи",
}
PRICE_COLUMN = "Ціна"
DEFAULT_PRICE = 1
TAX_GROUP_COLUMN = "Податкова група"
TAX_GROUP_VALUE = "Без ПДВ"
MAX_SKIPPED_ROW_LOGS = 5

LOG_PATH = Path(r"C:\FullStack\PriceFeedPipeline\logs\products_export_vchasno.log")

REQUIRED_SOURCE_COLUMNS = frozenset(FIELD_MAP.values())
REQUIRED_DESTINATION_COLUMNS = frozenset((*FIELD_MAP, PRICE_COLUMN, TAX_GROUP_COLUMN))

LOGGER = logging.getLogger(__name__)


@dataclass(slots=True)
class ExportStats:
    """Лічильники завершеного експорту."""

    written: int = 0
    skipped: int = 0


class DuplicateProductCodeError(ValueError):
    """Помилка, коли у джерелі повторюється Код_товару."""


def _configure_logging() -> None:
    """Налаштувати компактне логування в консоль і файл для прямого запуску."""
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    formatter = logging.Formatter(
        fmt="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    file_handler = logging.FileHandler(LOG_PATH, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logging.basicConfig(level=logging.INFO, handlers=[console_handler, file_handler], force=True)


def _text(value: Any) -> str:
    """Повернути текст без пробілів; порожні комірки перетворити на порожній рядок."""
    return str(value).strip() if value is not None else ""


def _header_indexes(header: Sequence[Any], required_columns: frozenset[str]) -> dict[str, int]:
    """Повернути індекси обов'язкових заголовків, використовуючи їх першу появу."""
    positions: dict[str, int] = {}
    for index, value in enumerate(header):
        name = _text(value)
        if name:
            positions.setdefault(name, index)

    missing = sorted(required_columns - positions.keys())
    if missing:
        raise ValueError(f"Відсутні обов'язкові колонки: {', '.join(missing)}")
    return {name: positions[name] for name in required_columns}


def _source_sheet(workbook: Workbook) -> Worksheet:
    """Повернути налаштований аркуш Prom або повідомити про доступні аркуші."""
    if SOURCE_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Аркуш '{SOURCE_SHEET_NAME}' відсутній у {SOURCE_XLSX.name}. "
            f"Наявні аркуші: {', '.join(workbook.sheetnames)}"
        )
    return workbook[SOURCE_SHEET_NAME]


def _destination_sheet(workbook: Workbook) -> Worksheet:
    """Повернути аркуш Вчасно або використати активний аркуш шаблону."""
    if DEST_SHEET_NAME in workbook.sheetnames:
        return workbook[DEST_SHEET_NAME]

    LOGGER.warning(
        "Аркуш '%s' не знайдено в шаблоні; використовую активний аркуш '%s'.",
        DEST_SHEET_NAME,
        workbook.active.title,
    )
    return workbook.active


def _clear_data_rows(worksheet: Worksheet) -> None:
    """Видалити лише рядки товарів, зберігши заголовки й налаштування шаблону."""
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)


def _write_products(
    source_sheet: Worksheet,
    source_columns: dict[str, int],
    destination_sheet: Worksheet,
    destination_columns: dict[str, int],
) -> ExportStats:
    """Потоково записати валідні рядки джерела безпосередньо до аркуша Вчасно."""
    source_name_column = source_columns[FIELD_MAP["Назва товару"]]
    source_code_column = source_columns[FIELD_MAP["Артикул товару"]]
    source_category_column = source_columns[FIELD_MAP["Категорія товарів"]]
    source_last_column = max(source_columns.values()) + 1

    destination_name_column = destination_columns["Назва товару"] + 1
    destination_price_column = destination_columns[PRICE_COLUMN] + 1
    destination_tax_group_column = destination_columns[TAX_GROUP_COLUMN] + 1
    destination_code_column = destination_columns["Артикул товару"] + 1
    destination_category_column = destination_columns["Категорія товарів"] + 1

    stats = ExportStats()
    logged_skips = 0
    first_row_by_code: dict[str, int] = {}
    duplicate_rows_by_code: dict[str, list[int]] = {}
    _clear_data_rows(destination_sheet)

    for row_number, row in enumerate(
        source_sheet.iter_rows(min_row=2, max_col=source_last_column, values_only=True),
        start=2,
    ):
        name = _text(row[source_name_column])
        code = _text(row[source_code_column])
        if not name or not code:
            stats.skipped += 1
            if logged_skips < MAX_SKIPPED_ROW_LOGS:
                LOGGER.warning("Рядок %d пропущено: назва=%r, код=%r", row_number, name, code)
                logged_skips += 1
            continue

        first_row = first_row_by_code.get(code)
        if first_row is None:
            first_row_by_code[code] = row_number
        else:
            duplicate_rows_by_code.setdefault(code, [first_row]).append(row_number)

        destination_sheet.append(
            {
                destination_name_column: name,
                destination_price_column: DEFAULT_PRICE,
                destination_tax_group_column: TAX_GROUP_VALUE,
                destination_code_column: code,
                destination_category_column: _text(row[source_category_column]),
            }
        )
        stats.written += 1

    if stats.skipped > logged_skips:
        LOGGER.warning(
            "Ще %d рядків пропущено; у журналі показано лише перші %d.",
            stats.skipped - logged_skips,
            MAX_SKIPPED_ROW_LOGS,
        )

    if duplicate_rows_by_code:
        for code, row_numbers in sorted(duplicate_rows_by_code.items()):
            LOGGER.error(
                "Код_товару '%s' повторюється у рядках Prom: %s",
                code,
                ", ".join(map(str, row_numbers)),
            )
        raise DuplicateProductCodeError(
            f"Виявлено повтори Код_товару: {len(duplicate_rows_by_code)}. "
            "Файл Вчасно не перезаписано."
        )
    return stats


def _save_atomically(workbook: Workbook, destination_path: Path) -> None:
    """Зберегти повну книгу до заміни попереднього файлу призначення."""
    with tempfile.NamedTemporaryFile(
        mode="wb",
        suffix=destination_path.suffix,
        prefix=f".{destination_path.stem}.",
        dir=destination_path.parent,
        delete=False,
    ) as temporary_file:
        temporary_path = Path(temporary_file.name)

    try:
        workbook.save(temporary_path)
        temporary_path.replace(destination_path)
    finally:
        temporary_path.unlink(missing_ok=True)


def export_products(source_path: Path = SOURCE_XLSX, destination_path: Path = DEST_XLSX) -> ExportStats:
    """Перевірити, зіставити й експортувати товари Prom до шаблону Вчасно."""
    if not source_path.is_file():
        raise FileNotFoundError(f"Файл джерела не знайдено: {source_path}")
    if not destination_path.is_file():
        raise FileNotFoundError(f"Шаблон Вчасно не знайдено: {destination_path}")

    source_workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    destination_workbook = openpyxl.load_workbook(destination_path)
    try:
        source_sheet = _source_sheet(source_workbook)
        destination_sheet = _destination_sheet(destination_workbook)

        source_header = next(source_sheet.iter_rows(max_row=1, values_only=True), ())
        destination_header = next(destination_sheet.iter_rows(max_row=1, values_only=True), ())
        source_columns = _header_indexes(source_header, REQUIRED_SOURCE_COLUMNS)
        destination_columns = _header_indexes(destination_header, REQUIRED_DESTINATION_COLUMNS)

        LOGGER.info("Джерело: %s / %s", source_path, source_sheet.title)
        LOGGER.info("Призначення: %s / %s", destination_path, destination_sheet.title)
        stats = _write_products(source_sheet, source_columns, destination_sheet, destination_columns)
        if not stats.written:
            raise ValueError("Не знайдено жодного товару з назвою та кодом; шаблон не змінено.")

        _save_atomically(destination_workbook, destination_path)
        return stats
    finally:
        source_workbook.close()
        destination_workbook.close()


def main() -> int:
    _configure_logging()
    started_at = time.monotonic()
    try:
        stats = export_products()
    except (FileNotFoundError, PermissionError, ValueError, OSError) as error:
        LOGGER.error("Експорт не виконано: %s", error)
        return 1

    elapsed = time.monotonic() - started_at
    LOGGER.info(
        "Готово: записано %d товарів, пропущено %d; час %.2f с.",
        stats.written,
        stats.skipped,
        elapsed,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
