"""
epicenter_export_coef.py
────────────────────────
Генерує рядки epicenter_coefficients.csv з листа «Маппінг»
та заповнює threshold на основі epicenter_royalty.xlsx.

Алгоритм:
  1. З листа «Маппінг» читаємо prom_category_id, prom_category_name,
     epicenter_category_id — вони стають рядками CSV
  2. В epicenter_royalty.xlsx шукаємо збіг по стовпцю «ID категорії»
     → беремо Відсоток роялті = X
     Якщо категорію не знайдено → беремо coef_uncategorized зі рядка дефолтів
  3. threshold = calc_coef(X)  # формула в services/market_formula_coef.py
  4. Записуємо threshold у потрібний рядок CSV
     (рядок дефолтів з порожнім prom_category_id не чіпається)
  5. Ручні coef_viatec / coef_secur / coef_lp переносяться зі старого файлу
     без змін (services/coef_export_service.py) — скрипт перебудовує рядки
     з мапінгу при кожному запуску, тому без цього кроку вручну введені
     коефіцієнти постачальників губилися б на кожному наступному запуску.

Запуск:
    python scripts/epicenter_export_coef.py
"""

from __future__ import annotations

import csv
import io
import logging
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

import openpyxl

from services.coef_export_service import SUPPLIER_COEF_FIELDS, read_manual_overrides
from services.market_formula_coef import calc_coef

# ─────────────────────────────── config ───────────────────────────────────────

BASE_DIR = Path(r"C:\FullStack\PriceFeedPipeline\data\markets")

MAPPINGS_PATH  = BASE_DIR / "epicenter_mappings.xlsx"
ROYALTY_PATH   = BASE_DIR / "epicenter_royalty.xlsx"
CSV_PATH       = BASE_DIR / "epicenter_coefficients.csv"

MAPPINGS_SHEET = "Маппінг"

# Заголовки стовпців (шукаємо позицію динамічно — стійко до зсувів колонок)
MAPPINGS_COL_PROM_ID      = "prom_category_id"
MAPPINGS_COL_PROM_NAME    = "Категорія Прому"
MAPPINGS_COL_EPICENTER_ID = "epicenter_category_id"

ROYALTY_COL_CATEGORY_ID = "ID категорії"
ROYALTY_COL_PERCENT     = "Відсоток роялті"

CSV_COL_CAT_ID             = "prom_category_id"
CSV_COL_CAT_NAME           = "prom_category_name"
CSV_COL_THRESHOLD          = "threshold"           # авторахований коефіцієнт (було "coef" — колонку перейменували)
CSV_COL_COEF_UNCATEGORIZED = "coef_uncategorized"  # джерело fallback-значення

CSV_DELIMITER = ";"
CSV_ENCODING  = "utf-8-sig"    # обробляє BOM автоматично

# ─────────────────────────────── logging ──────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ─────────────────────────────── helpers ──────────────────────────────────────

def _find_col_index(header_row: tuple, col_name: str, source: str) -> int:
    """
    Повертає 0-based індекс стовпця за іменем заголовка.
    Пошук case-insensitive, з видаленням пробілів.
    Кидає ValueError якщо стовпець не знайдено.
    """
    normalized = col_name.strip().lower()
    for idx, cell in enumerate(header_row):
        if cell is not None and str(cell).strip().lower() == normalized:
            return idx
    raise ValueError(
        f"[{source}] Стовпець '{col_name}' не знайдено. "
        f"Доступні: {[c for c in header_row if c is not None]}"
    )


def _to_int(value: object, label: str) -> Optional[int]:
    """Безпечне приведення до int. Повертає None при помилці."""
    try:
        return int(value)  # type: ignore[arg-type]
    except (ValueError, TypeError):
        log.warning("Не вдалося перетворити '%s' на int (%s)", value, label)
        return None


def _to_decimal(value: object, label: str) -> Optional[Decimal]:
    """Безпечне приведення до Decimal. Повертає None при помилці."""
    try:
        text = str(value).replace(",", ".").strip()
        return Decimal(text)
    except InvalidOperation:
        log.warning("Не вдалося перетворити '%s' на Decimal (%s)", value, label)
        return None


def read_fallback_coef(csv_path: Path) -> str | None:
    """
    Читає значення coef_uncategorized зі рядка дефолтів (порожній prom_category_id).
    Повертає рядок (наприклад '1.45') або None якщо не знайдено / порожньо.
    """
    raw    = csv_path.read_text(encoding=CSV_ENCODING)
    reader = csv.DictReader(io.StringIO(raw), delimiter=CSV_DELIMITER)

    if reader.fieldnames is None or CSV_COL_COEF_UNCATEGORIZED not in reader.fieldnames:
        return None

    for row in reader:
        if not row.get(CSV_COL_CAT_ID, "").strip():  # рядок дефолтів
            coef = row.get(CSV_COL_COEF_UNCATEGORIZED, "").strip()
            return coef if coef else None

    return None


# ─────────────────────────────── loaders ─────────────────────────────────────

def load_mappings(path: Path, sheet: str) -> dict[int, tuple[str, int]]:
    """
    Читає лист «Маппінг» і повертає
    {prom_category_id: (prom_category_name, epicenter_category_id)}.
    Позиції стовпців визначаються за заголовками — стійко до додавання колонок.
    Пропускає рядки з відсутніми або невалідними ID.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    if sheet not in wb.sheetnames:
        raise ValueError(
            f"Лист '{sheet}' не знайдено в {path}. "
            f"Доступні листи: {wb.sheetnames}"
        )

    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)

    header = next(rows, None)
    if header is None:
        raise RuntimeError(f"Лист '{sheet}' в {path} порожній")

    prom_col      = _find_col_index(header, MAPPINGS_COL_PROM_ID,      "epicenter_mappings")
    prom_name_col = _find_col_index(header, MAPPINGS_COL_PROM_NAME,    "epicenter_mappings")
    epicenter_col = _find_col_index(header, MAPPINGS_COL_EPICENTER_ID, "epicenter_mappings")

    result: dict[int, tuple[str, int]] = {}

    for row_idx, row in enumerate(rows, start=2):  # start=2 — реальний номер рядка у файлі
        prom_id      = _to_int(row[prom_col],      f"prom_category_id row={row_idx}")
        epicenter_id = _to_int(row[epicenter_col], f"epicenter_category_id row={row_idx}")

        if prom_id is None or epicenter_id is None:
            continue

        prom_name = str(row[prom_name_col] or "").strip()
        result[prom_id] = (prom_name, epicenter_id)

    wb.close()
    log.info("mappings: завантажено %d записів (prom_id -> name, epicenter_id)", len(result))
    return result


def load_royalty(path: Path) -> dict[int, Decimal]:
    """
    Читає epicenter_royalty.xlsx і повертає {epicenter_category_id: royalty_percent}.
    Лист визначається автоматично (активний).
    При дублюючих ID беремо максимальний відсоток.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)

    header = next(rows, None)
    if header is None:
        raise RuntimeError(f"Файл {path} порожній")

    cat_col     = _find_col_index(header, ROYALTY_COL_CATEGORY_ID, "royalty_epicenter")
    percent_col = _find_col_index(header, ROYALTY_COL_PERCENT,     "royalty_epicenter")

    royalty_map: dict[int, list[Decimal]] = {}

    for row_idx, row in enumerate(rows, start=2):
        cat_id  = _to_int(row[cat_col],         f"ID категорії row={row_idx}")
        percent = _to_decimal(row[percent_col],  f"Відсоток роялті row={row_idx}")

        if cat_id is None or percent is None:
            continue

        royalty_map.setdefault(cat_id, []).append(percent)

    wb.close()

    result = {cat_id: max(vals) for cat_id, vals in royalty_map.items()}
    log.info("royalty: завантажено %d унікальних категорій Epicenter", len(result))
    return result


# ─────────────────────────────── CSV processing ───────────────────────────────

def process_csv(
    csv_path: Path,
    mappings: dict[int, tuple[str, int]],
    royalty: dict[int, Decimal],
    fallback_coef: str | None,
) -> tuple[int, int]:
    """
    Генерує рядки CSV з маппінгу та заповнює threshold, перезаписує файл.

    Правила для threshold:
      - знайдено в royalty       → обчислюємо через calc_coef()
      - не знайдено в royalty    → fallback з coef_uncategorized рядка дефолтів

    coef_viatec / coef_secur / coef_lp — вручні, цим скриптом НЕ обчислюються —
    переносяться зі старого файлу ідемпотентно (read_manual_overrides), інакше
    вони б губилися на кожному запуску, бо рядки перебудовуються з маппінгу з нуля.

    Порядок запису у файлі:
      1. Рядок дефолтів (порожній prom_category_id) — без змін
      2. Дані з маппінгу, відсортовані за prom_category_id

    Повертає (updated, fallback_used).
    """
    raw = csv_path.read_text(encoding=CSV_ENCODING)
    reader = csv.DictReader(io.StringIO(raw), delimiter=CSV_DELIMITER)
    fieldnames = list(reader.fieldnames or [])

    if not fieldnames:
        raise RuntimeError(f"CSV {csv_path} порожній або не читається")
    for col in (CSV_COL_CAT_ID, CSV_COL_CAT_NAME, CSV_COL_THRESHOLD):
        if col not in fieldnames:
            raise RuntimeError(f"Стовпець '{col}' не знайдено в {csv_path}")

    # Вручні coef_viatec/coef_secur/coef_lp зі старого файлу — ключ один на категорію
    # (таблиця Epicenter плоска — одне правило на категорію).
    supplier_overrides = read_manual_overrides(
        csv_path,
        key_fields=(CSV_COL_CAT_ID,),
        preserve_fields=SUPPLIER_COEF_FIELDS,
    )

    # Зберігаємо рядок дефолтів (порожній prom_category_id)
    defaults_rows = [
        row for row in reader
        if not row.get(CSV_COL_CAT_ID, "").strip()
    ]

    updated       = 0
    fallback_used = 0
    data_rows: list[dict[str, str]] = []

    for prom_id, (prom_name, epicenter_id) in sorted(mappings.items()):
        royalty_percent = royalty.get(epicenter_id)

        if royalty_percent is not None:
            try:
                threshold_str = str(calc_coef(royalty_percent))
            except ValueError as exc:
                log.error(
                    "prom_id=%d epicenter_id=%d: %s — пропускаємо",
                    prom_id, epicenter_id, exc,
                )
                continue
            log.debug(
                "prom_id=%-6d  epicenter_id=%-6d  royalty=%s  threshold=%s",
                prom_id, epicenter_id, royalty_percent, threshold_str,
            )
        else:
            if fallback_coef is None:
                log.warning(
                    "prom_category_id=%d -> epicenter_category_id=%d: "
                    "не знайдено в royalty, fallback (coef_uncategorized) не задано — пропускаємо",
                    prom_id, epicenter_id,
                )
                continue
            threshold_str = fallback_coef
            log.warning(
                "prom_category_id=%d -> epicenter_category_id=%d: "
                "не знайдено в royalty → fallback coef_uncategorized=%s",
                prom_id, epicenter_id, threshold_str,
            )
            fallback_used += 1

        row: dict[str, str] = {fn: "" for fn in fieldnames}
        row[CSV_COL_CAT_ID]    = str(prom_id)
        row[CSV_COL_CAT_NAME]  = prom_name
        row[CSV_COL_THRESHOLD] = threshold_str
        row.update(supplier_overrides.get((str(prom_id),), {}))
        data_rows.append(row)
        updated += 1

    # --- перезаписуємо файл: дефолти → дані ---
    out = io.StringIO()
    writer = csv.DictWriter(
        out,
        fieldnames=fieldnames,
        delimiter=CSV_DELIMITER,
        lineterminator="\n",
        extrasaction="ignore",
    )
    writer.writeheader()
    writer.writerows(defaults_rows)
    writer.writerows(data_rows)

    csv_path.write_text(out.getvalue(), encoding=CSV_ENCODING)
    return updated, fallback_used


# ─────────────────────────────── main ─────────────────────────────────────────

def main() -> None:
    log.info("=== epicenter_export_coef старт ===")

    for path in (MAPPINGS_PATH, ROYALTY_PATH, CSV_PATH):
        if not path.exists():
            log.error("Файл не знайдено: %s", path)
            sys.exit(1)

    mappings = load_mappings(MAPPINGS_PATH, MAPPINGS_SHEET)
    royalty  = load_royalty(ROYALTY_PATH)

    fallback_coef = read_fallback_coef(CSV_PATH)
    if fallback_coef:
        log.info("Fallback зі рядка дефолтів: coef_uncategorized=%s", fallback_coef)
    else:
        log.warning(
            "Fallback coef_uncategorized не знайдено в рядку дефолтів CSV — "
            "категорії без роялті будуть пропущені"
        )

    updated, fallback_used = process_csv(CSV_PATH, mappings, royalty, fallback_coef)

    log.info(
        "=== Готово: записано=%d, fallback_використано=%d ===",
        updated, fallback_used,
    )


if __name__ == "__main__":
    main()
