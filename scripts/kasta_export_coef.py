"""
kasta_export_coef.py
────────────────────
Генерує data/markets/kasta_coefficients.csv з:
  - data/markets/mappings.xlsx, аркуш «Категорія+»
  - data/markets/kasta_royalty.xlsx, аркуш «Роялті»

Алгоритм:
  1. З маппінгу читаємо prom_category_id → ключі зіставлення (Приналежність, Група, Вид)
  2. З таблиці роялті збираємо правила цінових діапазонів [price_from, price_to) → royalty_percent
  3. Для кожного збігу обчислюємо threshold = calc_coef(royalty_percent)  # services/market_formula_coef.py
     (threshold — множник дилерської ціни; узгоджено з prom/epicenter/rozetka)
  4. Вивід містить правила діапазонів, а не один максимальний threshold на категорію
  5. Стовпці coef / coef_viatec / coef_secur / coef_lp — вручні, цим скриптом НЕ
     обчислюються — лише додаються до OUTPUT_FIELDS. Нові рядки отримують порожні
     значення (не "1" і не будь-який інший дефолт) — порожній coef/coef_{supplier}
     при заповненому threshold це свідома критична відсутність ручного
     коефіцієнта (missing_manual_coef в pricing_rules/kasta.py::apply_prices).
     Вже заповнені вручну значення (включно coef_viatec/secur/lp) зберігаються
     ідемпотентно через services/coef_export_service.py::read_manual_overrides за ключем
     (prom_category_id, price_from, price_to) — раніше coef_viatec/secur/lp взагалі
     втрачалися на кожному запуску, бо їх не було в OUTPUT_FIELDS — виправлено.

Зіставлення відбувається за спільними заголовками обох файлів: Приналежність, Група, Вид.

Запуск:
    python scripts/kasta_export_coef.py
"""

from __future__ import annotations

import csv
import io
import logging
import re
import sys
import warnings
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from services.coef_export_service import SUPPLIER_COEF_FIELDS, read_manual_overrides
from services.market_formula_coef import calc_coef

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style",
    category=UserWarning,
    module="openpyxl.styles.stylesheet",
)


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parents[1]
BASE_DIR = ROOT / "data" / "markets"

MAPPINGS_PATH = BASE_DIR / "mappings.xlsx"
ROYALTY_PATH = BASE_DIR / "kasta_royalty.xlsx"
OUTPUT_CSV_PATH = BASE_DIR / "kasta_coefficients.csv"

MAPPINGS_SHEET = "Категорія+"
ROYALTY_SHEET = "Роялті"

CSV_DELIMITER = ";"
CSV_ENCODING = "utf-8-sig"

OUTPUT_FIELDS = [
    "prom_category_id",
    "prom_category_name",
    "Приналежність*:6",
    "Група*:13",
    "Вид*:21",
    "royalty_percent",
    "price_from",
    "price_to",
    "threshold",            # множник дилерської ціни: calc_coef(royalty_percent)
    "coef_uncategorized",  # J — є оптова ціна, але немає правила для категорії
    "coef_no_base",         # K — немає оптової ціни → базою стає ціна з XML-фіду
    "coef_viatec",          # ручний коефіцієнт діапазону для viatec — зберігається ідемпотентно
    "coef_secur",           # ручний коефіцієнт діапазону для secur — зберігається ідемпотентно
    "coef_lp",              # ручний коефіцієнт діапазону для lp — зберігається ідемпотентно
    "coef",                 # множник роздрібної ціни — застарілий, цим скриптом НЕ відновлюється (буде порожнім у кожному новому рядку)
]

# Усі вручні стовпці постачальників, що повинні переживати повторні запуски скрипта (ідемпотентно).
# Єдине джерело правди — SUPPLIER_COEF_FIELDS (services/coef_export_service.py).
# "coef" (загальний множник роздрібної ціни) більше НЕ відновлюється цим скриптом —
# тепер єдина джерело правди для всіх маркетплейсів — coef_viatec/secur/lp.

OUTPUT_MAPPING_FIELDS = [
    "Приналежність*:6",
    "Група*:13",
    "Вид*:21",
]

DEFAULT_UNCATEGORIZED_COEF = Decimal("1.45")
DEFAULT_NO_BASE_COEF       = Decimal("1.2")

ROYALTY_MEASURE_HEADERS = {
    "відсоток роялті",
    "ціна від (включно)",
    "ціна до (не включає)",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class MappingEntry:
    category_id: str
    category_name: str
    match_values: dict[str, str]
    output_values: dict[str, str]


@dataclass(frozen=True)
class RoyaltyRule:
    royalty_percent: Decimal
    price_from: Decimal
    price_to: Decimal


# ---------------------------------------------------------------------------
# Text / number helpers
# ---------------------------------------------------------------------------


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\ufeff", "").replace("\xa0", " ").strip().lower()
    return re.sub(r"\s+", " ", text)


def normalize_header(value: object) -> str:
    text = normalize_text(value)
    return re.sub(r"\*?:\d+\s*$", "", text).strip()


def cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_decimal(value: object, default: Decimal | None = None) -> Decimal:
    if value is None or value == "":
        if default is not None:
            return default
        raise InvalidOperation("empty decimal")

    if isinstance(value, Decimal):
        return value

    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
    if not text:
        if default is not None:
            return default
        raise InvalidOperation("empty decimal")

    return Decimal(text)


def format_decimal(value: Decimal) -> str:
    if value.is_infinite():
        return ""
    if value == value.to_integral_value():
        return str(value.quantize(Decimal("1")))
    return format(value.normalize(), "f").rstrip("0").rstrip(".")



# ---------------------------------------------------------------------------
# XLSX / CSV loading
# ---------------------------------------------------------------------------


def load_sheet_rows(path: Path, sheet: str) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{sheet}' not found in {path}. Available: {workbook.sheetnames}"
            )
        worksheet = workbook[sheet]

        # Some exported XLSX files store incorrect dimensions (A1:A1).
        # reset_dimensions() forces openpyxl to stream the real used range.
        if hasattr(worksheet, "reset_dimensions"):
            worksheet.reset_dimensions()

        return list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def header_index(header_row: tuple) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(header_row):
        normalized = normalize_header(value)
        if normalized and normalized not in result:
            result[normalized] = index
    return result


def get_cell(row: tuple, index: int) -> object:
    return row[index] if index < len(row) else None


def read_existing_defaults(path: Path) -> tuple[Decimal, Decimal]:
    """Читає coef_uncategorized та coef_no_base зі рядка дефолтів (порожній prom_category_id)."""
    if not path.exists():
        return DEFAULT_UNCATEGORIZED_COEF, DEFAULT_NO_BASE_COEF

    with path.open(encoding=CSV_ENCODING, errors="replace", newline="") as file:
        reader = csv.DictReader(file, delimiter=CSV_DELIMITER)
        if not reader.fieldnames:
            return DEFAULT_UNCATEGORIZED_COEF, DEFAULT_NO_BASE_COEF

        coef_unc: Decimal | None = None
        coef_nb:  Decimal | None = None

        for row in reader:
            if coef_unc is None:
                raw = (row.get("coef_uncategorized") or "").strip()
                if raw:
                    try:
                        coef_unc = parse_decimal(raw)
                    except InvalidOperation:
                        log.warning("Invalid coef_uncategorized=%r in %s", raw, path)

            if coef_nb is None:
                raw = (row.get("coef_no_base") or "").strip()
                if raw:
                    try:
                        coef_nb = parse_decimal(raw)
                    except InvalidOperation:
                        log.warning("Invalid coef_no_base=%r in %s", raw, path)

            if coef_unc is not None and coef_nb is not None:
                break

    return (
        coef_unc if coef_unc is not None else DEFAULT_UNCATEGORIZED_COEF,
        coef_nb  if coef_nb  is not None else DEFAULT_NO_BASE_COEF,
    )



def load_mappings(
    rows: list[tuple],
    dimensions: list[str],
    mapping_headers: dict[str, int],
) -> list[MappingEntry]:
    id_col = mapping_headers.get("іd категорії фіду", 0)
    name_col = mapping_headers.get("категорії фіду", 1)
    output_cols = {
        field: mapping_headers.get(normalize_header(field))
        for field in OUTPUT_MAPPING_FIELDS
    }

    result: list[MappingEntry] = []

    for row_number, row in enumerate(rows[1:], start=2):
        category_id = cell_text(get_cell(row, id_col))
        if not category_id:
            continue

        match_values: dict[str, str] = {}
        missing_dimension = False

        for dimension in dimensions:
            col = mapping_headers[dimension]
            value = normalize_text(get_cell(row, col))
            if not value:
                missing_dimension = True
                break
            match_values[dimension] = value

        if missing_dimension:
            log.warning("mappings row %d: missing Kasta dimension, skipped", row_number)
            continue

        output_values = {
            field: cell_text(get_cell(row, col)) if col is not None else ""
            for field, col in output_cols.items()
        }

        result.append(
            MappingEntry(
                category_id=category_id,
                category_name=cell_text(get_cell(row, name_col)),
                match_values=match_values,
                output_values=output_values,
            )
        )

    log.info("mappings: loaded %d categories", len(result))
    return result


def load_royalty_index(
    rows: list[tuple],
    dimensions: list[str],
    royalty_headers: dict[str, int],
) -> dict[tuple[str, ...], list[RoyaltyRule]]:
    percent_col = royalty_headers["відсоток роялті"]
    price_from_col = royalty_headers["ціна від (включно)"]
    price_to_col = royalty_headers["ціна до (не включає)"]

    result: dict[tuple[str, ...], list[RoyaltyRule]] = {}
    skipped = 0

    for row_number, row in enumerate(rows[1:], start=2):
        key_parts = [normalize_text(get_cell(row, royalty_headers[d])) for d in dimensions]
        if not all(key_parts):
            skipped += 1
            continue

        try:
            rule = RoyaltyRule(
                royalty_percent=parse_decimal(get_cell(row, percent_col)),
                price_from=parse_decimal(get_cell(row, price_from_col), Decimal("0")),
                price_to=parse_decimal(get_cell(row, price_to_col), Decimal("Infinity")),
            )
        except InvalidOperation:
            log.warning("royalty row %d: invalid numeric values, skipped", row_number)
            skipped += 1
            continue

        result.setdefault(tuple(key_parts), []).append(rule)

    for rules in result.values():
        rules.sort(key=lambda item: (item.price_from, item.price_to, item.royalty_percent))

    log.info(
        "royalty: loaded %d unique Kasta keys, skipped rows=%d",
        len(result),
        skipped,
    )
    return result


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------


def build_output_rows(
    mappings: list[MappingEntry],
    royalty_index: dict[tuple[str, ...], list[RoyaltyRule]],
    dimensions: list[str],
    default_coef: Decimal,
    no_base_coef: Decimal,
    manual_overrides: dict[tuple[str, ...], dict[str, str]],
) -> tuple[list[dict[str, str]], int, int, int]:
    rows: list[dict[str, str]] = []
    rows.append({field: "" for field in OUTPUT_FIELDS})
    rows[0]["coef_uncategorized"] = format_decimal(default_coef)
    rows[0]["coef_no_base"]       = format_decimal(no_base_coef)

    matched_categories = 0
    unmatched_categories = 0
    generated_rules = 0

    for mapping in mappings:
        key = tuple(mapping.match_values[d] for d in dimensions)
        rules = royalty_index.get(key)
        if not rules:
            unmatched_categories += 1
            log.warning(
                "category_id=%s: no Kasta royalty match for key=%s",
                mapping.category_id,
                key,
            )
            continue

        matched_categories += 1
        for rule in rules:
            try:
                threshold = calc_coef(rule.royalty_percent)
            except ValueError as exc:
                log.warning("category_id=%s: %s, skipped", mapping.category_id, exc)
                continue

            price_from_str = format_decimal(rule.price_from)
            price_to_str   = format_decimal(rule.price_to)
            override_key   = (mapping.category_id, price_from_str, price_to_str)

            output_row = {field: "" for field in OUTPUT_FIELDS}
            output_row.update(
                {
                    "prom_category_id": mapping.category_id,
                    "prom_category_name": mapping.category_name,
                    "royalty_percent": format_decimal(rule.royalty_percent),
                    "price_from": price_from_str,
                    "price_to": price_to_str,
                    "threshold": format_decimal(threshold),
                }
            )
            for field in OUTPUT_MAPPING_FIELDS:
                output_row[field] = mapping.output_values.get(field, "")

            # Ні в якому випадку НЕ підставляється числовий дефолт: порожні coef/coef_{supplier}
            # при заповненому threshold — ознака для apply_prices, що коефіцієнт ще не перевірено вручну.
            output_row.update(manual_overrides.get(override_key, {}))

            rows.append(output_row)
            generated_rules += 1

    return rows, matched_categories, unmatched_categories, generated_rules


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=OUTPUT_FIELDS,
        delimiter=CSV_DELIMITER,
        lineterminator="\n",
    )
    writer.writeheader()
    writer.writerows(rows)
    content = output.getvalue()

    if path.exists():
        with path.open("r+", encoding=CSV_ENCODING, newline="") as file:
            file.seek(0)
            file.write(content)
            file.truncate()
    else:
        path.write_text(content, encoding=CSV_ENCODING)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    log.info("=== kasta coefficients export start ===")

    for path in (MAPPINGS_PATH, ROYALTY_PATH):
        if not path.exists():
            log.error("File not found: %s", path)
            sys.exit(1)

    mapping_rows = load_sheet_rows(MAPPINGS_PATH, MAPPINGS_SHEET)
    royalty_rows = load_sheet_rows(ROYALTY_PATH, ROYALTY_SHEET)

    if not mapping_rows or not royalty_rows:
        log.error("Input workbook has no rows")
        sys.exit(1)

    mapping_headers = header_index(mapping_rows[0])
    royalty_headers = header_index(royalty_rows[0])

    missing = ROYALTY_MEASURE_HEADERS - set(royalty_headers)
    if missing:
        log.error("Missing required royalty columns: %s", ", ".join(sorted(missing)))
        sys.exit(1)

    dimensions = [
        header
        for header in royalty_headers
        if header in mapping_headers and header not in ROYALTY_MEASURE_HEADERS
    ]
    if not dimensions:
        log.error("No common Kasta mapping dimensions between mappings and royalty files")
        sys.exit(1)

    log.info("matching dimensions: %s", ", ".join(dimensions))

    default_coef, no_base_coef = read_existing_defaults(OUTPUT_CSV_PATH)
    manual_overrides = read_manual_overrides(
        OUTPUT_CSV_PATH,
        key_fields=("prom_category_id", "price_from", "price_to"),
        preserve_fields=SUPPLIER_COEF_FIELDS,
    )
    mappings = load_mappings(mapping_rows, dimensions, mapping_headers)
    royalty_index = load_royalty_index(royalty_rows, dimensions, royalty_headers)

    rows, matched, unmatched, rules = build_output_rows(
        mappings,
        royalty_index,
        dimensions,
        default_coef,
        no_base_coef,
        manual_overrides,
    )
    write_csv(OUTPUT_CSV_PATH, rows)

    log.info(
        "=== done: matched_categories=%d, unmatched_categories=%d, rules=%d, "
        "coef_uncategorized=%s, coef_no_base=%s ===",
        matched,
        unmatched,
        rules,
        format_decimal(default_coef),
        format_decimal(no_base_coef),
    )
    log.info("saved: %s", OUTPUT_CSV_PATH)


if __name__ == "__main__":
    main()
