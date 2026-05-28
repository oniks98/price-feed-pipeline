"""
rozetka_export_coef.py
──────────────────────────
Генерує data/markets/rozetka_coefficients.csv з:
  - data/markets/rozetka_mappings.xlsx, аркуші «Маппінг» і «Категорії Розетки»
  - data/markets/rozetka_royalty.xlsx, аркуш «Тариф»

Алгоритм матчингу (ієрархічний bubble-up):
  1. Для кожного запису з «Маппінг» беремо rozetka_category_id.
  2. Шукаємо rozetka_category_id напряму в індексі роялті → є → стоп, фіксуємо.
  3. Немає → переходимо до «Категорії Розетки», беремо parentCode цієї категорії.
  4. Перевіряємо parentCode в роялті → є → стоп, фіксуємо.
  5. Немає → піднімаємось ще на рівень вгору, і так далі.
  6. Якщо жоден рівень ієрархії не знайдено → warning, запис пропускається.

Заповнення першого діапазону (Rozetka rule):
  Якщо знайдені правила НЕ мають рядка «-» (ставка для всіх цін) і мінімальна
  ціна правил > 0 → для діапазону [0, min_price_from-1] застосовується ставка
  батьківської категорії (піднімаємося вгору по ієрархії до першого предка,
  що має правило для ціни 0; 0-категорія гарантовано має таке правило).

Особливості формату роялті:
  - Порожній «ID категорії» → forward-fill з попереднього рядка
  - Бренд «-»       → "" (правило для всіх брендів)
  - Діапазон «-»    → price_from="", price_to="" (ставка для всіх цін)
  - Діапазон «FROM-TO» → розбивається на дві окремі цифри
  - Відсоток «7,5»  → кома замінюється на крапку

Запуск:
    python scripts/rozetka_export_coef.py
"""

from __future__ import annotations

import csv
import io
import logging
import re
import sys
import warnings
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from services.market_formula_coef import calc_coef

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style",
    category=UserWarning,
    module="openpyxl.styles.stylesheet",
)


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parents[1]
BASE_DIR = ROOT / "data" / "markets"

MAPPINGS_PATH   = BASE_DIR / "rozetka_mappings.xlsx"
ROYALTY_PATH    = BASE_DIR / "rozetka_royalty.xlsx"
OUTPUT_CSV_PATH = BASE_DIR / "rozetka_coefficients.csv"

MAPPINGS_SHEET   = "Маппінг"
CATEGORIES_SHEET = "Категорії Розетки"
ROYALTY_SHEET    = "Тариф"

CSV_DELIMITER = ";"
CSV_ENCODING  = "utf-8-sig"

OUTPUT_FIELDS = [
    "prom_category_id",
    "prom_category_name",
    "rozetka_category_id",
    "rozetka_category_name",
    "matched_royalty_id",      # ID категорії в роялті (може бути предком)
    "matched_royalty_name",    # Назва тієї категорії (для діагностики)
    "match_level",             # "0"=прямий, "1+"=рівень предка, "gap:ID"=синтетичний
    "brand",
    "royalty_percent",
    "price_from",
    "price_to",
    "coef",
    "coef_uncategorized",
    "coef_no_base",
]

# Назви стовпців у «Маппінг» (після normalize_header)
_COL_PROM_ID   = "prom_category_id"
_COL_PROM_NAME = "категорія прому"
_COL_ROZ_ID    = "rozetka_category_id"
_COL_ROZ_NAME  = "назва категорії розетки"

# Назви стовпців у «Категорії Розетки» (після normalize_header)
_COL_CAT_ID        = "rozetka_category_id"
_COL_CAT_NAME      = "назва категорії розетки"
_COL_CAT_PARENT    = "parentcode"
_COL_CAT_LEVEL     = "level"
_COL_CAT_FULL_PATH = "повний шлях категорії"

# Назви стовпців у роялті (після normalize_header)
_COL_ROYALTY_CAT_ID  = "id категорії"
_COL_ROYALTY_BRAND   = "бренд"
_COL_ROYALTY_RANGE   = "діапазон цін"
_COL_ROYALTY_PERCENT = "відсоток комісії"

REQUIRED_MAPPING_COLS: frozenset[str] = frozenset({_COL_PROM_ID, _COL_ROZ_ID})
REQUIRED_CATEGORY_COLS: frozenset[str] = frozenset({_COL_CAT_ID, _COL_CAT_PARENT})
REQUIRED_ROYALTY_COLS: frozenset[str] = frozenset({
    _COL_ROYALTY_CAT_ID,
    _COL_ROYALTY_BRAND,
    _COL_ROYALTY_RANGE,
    _COL_ROYALTY_PERCENT,
})

DEFAULT_UNCATEGORIZED_COEF = Decimal("1.45")
DEFAULT_NO_BASE_COEF       = Decimal("1.2")

# Обмеження глибини підйому по ієрархії (захист від циклів / нескінченних дерев)
MAX_HIERARCHY_DEPTH = 10

_BRAND_ANY  = "-"
_RANGE_NONE = "-"

# Ціна для пошуку «ставки першого діапазону» в батьківській категорії
_PRICE_ZERO = Decimal("0")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class MappingEntry:
    category_id:   str  # prom_category_id
    category_name: str
    rozetka_id:    str
    rozetka_name:  str


@dataclass(frozen=True)
class RozetkaCategory:
    category_id: str
    name:        str
    parent_id:   str   # "" якщо кореневий
    level:       int
    full_path:   str


@dataclass(frozen=True)
class RoyaltyRule:
    brand:           str
    price_from:      str     # "" означає «з нуля» (рядок «-» у файлі)
    price_to:        str     # "" означає «до нескінченності»
    price_from_dec:  Decimal
    price_to_dec:    Decimal
    royalty_percent: Decimal


@dataclass(frozen=True)
class MatchResult:
    rules:             list[RoyaltyRule]
    matched_id:        str   # rozetka_category_id у роялті
    matched_name:      str   # назва тієї категорії
    match_level:       int   # 0 = прямий; 1+ = рівень предка


@dataclass(frozen=True)
class ResolvedRule:
    """
    Правило роялті зі збагаченою інформацією про джерело.

    Для звичайних правил matched_id / matched_name = match.matched_id / match.matched_name.
    Для синтетичного правила (gap-fill) вони вказують на БАТЬКІВСЬКУ категорію,
    з якої взята ставка, а match_level_str = "gap:<match.matched_id>".
    """
    rule:            RoyaltyRule
    matched_id:      str
    matched_name:    str
    match_level_str: str   # "0", "1", "gap:12345" тощо


# ---------------------------------------------------------------------------
# Text / number helpers
# ---------------------------------------------------------------------------


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\ufeff", "").replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", text)


def normalize_header(value: object) -> str:
    text = normalize_text(value).lower()
    return re.sub(r"\*?:\d+\s*$", "", text).strip()


def cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_id(raw: object) -> str:
    """«131143.0» → «131143»; рядок без змін; None → ""."""
    value = cell_text(raw)
    if not value:
        return ""
    try:
        return str(int(float(value)))
    except (ValueError, OverflowError):
        return value


def parse_decimal(raw: object, default: Decimal | None = None) -> Decimal:
    if raw is None or raw == "":
        if default is not None:
            return default
        raise InvalidOperation("empty decimal")
    text = (
        str(raw)
        .replace("\xa0", "")
        .replace(" ", "")
        .replace(",", ".")
        .strip()
    )
    if not text:
        if default is not None:
            return default
        raise InvalidOperation("empty decimal after clean")
    return Decimal(text)


def format_decimal(value: Decimal) -> str:
    if not value.is_finite():
        return ""
    if value == value.to_integral_value():
        return str(value.quantize(Decimal("1")))
    return format(value.normalize(), "f").rstrip("0").rstrip(".")


# ---------------------------------------------------------------------------
# XLSX helpers
# ---------------------------------------------------------------------------


def _open_sheet(path: Path, preferred_sheet: str) -> tuple:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if preferred_sheet in wb.sheetnames:
        ws = wb[preferred_sheet]
    else:
        log.warning(
            "Sheet '%s' not found in %s, using first sheet '%s'",
            preferred_sheet, path.name, wb.sheetnames[0],
        )
        ws = wb[wb.sheetnames[0]]
    if hasattr(ws, "reset_dimensions"):
        ws.reset_dimensions()
    return wb, ws


def _header_index(header_row: tuple) -> dict[str, int]:
    result: dict[str, int] = {}
    for idx, val in enumerate(header_row):
        key = normalize_header(val)
        if key and key not in result:
            result[key] = idx
    return result


def _check_required(
    headers: dict[str, int],
    required: frozenset[str],
    source: str,
) -> None:
    missing = required - set(headers)
    if missing:
        log.error("Missing required columns in %s: %s", source, sorted(missing))
        sys.exit(1)


# ---------------------------------------------------------------------------
# Loading: mappings  (аркуш «Маппінг»)
# ---------------------------------------------------------------------------


def load_mappings(path: Path, sheet: str) -> list[MappingEntry]:
    wb, ws = _open_sheet(path, sheet)
    try:
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        log.error("Mappings sheet is empty: %s / %s", path.name, sheet)
        sys.exit(1)

    hdrs = _header_index(rows[0])
    _check_required(hdrs, REQUIRED_MAPPING_COLS, f"{path.name}/{sheet}")

    id_col       = hdrs[_COL_PROM_ID]
    name_col     = hdrs.get(_COL_PROM_NAME, 1)
    roz_id_col   = hdrs[_COL_ROZ_ID]
    roz_name_col = hdrs.get(_COL_ROZ_NAME, 3)

    entries: list[MappingEntry] = []
    skipped = 0

    for row_num, row in enumerate(rows[1:], start=2):
        def _get(col: int) -> str:
            return cell_text(row[col] if col < len(row) else None)

        prom_id    = normalize_id(row[id_col] if id_col < len(row) else None)
        rozetka_id = normalize_id(row[roz_id_col] if roz_id_col < len(row) else None)

        if not prom_id:
            continue
        if not rozetka_id:
            log.warning(
                "mappings row %d: empty rozetka_category_id, skipped (prom_id=%s)",
                row_num, prom_id,
            )
            skipped += 1
            continue

        entries.append(MappingEntry(
            category_id   = prom_id,
            category_name = _get(name_col),
            rozetka_id    = rozetka_id,
            rozetka_name  = _get(roz_name_col),
        ))

    log.info("mappings: loaded %d entries, skipped %d", len(entries), skipped)
    return entries


# ---------------------------------------------------------------------------
# Loading: category tree  (аркуш «Категорії Розетки»)
# ---------------------------------------------------------------------------


def load_category_tree(path: Path, sheet: str) -> dict[str, RozetkaCategory]:
    """
    Повертає: rozetka_category_id → RozetkaCategory.
    Використовується для bubble-up по ієрархії та для gap-fill.
    """
    wb, ws = _open_sheet(path, sheet)
    try:
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        log.error("Categories sheet is empty: %s / %s", path.name, sheet)
        sys.exit(1)

    hdrs = _header_index(rows[0])
    _check_required(hdrs, REQUIRED_CATEGORY_COLS, f"{path.name}/{sheet}")

    id_col        = hdrs[_COL_CAT_ID]
    name_col      = hdrs.get(_COL_CAT_NAME, 1)
    parent_col    = hdrs[_COL_CAT_PARENT]
    level_col     = hdrs.get(_COL_CAT_LEVEL)
    full_path_col = hdrs.get(_COL_CAT_FULL_PATH)

    tree: dict[str, RozetkaCategory] = {}
    skipped = 0

    for row_num, row in enumerate(rows[1:], start=2):
        cat_id    = normalize_id(row[id_col] if id_col < len(row) else None)
        parent_id = normalize_id(row[parent_col] if parent_col < len(row) else None)

        if not cat_id:
            skipped += 1
            continue

        name = normalize_text(row[name_col] if name_col < len(row) else None)

        level: int = 0
        if level_col is not None and level_col < len(row):
            try:
                level = int(row[level_col] or 0)
            except (ValueError, TypeError):
                level = 0

        full_path = ""
        if full_path_col is not None and full_path_col < len(row):
            full_path = normalize_text(row[full_path_col])

        tree[cat_id] = RozetkaCategory(
            category_id = cat_id,
            name        = name,
            parent_id   = parent_id,
            level       = level,
            full_path   = full_path,
        )

    log.info("category tree: loaded %d categories, skipped %d", len(tree), skipped)
    return tree


# ---------------------------------------------------------------------------
# Loading: royalty  (аркуш «Тариф»)
# ---------------------------------------------------------------------------


def _parse_range(raw: str) -> tuple[str, str, Decimal, Decimal]:
    """
    Розбирає рядок діапазону цін.

    «-»         → ("", "", 0, Infinity)  — ставка для всіх цін (fallback рядка)
    «2000-6999» → ("2000", "6999", 2000, 6999)
    """
    value = normalize_text(raw)
    if not value or value == _RANGE_NONE:
        return "", "", Decimal("0"), Decimal("Infinity")
    match = re.match(r"^(\d+)-(\d+)$", value)
    if not match:
        raise ValueError(f"Unexpected range format: {value!r}")
    from_str, to_str = match.group(1), match.group(2)
    return from_str, to_str, Decimal(from_str), Decimal(to_str)


def load_royalty_index(path: Path, sheet: str) -> dict[str, list[RoyaltyRule]]:
    """
    Повертає: rozetka_category_id → відсортований список RoyaltyRule.
    Порожній «ID категорії» → forward-fill з попереднього рядка.
    """
    wb, ws = _open_sheet(path, sheet)
    try:
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        log.error("Royalty sheet is empty: %s", path.name)
        sys.exit(1)

    hdrs = _header_index(rows[0])
    _check_required(hdrs, REQUIRED_ROYALTY_COLS, path.name)

    cat_id_col  = hdrs[_COL_ROYALTY_CAT_ID]
    brand_col   = hdrs[_COL_ROYALTY_BRAND]
    range_col   = hdrs[_COL_ROYALTY_RANGE]
    percent_col = hdrs[_COL_ROYALTY_PERCENT]

    index: dict[str, list[RoyaltyRule]] = {}
    last_cat_id = ""
    skipped = 0

    for row_num, row in enumerate(rows[1:], start=2):
        def _get(col: int) -> str:
            return normalize_text(row[col] if col < len(row) else None)

        raw_cat_id = normalize_id(row[cat_id_col] if cat_id_col < len(row) else None)
        if raw_cat_id:
            last_cat_id = raw_cat_id
        cat_id = last_cat_id

        if not cat_id:
            skipped += 1
            continue

        raw_brand   = _get(brand_col)
        raw_range   = _get(range_col)
        raw_percent = _get(percent_col)

        brand = "" if (not raw_brand or raw_brand == _BRAND_ANY) else raw_brand

        try:
            pf_str, pt_str, pf_dec, pt_dec = _parse_range(raw_range)
        except ValueError as exc:
            log.warning("royalty row %d: %s, skipped", row_num, exc)
            skipped += 1
            continue

        try:
            percent = parse_decimal(raw_percent)
        except InvalidOperation:
            log.warning("royalty row %d: invalid percent %r, skipped", row_num, raw_percent)
            skipped += 1
            continue

        index.setdefault(cat_id, []).append(RoyaltyRule(
            brand           = brand,
            price_from      = pf_str,
            price_to        = pt_str,
            price_from_dec  = pf_dec,
            price_to_dec    = pt_dec,
            royalty_percent = percent,
        ))

    for rules in index.values():
        rules.sort(key=lambda r: (0 if r.brand == "" else 1, r.price_from_dec, r.price_to_dec))

    log.info(
        "royalty: loaded %d unique category keys, skipped rows=%d",
        len(index), skipped,
    )
    return index


# ---------------------------------------------------------------------------
# Hierarchy matching (bubble-up)
# ---------------------------------------------------------------------------


def find_rules_by_hierarchy(
    rozetka_id:    str,
    category_tree: dict[str, RozetkaCategory],
    royalty_index: dict[str, list[RoyaltyRule]],
) -> MatchResult | None:
    """
    Шукає правила роялті для rozetka_id.

    Алгоритм:
      1. Перевіряємо rozetka_id напряму в royalty_index.
      2. Якщо немає — беремо parentCode з category_tree і перевіряємо.
      3. Повторюємо вгору до MAX_HIERARCHY_DEPTH.
      4. Якщо нічого не знайдено → повертаємо None.

    Повертає MatchResult із першим (найближчим) знайденим набором правил.
    """
    visited:     set[str] = set()
    current_id = rozetka_id
    level      = 0

    while current_id and level < MAX_HIERARCHY_DEPTH:
        if current_id in visited:
            log.warning("category_tree: cycle detected at id=%s", current_id)
            break
        visited.add(current_id)

        rules = royalty_index.get(current_id)
        if rules:
            cat      = category_tree.get(current_id)
            cat_name = cat.name if cat else current_id
            return MatchResult(
                rules        = rules,
                matched_id   = current_id,
                matched_name = cat_name,
                match_level  = level,
            )

        cat = category_tree.get(current_id)
        if cat is None or not cat.parent_id:
            break
        current_id = cat.parent_id
        level += 1

    return None


# ---------------------------------------------------------------------------
# Gap-fill: перший діапазон від батьківської категорії
# ---------------------------------------------------------------------------


def find_gap_fill_rate(
    source_cat_id: str,
    category_tree: dict[str, RozetkaCategory],
    royalty_index: dict[str, list[RoyaltyRule]],
) -> tuple[Decimal, str, str] | None:
    """
    Шукає % комісії для ціни 0 у батьківській категорії source_cat_id,
    піднімаючись вгору по ієрархії до першого предка, що покриває ціну 0.

    Правило Розетки:
      «У разі відсутності в певній категорії відсоткової ставки для першого
      діапазону (від 0 до другого діапазону), для нього діє ставка категорії
      вищого рівня.»

    Покриття ціни 0 є якщо у предка є:
      - рядок-прочерк («-», тобто price_from_dec=0, price_to_dec=Infinity)
      - або конкретний діапазон що починається з 0 (price_from_dec=0)

    Повертає (royalty_percent, cat_id, cat_name) або None якщо не знайдено.
    """
    visited:    set[str] = set()
    cat        = category_tree.get(source_cat_id)
    current_id = cat.parent_id if cat else ""
    depth      = 0

    while current_id and depth < MAX_HIERARCHY_DEPTH:
        if current_id in visited:
            log.warning("gap_fill: cycle detected at id=%s", current_id)
            break
        visited.add(current_id)

        rules = royalty_index.get(current_id, [])
        for rule in rules:
            # Перевіряємо лише правила для «всіх брендів» (brand="")
            if rule.brand != "":
                continue
            # Правило покриває ціну 0 якщо price_from_dec == 0
            # (як dash-рядок з price_from_dec=0,Inf, так і діапазон 0-X)
            if rule.price_from_dec == _PRICE_ZERO:
                cat_info = category_tree.get(current_id)
                cat_name = cat_info.name if cat_info else current_id
                return rule.royalty_percent, current_id, cat_name

        parent_cat = category_tree.get(current_id)
        if parent_cat is None or not parent_cat.parent_id:
            break
        current_id = parent_cat.parent_id
        depth += 1

    return None


def _has_gap_at_zero(rules: list[RoyaltyRule]) -> tuple[bool, Decimal]:
    """
    Визначає чи є «gap» у правилах для ціни від 0.

    Gap існує якщо:
      - немає рядка-прочерка (dash rule: price_from_dec=0, price_to_dec=Infinity)
      - і мінімальна price_from_dec серед конкретних діапазонів > 0

    Повертає (has_gap, min_price_from).
    min_price_from — нижня межа першого конкретного діапазону (або 0 якщо gap=False).
    """
    has_dash_rule = any(
        r.price_from_dec == _PRICE_ZERO and r.price_to_dec == Decimal("Infinity")
        for r in rules
        if r.brand == ""
    )
    if has_dash_rule:
        return False, _PRICE_ZERO

    specific = [r for r in rules if r.brand == "" and r.price_from != ""]
    if not specific:
        return False, _PRICE_ZERO

    min_from = min(r.price_from_dec for r in specific)
    return min_from > _PRICE_ZERO, min_from


def resolve_match_rules(
    match:         MatchResult,
    category_tree: dict[str, RozetkaCategory],
    royalty_index: dict[str, list[RoyaltyRule]],
) -> list[ResolvedRule]:
    """
    Конвертує MatchResult у список ResolvedRule, додаючи синтетичне правило
    для першого діапазону [0, min_price_from-1] якщо він відсутній у маппінгу.

    Логіка gap-fill:
      1. Визначаємо чи є gap (_has_gap_at_zero).
      2. Якщо є → шукаємо ставку у батьківській категорії (find_gap_fill_rate).
      3. Синтетичне правило: price_from=0, price_to=min_from-1, % з батька.
      4. Синтетичне правило додається ПЕРШИМ у список (найменший діапазон).
      5. В match_level_str синтетичного правила вказується "gap:<source_cat_id>"
         щоб було видно звідки взята ставка.

    Якщо батьківська ставка не знайдена (нестандартна ієрархія) — логується warning,
    синтетичне правило НЕ додається (безпечний fallback: інші правила виводяться як є).
    """
    resolved: list[ResolvedRule] = []
    level_str = str(match.match_level)

    has_gap, min_from = _has_gap_at_zero(match.rules)

    if has_gap:
        gap_price_to = min_from - Decimal("1")
        fill = find_gap_fill_rate(match.matched_id, category_tree, royalty_index)

        if fill:
            fill_percent, fill_cat_id, fill_cat_name = fill
            synthetic = RoyaltyRule(
                brand           = "",
                price_from      = "0",
                price_to        = format_decimal(gap_price_to),
                price_from_dec  = _PRICE_ZERO,
                price_to_dec    = gap_price_to,
                royalty_percent = fill_percent,
            )
            resolved.append(ResolvedRule(
                rule            = synthetic,
                matched_id      = fill_cat_id,
                matched_name    = fill_cat_name,
                match_level_str = f"gap:{match.matched_id}",
            ))
            log.info(
                "gap_fill: cat=%s (%s) | gap [0-%s] → %s%% з батьківської %s (%s)",
                match.matched_id, match.matched_name,
                format_decimal(gap_price_to),
                format_decimal(fill_percent),
                fill_cat_id, fill_cat_name,
            )
        else:
            log.warning(
                "gap_fill: cat=%s (%s) | gap [0-%s] → батьківська ставка НЕ знайдена",
                match.matched_id, match.matched_name,
                format_decimal(gap_price_to),
            )

    for rule in match.rules:
        resolved.append(ResolvedRule(
            rule            = rule,
            matched_id      = match.matched_id,
            matched_name    = match.matched_name,
            match_level_str = level_str,
        ))

    return resolved


# ---------------------------------------------------------------------------
# Existing defaults
# ---------------------------------------------------------------------------


def read_existing_defaults(path: Path) -> tuple[Decimal, Decimal]:
    if not path.exists():
        return DEFAULT_UNCATEGORIZED_COEF, DEFAULT_NO_BASE_COEF

    coef_unc: Decimal | None = None
    coef_nb:  Decimal | None = None

    with path.open(encoding=CSV_ENCODING, errors="replace", newline="") as fh:
        reader = csv.DictReader(fh, delimiter=CSV_DELIMITER)
        if not reader.fieldnames:
            return DEFAULT_UNCATEGORIZED_COEF, DEFAULT_NO_BASE_COEF

        for row in reader:
            if coef_unc is None:
                raw = (row.get("coef_uncategorized") or "").strip()
                if raw:
                    try:
                        coef_unc = parse_decimal(raw)
                    except InvalidOperation:
                        log.warning("Invalid coef_uncategorized=%r in %s", raw, path.name)

            if coef_nb is None:
                raw = (row.get("coef_no_base") or "").strip()
                if raw:
                    try:
                        coef_nb = parse_decimal(raw)
                    except InvalidOperation:
                        log.warning("Invalid coef_no_base=%r in %s", raw, path.name)

            if coef_unc is not None and coef_nb is not None:
                break

    return (
        coef_unc if coef_unc is not None else DEFAULT_UNCATEGORIZED_COEF,
        coef_nb  if coef_nb  is not None else DEFAULT_NO_BASE_COEF,
    )


# ---------------------------------------------------------------------------
# Build output rows
# ---------------------------------------------------------------------------


def build_output_rows(
    mappings:      list[MappingEntry],
    category_tree: dict[str, RozetkaCategory],
    royalty_index: dict[str, list[RoyaltyRule]],
    default_coef:  Decimal,
    no_base_coef:  Decimal,
) -> tuple[list[dict[str, str]], int, int, int, int, int]:
    """
    Повертає (rows, matched_direct, matched_ancestor, unmatched, generated_rules, gap_fills).
    Перший рядок — рядок дефолтів (порожній prom_category_id).
    """
    rows: list[dict[str, str]] = []

    defaults_row: dict[str, str] = {f: "" for f in OUTPUT_FIELDS}
    defaults_row["coef_uncategorized"] = format_decimal(default_coef)
    defaults_row["coef_no_base"]       = format_decimal(no_base_coef)
    rows.append(defaults_row)

    matched_direct   = 0
    matched_ancestor = 0
    unmatched        = 0
    generated        = 0
    gap_fills        = 0

    for entry in mappings:
        match = find_rules_by_hierarchy(entry.rozetka_id, category_tree, royalty_index)

        if match is None:
            unmatched += 1
            cat = category_tree.get(entry.rozetka_id)
            full_path = cat.full_path if cat else "unknown"
            log.warning(
                "prom_id=%s rozetka_id=%s (%s): no royalty match in hierarchy | path: %s",
                entry.category_id, entry.rozetka_id, entry.rozetka_name, full_path,
            )
            continue

        if match.match_level == 0:
            matched_direct += 1
        else:
            matched_ancestor += 1
            cat = category_tree.get(entry.rozetka_id)
            full_path = cat.full_path if cat else "unknown"
            log.info(
                "prom_id=%s rozetka_id=%s (%s): matched via ancestor "
                "id=%s (%s) level=%d | path: %s",
                entry.category_id, entry.rozetka_id, entry.rozetka_name,
                match.matched_id, match.matched_name, match.match_level,
                full_path,
            )

        # Розгортаємо правила з gap-fill якщо потрібно
        resolved_rules = resolve_match_rules(match, category_tree, royalty_index)

        for resolved in resolved_rules:
            rule = resolved.rule

            try:
                coef = calc_coef(rule.royalty_percent)
            except ValueError as exc:
                log.warning(
                    "prom_id=%s brand=%r range=%s-%s: %s, skipped",
                    entry.category_id, rule.brand, rule.price_from, rule.price_to, exc,
                )
                continue

            is_gap_fill = resolved.match_level_str.startswith("gap:")
            if is_gap_fill:
                gap_fills += 1

            row: dict[str, str] = {f: "" for f in OUTPUT_FIELDS}
            row["prom_category_id"]      = entry.category_id
            row["prom_category_name"]    = entry.category_name
            row["rozetka_category_id"]   = entry.rozetka_id
            row["rozetka_category_name"] = entry.rozetka_name
            row["matched_royalty_id"]    = resolved.matched_id
            row["matched_royalty_name"]  = resolved.matched_name
            row["match_level"]           = resolved.match_level_str
            row["brand"]                 = rule.brand
            row["royalty_percent"]       = format_decimal(rule.royalty_percent)
            row["price_from"]            = rule.price_from
            row["price_to"]              = rule.price_to
            row["coef"]                  = format_decimal(coef)
            rows.append(row)
            generated += 1

    return rows, matched_direct, matched_ancestor, unmatched, generated, gap_fills


# ---------------------------------------------------------------------------
# CSV write (idempotent)
# ---------------------------------------------------------------------------


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    buf = io.StringIO()
    writer = csv.DictWriter(
        buf,
        fieldnames=OUTPUT_FIELDS,
        delimiter=CSV_DELIMITER,
        lineterminator="\n",
    )
    writer.writeheader()
    writer.writerows(rows)
    content = buf.getvalue()

    if path.exists():
        with path.open("r+", encoding=CSV_ENCODING, newline="") as fh:
            fh.seek(0)
            fh.write(content)
            fh.truncate()
    else:
        path.write_text(content, encoding=CSV_ENCODING)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    log.info("=== rozetka coefficients export start ===")

    for p in (MAPPINGS_PATH, ROYALTY_PATH):
        if not p.exists():
            log.error("File not found: %s", p)
            sys.exit(1)

    default_coef, no_base_coef = read_existing_defaults(OUTPUT_CSV_PATH)

    mappings      = load_mappings(MAPPINGS_PATH, MAPPINGS_SHEET)
    category_tree = load_category_tree(MAPPINGS_PATH, CATEGORIES_SHEET)
    royalty_index = load_royalty_index(ROYALTY_PATH, ROYALTY_SHEET)

    rows, matched_direct, matched_ancestor, unmatched, generated, gap_fills = build_output_rows(
        mappings, category_tree, royalty_index, default_coef, no_base_coef,
    )

    write_csv(OUTPUT_CSV_PATH, rows)

    log.info(
        "=== done: matched_direct=%d, matched_ancestor=%d, unmatched=%d, "
        "rules=%d (gap_fills=%d), "
        "coef_uncategorized=%s, coef_no_base=%s ===",
        matched_direct, matched_ancestor, unmatched,
        generated, gap_fills,
        format_decimal(default_coef),
        format_decimal(no_base_coef),
    )
    log.info("saved: %s", OUTPUT_CSV_PATH)


if __name__ == "__main__":
    main()
