"""
epicenter_export_attr_options.py
---------------------------------
КРОК 5 з 5 пайплайну маппінгу Prom → Epicenter.

Що робить:
  Завантажує з API опції атрибутів Epicenter і записує в лист «Опції атрибутів»
  ТІЛЬКИ для тих пар (set_code, attr_code), у яких isRequired = TRUE
  у листі «Сети атрибутів» — і тільки для set_codes з «Маппінгу».

Передумова:
  • Лист «Маппінг» — заповнені epicenter_category_id.
  • Лист «Сети атрибутів» — наявні рядки з isRequired = TRUE.

Архітектура агрегації:
  GLOBAL_ATTR_CODES — агрегуємо всі set_codes в один рядок.
    Ці атрибути однозначні для всіх категорій (одиниця виміру, країна, бренд тощо).
    Дефолтна опція одна на всі категорії — економить місце.

  Решта attr_codes — окремий рядок на кожен (set_code, attr_code).
    Дозволяє задати різний default_option_code для різних категорій
    (наприклад, «режим роботи» у дрелі і картці доступу — різні опції).

Інкрементальна логіка:
  Якщо «Опції атрибутів» вже існує — дописує тільки нові set_codes,
  існуючі рядки не чіпає.

  Для GLOBAL attrs — оновлює set_codes in-place, нових рядків не створює.
  Для scoped attrs — перевіряє пару (set_code, attr_code); якщо вже є — пропускає.

Наступний крок (КРОК 6):
  Заповни prom_option_name у «Опції атрибутів» (колонка H) вручну або скриптом.

Запуск:
    python scripts/epicenter_export_attr_options.py
"""

from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from threading import Lock
import time

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_TOKEN = "5a6489d1a5c48c9d174bd31f2a0a8fd0"
BASE_URL  = "https://api.epicentrm.com.ua/v2/pim"
HEADERS   = {
    "Authorization": f"Bearer {API_TOKEN}",
    "Content-Type": "application/json",
    "Accept": "application/json",
}

ROOT        = Path(__file__).parents[1]
OUTPUT_PATH = ROOT / "data" / "markets" / "epicenter_mappings.xlsx"

OPTION_TYPES           = {"select", "multiselect"}
NON_OPTION_TYPES       = {"float", "int", "text", "string"}
EXTRA_NON_OPTION_TYPES = {"boolean", "bool", "date", "datetime", "color",
                          "image", "file", "price", "range",
                          "richtext", "rich_text", "textarea", "array"}
ALL_NON_OPTION_TYPES   = NON_OPTION_TYPES | EXTRA_NON_OPTION_TYPES
ALL_KNOWN_TYPES        = OPTION_TYPES | ALL_NON_OPTION_TYPES

# ---------------------------------------------------------------------------
# Стратегія агрегації
# ---------------------------------------------------------------------------
#
# GLOBAL_ATTR_CODES — атрибути, значення яких однозначні для всіх категорій.
# Для них всі set_codes збираємо в один рядок (компактно, дефолт один на всіх).
#
# Решта attr_codes — «scoped»: окремий рядок на кожен set_code, щоб для кожної
# категорії можна було поставити свій default_option_code.
# Наприклад, «режим роботи» у дрелі і картці доступу — різні опції за замовчуванням.
#
GLOBAL_ATTR_CODES: frozenset[str] = frozenset({
    "ratio",
    "length",
    "width",
    "weight",
    "height",
    "measure",
    "country_of_origin",
    "brand",
})

OPTIONS_WORKERS  = 8
REQ_TIMEOUT      = (10, 30)
MAX_PAGES        = 2000


# ─── Session ──────────────────────────────────────────────────────────────────

def _make_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(total=2, backoff_factor=0.5, status_forcelist=(429,), allowed_methods=["GET"])
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update(HEADERS)
    return session


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _norm_id(val: object) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if "." in s:
        try:
            f = float(s)
            if f == int(f):
                return str(int(f))
        except (ValueError, OverflowError):
            pass
    return s


def _get_translation(translations: list[dict], lang: str = "ua") -> str:
    for priority_lang in (lang, "ua", "uk", "ru", "en"):
        for t in translations:
            if t.get("languageCode") == priority_lang:
                val = t.get("value") or t.get("title") or ""
                if str(val).strip():
                    return str(val).strip()
    return ""


def _parse_option_name(opt: dict) -> str:
    translations = opt.get("translations", [])
    if translations:
        for lang in ("ua", "ru", "en", "uk"):
            for t in translations:
                if t.get("languageCode") == lang:
                    val = t.get("value") or t.get("title") or ""
                    if str(val).strip():
                        return str(val).strip()
        for t in translations:
            val = t.get("value") or t.get("title") or ""
            if str(val).strip():
                return str(val).strip()
    for field in ("name", "title", "label", "value"):
        val = opt.get(field, "")
        if val and str(val).strip():
            return str(val).strip()
    return ""


# ─── Styles ───────────────────────────────────────────────────────────────────

HDR_FILL    = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
ORANGE_FILL = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _hdr(cell, fill=HDR_FILL) -> None:
    cell.font      = HDR_FONT
    cell.fill      = fill
    cell.alignment = CENTER
    cell.border    = THIN_BORDER


# ─── Readers ──────────────────────────────────────────────────────────────────

def load_mapped_set_codes() -> set[str]:
    """Повертає заповнені epicenter_category_id з листа «Маппінг»."""
    if not OUTPUT_PATH.exists():
        return set()
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(OUTPUT_PATH, read_only=True, data_only=True)
        if "Маппінг" not in wb.sheetnames:
            wb.close()
            return set()
        rows = list(wb["Маппінг"].iter_rows(values_only=True))
        wb.close()
        if not rows:
            return set()
        headers = [str(c).strip() if c else "" for c in rows[0]]
        try:
            col = headers.index("epicenter_category_id")
        except ValueError:
            return set()
        codes = {
            _norm_id(row[col])
            for row in rows[1:]
            if len(row) > col and row[col]
        }
        print(f"   Знайдено {len(codes)} epicenter_category_id у «Маппінг».")
        return codes
    except Exception as e:
        print(f"⚠️  Не вдалося прочитати «Маппінг»: {e}")
        return set()


def load_attr_set_meta(filter_set_codes: set[str]) -> dict[tuple[str, str], dict]:
    """
    Читає «Сети атрибутів» і повертає метадані для пар (set_code, attr_code).
    Включає пару ТІЛЬКИ якщо isRequired = TRUE і set_code є в filter_set_codes.
    """
    if not OUTPUT_PATH.exists():
        return {}
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(OUTPUT_PATH, read_only=True, data_only=True)
        if "Сети атрибутів" not in wb.sheetnames:
            wb.close()
            return {}
        rows = list(wb["Сети атрибутів"].iter_rows(values_only=True))
        wb.close()
        if not rows:
            return {}
        headers = [str(c).strip() if c else "" for c in rows[0]]
        idx = {h: i for i, h in enumerate(headers)}
        required_cols = {"set_code", "set_name_uk", "attr_code", "attr_name_uk",
                         "attr_type", "isRequired", "prom_param_name"}
        if not required_cols.issubset(idx):
            missing = required_cols - idx.keys()
            print(f"⚠️  Відсутні колонки в «Сети атрибутів»: {missing}")
            return {}

        meta: dict[tuple[str, str], dict] = {}
        skipped_no_match: dict[str, int] = {}

        for row in rows[1:]:
            sc = _norm_id(row[idx["set_code"]])
            ac = _norm_id(row[idx["attr_code"]])
            if not sc or not ac or sc not in filter_set_codes:
                continue

            atype = str(row[idx["attr_type"]] or "").strip().lower()
            if atype not in ALL_KNOWN_TYPES:
                skipped_no_match[atype or "(порожній)"] = skipped_no_match.get(atype or "(порожній)", 0) + 1
                continue

            prom = str(row[idx["prom_param_name"]] or "").strip() \
                if len(row) > idx["prom_param_name"] else ""
            is_required = str(row[idx["isRequired"]] or "").strip().upper() in ("TRUE", "1", "YES")
            if not is_required:
                continue

            meta[(sc, ac)] = {
                "set_name":    str(row[idx["set_name_uk"]] or "").strip(),
                "attr_name":   str(row[idx["attr_name_uk"]] or "").strip(),
                "attr_type":   atype,
                "prom_param":  prom,
                "needs_default": not bool(prom),
            }

        with_opts     = sum(1 for m in meta.values() if m["attr_type"] in OPTION_TYPES)
        no_opts       = len(meta) - with_opts
        needs_default = sum(1 for m in meta.values() if m["needs_default"])
        has_prom      = len(meta) - needs_default
        print(
            f"   Пар isRequired=TRUE для обробки: {len(meta)} шт.\n"
            f"   • select/multiselect (завантажуємо опції з API): {with_opts}\n"
            f"   • float/int/text/string/ін. (рядок без опцій):  {no_opts}\n"
            f"   • без prom_param_name (needs_default=TRUE):      {needs_default}\n"
            f"   • з prom_param_name (needs_default=FALSE):       {has_prom}"
        )
        if skipped_no_match:
            total_skipped = sum(skipped_no_match.values())
            breakdown = ", ".join(
                f"{t!r}×{n}" for t, n in sorted(skipped_no_match.items(), key=lambda x: -x[1])
            )
            print(
                f"   ⚠️  Пропущено (невідомий тип): {total_skipped}\n"
                f"       Типи: {breakdown}\n"
                f"       → Додай в EXTRA_NON_OPTION_TYPES або OPTION_TYPES."
            )
        return meta
    except Exception as e:
        print(f"⚠️  Не вдалося прочитати метадані атрибутів: {e}")
        return {}


def load_set_codes_with_options() -> set[str]:
    """
    Повертає set_codes, для яких вже є рядки в «Опції атрибутів».
    Читає як колонку set_codes (comma-separated), так і застарілу set_code.
    """
    if not OUTPUT_PATH.exists():
        return set()
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(OUTPUT_PATH, read_only=True, data_only=True)
        if "Опції атрибутів" not in wb.sheetnames:
            wb.close()
            return set()
        rows = list(wb["Опції атрибутів"].iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2:
            return set()
        headers = [str(c).strip() if c else "" for c in rows[0]]
        col_name = "set_codes" if "set_codes" in headers else ("set_code" if "set_code" in headers else None)
        if not col_name:
            return set()
        sc_col = headers.index(col_name)
        codes: set[str] = set()
        for row in rows[1:]:
            if len(row) > sc_col and row[sc_col]:
                for sc in str(row[sc_col]).split(","):
                    sc = sc.strip()
                    if sc:
                        codes.add(sc)
        print(f"   set_codes вже з опціями: {len(codes)} шт.")
        return codes
    except Exception as e:
        print(f"⚠️  Не вдалося прочитати «Опції атрибутів»: {e}")
        return set()


def load_existing_global_attr_codes() -> set[str]:
    """
    Повертає attr_codes з GLOBAL_ATTR_CODES, які вже є в «Опції атрибутів».
    Використовується для in-place оновлення set_codes без нових рядків.
    """
    if not OUTPUT_PATH.exists():
        return set()
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(OUTPUT_PATH, read_only=True, data_only=True)
        if "Опції атрибутів" not in wb.sheetnames:
            wb.close()
            return set()
        rows = list(wb["Опції атрибутів"].iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2:
            return set()
        headers = [str(c).strip() if c else "" for c in rows[0]]
        try:
            ac_col = headers.index("attr_code")
        except ValueError:
            return set()
        codes = {
            str(row[ac_col]).strip()
            for row in rows[1:]
            if len(row) > ac_col and row[ac_col]
            and str(row[ac_col]).strip() in GLOBAL_ATTR_CODES
        }
        print(f"   Існуючих global attr_codes у «Опції атрибутів»: {len(codes)} шт.")
        return codes
    except Exception as e:
        print(f"⚠️  Не вдалося прочитати «Опції атрибутів»: {e}")
        return set()


def load_existing_scoped_sc_ac_pairs() -> set[tuple[str, str]]:
    """
    Повертає (set_code, attr_code) пари для NON-global атрибутів,
    які вже є в «Опції атрибутів».
    Використовується щоб не дублювати рядки при повторному запуску.
    """
    if not OUTPUT_PATH.exists():
        return set()
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(OUTPUT_PATH, read_only=True, data_only=True)
        if "Опції атрибутів" not in wb.sheetnames:
            wb.close()
            return set()
        rows = list(wb["Опції атрибутів"].iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2:
            return set()
        headers = [str(c).strip() if c else "" for c in rows[0]]
        try:
            ac_col = headers.index("attr_code")
            sc_col = headers.index("set_codes")
        except ValueError:
            return set()
        pairs: set[tuple[str, str]] = set()
        for row in rows[1:]:
            if len(row) <= max(ac_col, sc_col):
                continue
            ac = str(row[ac_col] or "").strip()
            if not ac or ac in GLOBAL_ATTR_CODES:
                continue
            for sc in str(row[sc_col] or "").split(","):
                sc = sc.strip()
                if sc:
                    pairs.add((sc, ac))
        print(f"   Існуючих scoped (set_code, attr_code) пар: {len(pairs)} шт.")
        return pairs
    except Exception as e:
        print(f"⚠️  Не вдалося прочитати scoped пари: {e}")
        return set()


# ─── Attr aggregation ─────────────────────────────────────────────────────────

class _AttrAgg:
    """Агрегат по attr_code або по (set_code, attr_code)."""
    __slots__ = ("attr_code", "first_set", "attr_name", "attr_type",
                 "set_codes", "prom_params", "needs_default")

    def __init__(self, ac: str, sc: str, meta: dict) -> None:
        self.attr_code     = ac
        self.first_set     = sc
        self.attr_name     = meta["attr_name"]
        self.attr_type     = meta["attr_type"]
        self.set_codes:    list[str] = [sc]
        self.prom_params:  list[str] = [meta["prom_param"]] if meta["prom_param"] else []
        self.needs_default = meta["needs_default"]

    def merge(self, sc: str, meta: dict) -> None:
        if sc not in self.set_codes:
            self.set_codes.append(sc)
        if meta["prom_param"] and meta["prom_param"] not in self.prom_params:
            self.prom_params.append(meta["prom_param"])
        if meta["needs_default"]:
            self.needs_default = True


def _build_agg_dicts(
    pair_meta: dict[tuple[str, str], dict],
) -> tuple[
    dict[str, _AttrAgg],            # global_ac_agg   — select/multiselect, key=attr_code
    dict[str, _AttrAgg],            # global_nopt_agg — numeric/text,       key=attr_code
    dict[tuple[str, str], _AttrAgg],# scoped_ac_agg   — select/multiselect, key=(sc, ac)
    dict[tuple[str, str], _AttrAgg],# scoped_nopt_agg — numeric/text,       key=(sc, ac)
]:
    """
    Розбиває pair_meta на 4 словники за стратегією агрегації.

    GLOBAL_ATTR_CODES → всі set_codes в один рядок (ключ = attr_code).
    Решта            → окремий рядок на кожен set_code (ключ = (sc, ac)).
    """
    global_ac_agg:   dict[str, _AttrAgg]             = {}
    global_nopt_agg: dict[str, _AttrAgg]             = {}
    scoped_ac_agg:   dict[tuple[str, str], _AttrAgg] = {}
    scoped_nopt_agg: dict[tuple[str, str], _AttrAgg] = {}

    for (sc, ac), meta in pair_meta.items():
        is_option_type = meta["attr_type"] in OPTION_TYPES
        is_global      = ac in GLOBAL_ATTR_CODES

        if is_global:
            target = global_ac_agg if is_option_type else global_nopt_agg
            if ac not in target:
                target[ac] = _AttrAgg(ac, sc, meta)
            else:
                target[ac].merge(sc, meta)
        else:
            target = scoped_ac_agg if is_option_type else scoped_nopt_agg
            key = (sc, ac)
            if key not in target:
                target[key] = _AttrAgg(ac, sc, meta)
            # scoped: один set_code на ключ — merge не потрібен

    return global_ac_agg, global_nopt_agg, scoped_ac_agg, scoped_nopt_agg


# ─── Streaming writer helpers ──────────────────────────────────────────────────

HEADERS_OPTIONS    = ["attr_code", "attr_name_uk", "attr_type",
                      "option_code", "option_name_uk", "prom_option_name",
                      "needs_default", "default_option_code",
                      "set_codes", "prom_param_name"]
COL_WIDTHS_OPTIONS = [30, 42, 16, 30, 45, 45, 14, 30, 60, 60]


def _row_vals(agg: _AttrAgg, opt: dict | None) -> list:
    """Будує список значень для одного рядка (порядок = HEADERS_OPTIONS)."""
    return [
        agg.attr_code,
        agg.attr_name,
        agg.attr_type,
        opt.get("code", "") if opt else "",
        _parse_option_name(opt) if opt else "",
        "",                           # prom_option_name — заповнює юзер
        agg.needs_default,
        "",                           # default_option_code — заповнює юзер
        ", ".join(agg.set_codes),
        ", ".join(agg.prom_params),
    ]


def _write_row_to_ws(ws, ri: int, vals: list) -> None:
    needs_default: bool = vals[6]
    for ci, val in enumerate(vals, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        if needs_default:
            cell.fill = ORANGE_FILL
        cell.font      = Font(name="Calibri", size=14)
        cell.alignment = LEFT


def _setup_options_sheet(wb) -> tuple:
    if "Опції атрибутів" not in wb.sheetnames:
        ws = wb.create_sheet("Опції атрибутів")
        for ci, (h, w) in enumerate(zip(HEADERS_OPTIONS, COL_WIDTHS_OPTIONS), 1):
            _hdr(ws.cell(row=1, column=ci, value=h))
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"
        return ws, 2
    ws = wb["Опції атрибутів"]
    return ws, ws.max_row + 1


def _update_existing_global_set_codes(
    ws,
    update_agg: dict[str, _AttrAgg],
) -> int:
    """
    Для GLOBAL attr_codes, які вже є в листі: дописує нові set_codes і prom_params
    до кожного рядка цього attr_code. Нові рядки НЕ створюються.

    Тільки GLOBAL_ATTR_CODES потрапляють сюди — для них один рядок покриває всі категорії.
    """
    if not update_agg:
        return 0

    headers = [
        str(ws.cell(row=1, column=c).value or "").strip()
        for c in range(1, ws.max_column + 1)
    ]
    try:
        ac_col = headers.index("attr_code")       + 1
        sc_col = headers.index("set_codes")       + 1
        pp_col = headers.index("prom_param_name") + 1
        nd_col = headers.index("needs_default")   + 1
    except ValueError as exc:
        print(f"⚠️  Не вдалося знайти колонку в «Опції атрибутів»: {exc}")
        return 0

    new_sc_by_ac: dict[str, list[str]] = {ac: agg.set_codes   for ac, agg in update_agg.items()}
    new_pp_by_ac: dict[str, list[str]] = {ac: agg.prom_params for ac, agg in update_agg.items()}
    nd_upgrade:   set[str]             = {ac for ac, agg in update_agg.items() if agg.needs_default}
    touched:      set[str]             = set()

    for row_idx in range(2, ws.max_row + 1):
        ac = str(ws.cell(row=row_idx, column=ac_col).value or "").strip()
        if ac not in update_agg:
            continue

        existing_sc_list = [s.strip() for s in str(ws.cell(row=row_idx, column=sc_col).value or "").split(",") if s.strip()]
        appended_sc = [s for s in new_sc_by_ac[ac] if s not in set(existing_sc_list)]
        if appended_sc:
            ws.cell(row=row_idx, column=sc_col).value = ", ".join(existing_sc_list + appended_sc)

        existing_pp_list = [p.strip() for p in str(ws.cell(row=row_idx, column=pp_col).value or "").split(",") if p.strip()]
        appended_pp = [p for p in new_pp_by_ac[ac] if p not in set(existing_pp_list)]
        if appended_pp:
            ws.cell(row=row_idx, column=pp_col).value = ", ".join(existing_pp_list + appended_pp)

        if ac in nd_upgrade:
            cell_nd = ws.cell(row=row_idx, column=nd_col)
            if not cell_nd.value:
                cell_nd.value = True
                cell_nd.fill  = ORANGE_FILL

        touched.add(ac)

    count = len(touched)
    if count:
        print(f"   ✅ Global: оновлено set_codes у {count} існуючих attr_codes: {sorted(touched)}")
    return count


# ─── API ──────────────────────────────────────────────────────────────────────

def _try_fetch_options(
    session: requests.Session, attr_code: str, set_code: str
) -> list[dict] | None:
    options: list[dict] = []
    total_pages = 1
    page = 1
    while page <= MAX_PAGES:
        try:
            resp = session.get(
                f"{BASE_URL}/attribute-sets/{set_code}/attributes/{attr_code}/options",
                params={"page": page},
                timeout=REQ_TIMEOUT,
            )
            if resp.status_code in (403, 404):
                return None
            resp.raise_for_status()
            data = resp.json()
        except Exception:
            return None
        batch = data.get("items", [])
        if not batch:
            break
        options.extend(batch)
        total_pages = data.get("pages", 1)
        if page % 50 == 0:
            print(f"      attr={attr_code}: стор. {page}/{total_pages} ({len(options)} опцій)...")
        if page >= total_pages:
            break
        page += 1
    if page > MAX_PAGES:
        print(f"      ⚠️  attr={attr_code} досягнуто safety cap {MAX_PAGES} стор. ({len(options)} опцій)")
    return options


def _fetch_options_one_attr(attr_code: str, set_codes: list[str]) -> list[dict]:
    """Завантажує опції, перебираючи set_codes до першого валідного."""
    session = _make_session()
    for set_code in set_codes:
        result = _try_fetch_options(session, attr_code, set_code)
        if result is not None:
            return result
    return []


# ─── Parallel fetch + streaming write ─────────────────────────────────────────

def _fetch_and_write_parallel(
    ac_agg_items: list[tuple[object, _AttrAgg]],
    nopt_agg_items: list[tuple[object, _AttrAgg]],
    ws,
    start_row: int,
    label: str,
    sort_by_set_code: bool = False,
) -> int:
    """
    Загальна функція паралельного завантаження + запису для будь-якого набору агрегатів.
    Приймає items (key, agg) — ключ використовується тільки для відображення прогресу.

    sort_by_set_code=True (для scoped атрибутів):
        Збирає всі результати в пам'яті, сортує по (set_code, attr_code),
        потім пише — рядки однієї категорії йдуть підряд.
        as_completed повертає futures у порядку завершення (не сабмісії),
        тому без сортування порядок недетермінований.

    sort_by_set_code=False (для global атрибутів):
        Пише одразу при завершенні future — RAM мінімальна, порядок не критичний.

    Повертає номер наступного вільного рядка.
    """
    total = len(ac_agg_items)
    print(
        f"\n⬇️  [{label}] select/multiselect: {total} attr_codes ({OPTIONS_WORKERS} потоків)"
        f"\n   [{label}] числові/текстові (без API): {len(nopt_agg_items)}"
    )

    current_row = start_row

    def _worker(key, agg: _AttrAgg) -> tuple[object, _AttrAgg, list[dict], float]:
        t0 = time.monotonic()
        opts = _fetch_options_one_attr(agg.attr_code, agg.set_codes)
        return key, agg, opts, time.monotonic() - t0

    fetch_start = time.monotonic()
    total_written = 0

    if sort_by_set_code:
        # ── Scoped: збираємо всі результати, сортуємо, пишемо ──────────────
        # Сортування потрібне бо as_completed повертає futures у довільному
        # порядку — рядки одного set_code перемішуються з іншими.
        # Після сортування по (set_code, attr_code) всі рядки однієї категорії
        # йдуть підряд — зручно для ручного заповнення default_option_code.

        # 1. Паралельно завантажуємо опції
        fetched: list[tuple[_AttrAgg, list[dict]]] = []
        done = 0
        with ThreadPoolExecutor(max_workers=OPTIONS_WORKERS) as pool:
            future_map = {pool.submit(_worker, k, agg): (k, agg) for k, agg in ac_agg_items}
            for future in as_completed(future_map):
                _, agg = future_map[future]
                try:
                    _, agg, opts, elapsed = future.result()
                except Exception as e:
                    opts = []
                    print(f"   ⚠️  ERROR attr={agg.attr_code}: {e}")
                done += 1
                fetched.append((agg, opts))
                print(f"   [{done}/{total}] attr={agg.attr_code} sc={agg.set_codes[0]} → {len(opts)} опцій")

        # 2. Сортуємо: спочатку по set_code (числово якщо можливо), потім по attr_code
        def _sort_key(item: tuple[_AttrAgg, list]) -> tuple:
            sc = item[0].set_codes[0]
            try:
                return (int(sc), item[0].attr_code)
            except ValueError:
                return (0, sc + item[0].attr_code)

        fetched.sort(key=_sort_key)

        # 3. Числові/текстові також сортуємо і пишемо разом зі своїм set_code
        nopt_sorted = sorted(
            nopt_agg_items,
            key=lambda item: (
                int(item[1].set_codes[0]) if item[1].set_codes[0].isdigit() else 0,
                item[1].attr_code,
            ),
        )

        # 4. Об'єднуємо select і nopt, сортуємо разом
        all_rows: list[tuple[_AttrAgg, list[dict] | list[None]]] = [
            *fetched,
            *((agg, [None]) for _, agg in nopt_sorted),
        ]
        all_rows.sort(key=_sort_key)

        for agg, opts in all_rows:
            for opt in (opts or [None]):
                _write_row_to_ws(ws, current_row, _row_vals(agg, opt))
                current_row += 1
                total_written += 1

    else:
        # ── Global: пишемо одразу при завершенні future ─────────────────────
        lock = Lock()
        done = [0]

        # Рядки-заглушки для нот-опшн — пишемо одразу
        for _, agg in nopt_agg_items:
            _write_row_to_ws(ws, current_row, _row_vals(agg, None))
            current_row += 1
            total_written += 1
        if nopt_agg_items:
            print(f"   [{label}] Числові/текстові записано: {len(nopt_agg_items)}")

        if ac_agg_items:
            with ThreadPoolExecutor(max_workers=OPTIONS_WORKERS) as pool:
                future_map = {pool.submit(_worker, k, agg): (k, agg) for k, agg in ac_agg_items}
                for future in as_completed(future_map):
                    _, agg = future_map[future]
                    try:
                        _, agg, opts, elapsed = future.result()
                        elapsed_s = f"{elapsed:.1f}s"
                    except Exception as e:
                        opts = []
                        elapsed_s = "err"
                        print(f"   ⚠️  ERROR attr={agg.attr_code}: {e}")

                    rows_for_attr = opts or [None]
                    with lock:
                        for opt in rows_for_attr:
                            _write_row_to_ws(ws, current_row, _row_vals(agg, opt))
                            current_row += 1
                        done[0] += 1
                        total_written += len(rows_for_attr)
                        sc_label = f"sc={agg.set_codes[0]}" if len(agg.set_codes) == 1 else f"sc×{len(agg.set_codes)}"
                        print(f"   [{done[0]}/{total}] attr={agg.attr_code} {sc_label} → {len(opts)} опцій ({elapsed_s})")

    print(f"   [{label}] ⏱️  {time.monotonic() - fetch_start:.1f}s | рядків: {total_written}")
    return current_row


# ─── Main write orchestration ──────────────────────────────────────────────────

def append_option_rows(
    pair_meta: dict[tuple[str, str], dict],
    existing_global_ac: set[str],
    existing_scoped_pairs: set[tuple[str, str]],
) -> bool:
    """
    Відкриває xlsx, оновлює/дописує рядки за стратегією агрегації, зберігає.

    Global attrs (GLOBAL_ATTR_CODES):
      • вже є в листі → in-place: дописуємо set_codes до існуючих рядків
      • новий          → API + нові рядки

    Scoped attrs (решта):
      • (set_code, attr_code) вже є → пропускаємо (не дублюємо)
      • новий                       → API + новий рядок (один на set_code)

    Повертає True якщо хоча б щось змінено.
    """
    import openpyxl as _xl
    wb = _xl.load_workbook(OUTPUT_PATH)
    ws, start_row = _setup_options_sheet(wb)

    global_ac_agg, global_nopt_agg, scoped_ac_agg, scoped_nopt_agg = _build_agg_dicts(pair_meta)

    # --- Global: розбиваємо на update і new ---
    global_update_agg   = {ac: agg for ac, agg in {**global_ac_agg, **global_nopt_agg}.items()
                           if ac in existing_global_ac}
    new_global_ac_agg   = {ac: agg for ac, agg in global_ac_agg.items()   if ac not in existing_global_ac}
    new_global_nopt_agg = {ac: agg for ac, agg in global_nopt_agg.items() if ac not in existing_global_ac}

    # --- Scoped: фільтруємо вже існуючі (sc, ac) пари ---
    new_scoped_ac_agg   = {k: agg for k, agg in scoped_ac_agg.items()   if k not in existing_scoped_pairs}
    new_scoped_nopt_agg = {k: agg for k, agg in scoped_nopt_agg.items() if k not in existing_scoped_pairs}

    print(
        f"\n📋 Розподіл:\n"
        f"   Global attrs ({len(GLOBAL_ATTR_CODES)} кодів):\n"
        f"     • in-place оновлення (вже є в листі):  {len(global_update_agg)}"
        + (f" {sorted(global_update_agg)}" if global_update_agg else "") + "\n"
        f"     • нові select/multiselect (API):        {len(new_global_ac_agg)}\n"
        f"     • нові числові/текстові (без API):      {len(new_global_nopt_agg)}\n"
        f"   Scoped attrs (по одному рядку на set_code):\n"
        f"     • нові select/multiselect (API):        {len(new_scoped_ac_agg)}\n"
        f"     • нові числові/текстові (без API):      {len(new_scoped_nopt_agg)}\n"
        f"     • пропущено (вже є в листі):            "
        f"{len(scoped_ac_agg) + len(scoped_nopt_agg) - len(new_scoped_ac_agg) - len(new_scoped_nopt_agg)}"
    )

    # In-place оновлення для global
    updated_count = _update_existing_global_set_codes(ws, global_update_agg)

    # Нові рядки: спочатку global, потім scoped
    current_row = fetch_and_write_parallel = start_row

    if new_global_ac_agg or new_global_nopt_agg:
        current_row = _fetch_and_write_parallel(
            list(new_global_ac_agg.items()),
            list(new_global_nopt_agg.items()),
            ws, current_row, "global",
        )

    if new_scoped_ac_agg or new_scoped_nopt_agg:
        current_row = _fetch_and_write_parallel(
            list(new_scoped_ac_agg.items()),
            list(new_scoped_nopt_agg.items()),
            ws, current_row, "scoped",
            sort_by_set_code=True,
        )

    rows_written = current_row - start_row
    if rows_written == 0 and updated_count == 0:
        wb.close()
        return False

    print(f"   💾 Збереження xlsx ({rows_written} нових рядків, {updated_count} global оновлено)...")
    save_start = time.monotonic()
    wb.save(OUTPUT_PATH)
    wb.close()
    print(f"   ✅ Збережено за {time.monotonic() - save_start:.1f}s")
    return True


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    main_start = time.monotonic()
    print("🚀 epicenter_export_attr_options.py — КРОК 5\n")

    if not OUTPUT_PATH.exists():
        print(
            "❌ epicenter_mappings.xlsx не знайдено.\n"
            "   Спочатку виконай КРОК 1: python scripts/epicenter_export_categories.py"
        )
        return

    mapped_set_codes = load_mapped_set_codes()
    if not mapped_set_codes:
        print(
            "\n⏭️  Немає заповнених epicenter_category_id у «Маппінг».\n"
            "   Виконай КРОК 2: python scripts/epicenter_map_categories.py"
        )
        return

    existing_with_options = load_set_codes_with_options()
    new_set_codes = mapped_set_codes - existing_with_options

    if not new_set_codes:
        print("   ✅ Опції вже завантажено для всіх категорій.")
        print("   КРОК 6: Заповни prom_option_name (колонка H) у «Опції атрибутів».")
        return

    print(f"   Нових set_codes для завантаження: {len(new_set_codes)} шт.")

    pair_meta = load_attr_set_meta(new_set_codes)
    if not pair_meta:
        print(
            "\n⏭️  Немає атрибутів з isRequired=TRUE для нових категорій.\n"
            "   Перевір лист «Сети атрибутів», колонку isRequired."
        )
        return

    existing_global_ac     = load_existing_global_attr_codes()
    existing_scoped_pairs  = load_existing_scoped_sc_ac_pairs()

    written = append_option_rows(pair_meta, existing_global_ac, existing_scoped_pairs)
    if not written:
        print("⚠️  Опцій не записано.")
        return

    total_elapsed = time.monotonic() - main_start
    print(
        f"\n✅ Оновлено: {OUTPUT_PATH}  (загальний час: {total_elapsed:.1f}s)\n"
        "   КРОК 6: Заповни prom_option_name (колонка H) у «Опції атрибутів» вручну або скриптом."
    )


if __name__ == "__main__":
    main()
