# -*- coding: utf-8 -*-
"""
update_kod_product.py
---------------------
Алгоритм:
  1. Завантажує sku_map.json: { "000030614": 100100, ... }
  2. Відкриває export-products.xlsx.
  3. Для кожного рядка перевіряє Ідентифікатор_товару:
       - шукає збіг з ключем sku_map напряму АБО з префіксом "prom_"
       - якщо знайдено → встановлює Код_товару з sku_map
       - якщо не знайдено → встановлює Код_товару = 777777
  4. Логує які ідентифікатори з sku_map ВІДСУТНІ в Excel.
  5. Зберігає файл.
"""

from __future__ import annotations

import json
import logging
import sys
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Конфігурація
# ---------------------------------------------------------------------------
EXCEL_PATH   = Path(r"C:\FullStack\PriceFeedPipeline\data\export-products.xlsx")
SKU_MAP_PATH = Path(r"C:\FullStack\PriceFeedPipeline\data\secur\sku_map.json")

COL_KOD          = "Код_товару"
COL_IDENTIFIER   = "Ідентифікатор_товару"
FALLBACK_KOD     = 777777
PROM_PREFIX      = "prom_"

# ---------------------------------------------------------------------------
# Логування
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Завантаження sku_map
# ---------------------------------------------------------------------------

def load_sku_map(path: Path) -> tuple[dict[str, int], list[str]]:
    """
    Завантажує sku_map.json і будує lookup-словник одразу для двох форматів ключа:
      "000030614"      → 100100
      "prom_000030614" → 100100

    Повертає (lookup, bare_keys) — де bare_keys це оригінальні ключі без префіксу.
    """
    if not path.exists():
        log.error("Файл sku_map не знайдено: %s", path)
        sys.exit(1)

    with path.open(encoding="utf-8") as fh:
        raw: dict[str, int] = json.load(fh)

    lookup: dict[str, int] = {}
    bare_keys: list[str] = []

    for sku, kod in raw.items():
        bare = sku.strip()
        bare_keys.append(bare)
        lookup[bare] = int(kod)
        lookup[f"{PROM_PREFIX}{bare}"] = int(kod)

    log.info("Завантажено записів з sku_map: %d (lookup-розмір: %d)", len(raw), len(lookup))
    return lookup, bare_keys


# ---------------------------------------------------------------------------
# Основна логіка
# ---------------------------------------------------------------------------

def resolve_kod(identifier: str | None, lookup: dict[str, int]) -> int:
    """Повертає Код_товару з lookup або FALLBACK_KOD якщо не знайдено."""
    if identifier is None:
        return FALLBACK_KOD
    key = str(identifier).strip()
    return lookup.get(key, FALLBACK_KOD)


def update_excel(
    excel_path: Path,
    lookup: dict[str, int],
    bare_keys: list[str],
) -> None:
    if not excel_path.exists():
        log.error("Файл Excel не знайдено: %s", excel_path)
        sys.exit(1)

    log.info("Відкриваємо файл: %s", excel_path)
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # --- визначаємо індекси стовпців (1-based для openpyxl) ---
    header_row = [cell.value for cell in ws[1]]

    if COL_KOD not in header_row:
        log.error("Стовпець '%s' не знайдено у файлі.", COL_KOD)
        sys.exit(1)

    if COL_IDENTIFIER not in header_row:
        log.error("Стовпець '%s' не знайдено у файлі.", COL_IDENTIFIER)
        sys.exit(1)

    col_kod_idx: int        = header_row.index(COL_KOD) + 1        # 1-based
    col_identifier_idx: int = header_row.index(COL_IDENTIFIER) + 1 # 1-based

    log.info(
        "Стовпець '%s' → колонка %d | '%s' → колонка %d",
        COL_KOD, col_kod_idx, COL_IDENTIFIER, col_identifier_idx,
    )

    # --- збираємо всі ідентифікатори з Excel (нормалізовані до bare-формату) ---
    excel_identifiers: set[str] = set()

    total    = 0
    matched  = 0
    fallback = 0

    for row_idx in range(2, ws.max_row + 1):
        identifier = ws.cell(row=row_idx, column=col_identifier_idx).value

        # пропускаємо повністю порожні рядки
        if all(
            ws.cell(row=row_idx, column=c).value is None
            for c in range(1, min(5, ws.max_column + 1))
        ):
            continue

        # зберігаємо нормалізований ідентифікатор для перехресної перевірки
        if identifier is not None:
            raw_id = str(identifier).strip()
            # зводимо до bare-форми (прибираємо prom_ якщо є)
            bare_id = raw_id.removeprefix(PROM_PREFIX)
            excel_identifiers.add(bare_id)

        new_kod = resolve_kod(identifier, lookup)
        ws.cell(row=row_idx, column=col_kod_idx).value = new_kod

        total += 1
        if new_kod == FALLBACK_KOD:
            fallback += 1
        else:
            matched += 1

    # --- ідентифікатори з sku_map, яких НЕМАЄ в Excel ---
    missing_in_excel: list[str] = sorted(
        key for key in bare_keys if key not in excel_identifiers
    )

    if missing_in_excel:
        log.info("--- ІДЕНТИФІКАТОРИ З sku_map, ВІДСУТНІ В EXCEL ---")
        for sku_id in missing_in_excel:
            kod_value = lookup.get(sku_id, "?")
            log.info(
                "  Ідентифікатор товару '%s' (Код_товару: %s) — не знайдено у файлі",
                sku_id, kod_value,
            )
        log.info("Всього відсутніх в Excel: %d", len(missing_in_excel))
    else:
        log.info("Всі ідентифікатори з sku_map присутні в Excel.")

    # --- збереження ---
    log.info("Зберігаємо файл: %s", excel_path)
    wb.save(excel_path)
    wb.close()

    # --- підсумок ---
    log.info("=== ПІДСУМОК ===")
    log.info("  Всього оброблено рядків           : %d", total)
    log.info("  Знайдено в sku_map (оновлено)     : %d", matched)
    log.info("  Не знайдено → %-7d            : %d", FALLBACK_KOD, fallback)
    log.info("  Відсутніх у Excel з sku_map       : %d", len(missing_in_excel))
    log.info("✅ Файл збережено успішно.")


# ---------------------------------------------------------------------------
# Точка входу
# ---------------------------------------------------------------------------

def main() -> None:
    log.info("Завантаження sku_map: %s", SKU_MAP_PATH)
    lookup, bare_keys = load_sku_map(SKU_MAP_PATH)

    update_excel(EXCEL_PATH, lookup, bare_keys)


if __name__ == "__main__":
    main()
