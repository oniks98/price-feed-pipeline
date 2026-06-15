from __future__ import annotations

import logging
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Final

import openpyxl

logger = logging.getLogger(__name__)

# Путь к маппингам
_XLSX_PATH: Final[Path] = (
    Path(__file__).parents[2] / "data" / "markets" / "epicenter_mappings.xlsx"
)

_SHEET_OPTIONS: Final[str] = "Опції атрибутів"

# Глобальные обязательные 8 кодов Эпицентра
GLOBAL_FLOATS: Final[frozenset[str]] = frozenset({"height", "length", "width", "weight", "ratio"})
GLOBAL_SELECTS: Final[frozenset[str]] = frozenset({"measure", "country_of_origin", "brand"})

# Индексы колонок (0-based)
_OPT_COL_ATTR_CODE: Final[int] = 0
_OPT_COL_ATTR_NAME: Final[int] = 1
_OPT_COL_ATTR_TYPE: Final[int] = 2
_OPT_COL_OPTION_CODE: Final[int] = 3
_OPT_COL_OPTION_NAME: Final[int] = 4
_OPT_COL_PROM_VALUE: Final[int] = 5
_OPT_COL_NEEDS_DEFAULT: Final[int] = 6
_OPT_COL_DEFAULT_CODE: Final[int] = 7
_OPT_COL_SET_CODES: Final[int] = 8
_OPT_COL_PROM_PARAMS: Final[int] = 9

NON_OPTION_TYPES: Final[frozenset[str]] = frozenset({"float", "int", "text", "string", "array"})


# ---------------------------------------------------------------------------
# Public types
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class AttrOption:
    """Маппінг одної опції select/multiselect атрибута."""
    attr_code: str
    attr_name: str
    option_code: str
    option_name: str


@dataclass(frozen=True)
class AttrMeta:
    """Метадані числового / текстового атрибута (без опцій)."""
    attr_code: str
    attr_name: str
    attr_type: str


OptionMap      = dict[str, dict[str, list[AttrOption]]]
SetOptionMap   = dict[str, OptionMap]
DefaultsMap    = dict[str, dict[str, AttrOption]]
NumericMap     = dict[str, list[AttrMeta]]
SetNumericMap  = dict[str, NumericMap]
AttrDefaultsMap  = dict[str, AttrOption]
FloatDefaultsMap = dict[str, str]
NumericDefaultsMap = dict[str, dict[str, tuple[AttrMeta, str]]]


@dataclass(frozen=True)
class CategoryAttrRules:
    """Готові правила атрибутів для одного set_code."""

    set_code: str
    option_map: OptionMap
    numeric_map: NumericMap
    select_defaults: dict[str, AttrOption]
    numeric_defaults: dict[str, tuple[AttrMeta, str]]
    global_select_defaults: AttrDefaultsMap
    global_non_option_defaults: FloatDefaultsMap

    def system_select_default(self, attr_code: str) -> AttrOption | None:
        return self.select_defaults.get(attr_code) or self.global_select_defaults.get(attr_code)

    def global_select_default(self, attr_code: str) -> AttrOption | None:
        return self.global_select_defaults.get(attr_code)

    def option_param_targets_attr(self, prom_param_name: str, attr_code: str) -> bool:
        return any(
            option.attr_code == attr_code
            for options in self.option_map.get(prom_param_name, {}).values()
            for option in options
        )

    def prom_names_for_attr(self, attr_code: str) -> frozenset[str]:
        return frozenset(
            prom_name
            for prom_name in self.option_map
            if self.option_param_targets_attr(prom_name, attr_code)
        )


@dataclass(frozen=True)
class AttrIndexes:
    """Raw cached workbook indexes used to build CategoryAttrRules."""
    option_map: OptionMap
    set_option_map: SetOptionMap
    defaults: DefaultsMap
    numeric_map: NumericMap
    set_numeric_map: SetNumericMap
    attr_defaults: AttrDefaultsMap
    float_defaults: FloatDefaultsMap
    numeric_defaults: NumericDefaultsMap


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _clean(value: object) -> str:
    return str(value).strip() if value is not None else ""

def _parse_set_codes(raw: object) -> list[str]:
    if not raw:
        return []
    return [s.strip() for s in str(raw).replace(";", ",").split(",") if s.strip()]

def _parse_aliases(raw: object) -> list[str]:
    if not raw:
        return []
    return [s.strip() for s in str(raw).replace(",", ";").split(";") if s.strip()]

def _load_workbook() -> openpyxl.Workbook:
    if not _XLSX_PATH.exists():
        raise FileNotFoundError(f"epicenter_mappings.xlsx не знайдено: {_XLSX_PATH}")
    return openpyxl.load_workbook(_XLSX_PATH, read_only=True, data_only=True)

def _merge_options(opts: list[AttrOption], attr_code: str, row_idx: int) -> AttrOption:
    if len(opts) == 1:
        return opts[0]
    return AttrOption(
        attr_code=opts[0].attr_code,
        attr_name=opts[0].attr_name,
        option_code=",".join(o.option_code for o in opts),
        option_name=", ".join(o.option_name for o in opts),
    )


# ---------------------------------------------------------------------------
# Main cached loader
# ---------------------------------------------------------------------------

@lru_cache(maxsize=1)
def _load_indexes() -> AttrIndexes:
    wb = _load_workbook()
    try:
        ws_opts = wb[_SHEET_OPTIONS]
    except KeyError:
        raise KeyError(f"Аркуш «{_SHEET_OPTIONS}» не знайдено у {_XLSX_PATH}")

    option_map: OptionMap = {}
    set_option_map: SetOptionMap = {}
    defaults: DefaultsMap = {}
    numeric_map: NumericMap = {}
    set_numeric_map: SetNumericMap = {}
    attr_defaults: AttrDefaultsMap = {}
    float_defaults: FloatDefaultsMap = {}
    numeric_defaults: NumericDefaultsMap = {}

    key_index: dict[tuple[str, str], AttrOption] = {}
    set_key_index: dict[tuple[str, str, str], AttrOption] = {}

    pending_global_defaults: list[tuple[str, str, int]] = []
    pending_set_defaults: list[tuple[str, str, str, int]] = []

    for row_idx, row in enumerate(ws_opts.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue

        attr_code = _clean(row[_OPT_COL_ATTR_CODE])
        if not attr_code:
            continue

        attr_name = _clean(row[_OPT_COL_ATTR_NAME])
        attr_type_raw = _clean(row[_OPT_COL_ATTR_TYPE]).lower()
        option_code = _clean(row[_OPT_COL_OPTION_CODE])
        option_name_uk = _clean(row[_OPT_COL_OPTION_NAME])
        prom_option_aliases = _parse_aliases(row[_OPT_COL_PROM_VALUE])
        default_code = _clean(row[_OPT_COL_DEFAULT_CODE])
        set_codes = _parse_set_codes(row[_OPT_COL_SET_CODES])
        prom_param_aliases = _parse_aliases(row[_OPT_COL_PROM_PARAMS])

        # --- КЕЙС 1: ГЛОБАЛЬНЫЕ ЧИСЛОВЫЕ АТРИБУТЫ (height, length, etc.) ---
        if attr_code in GLOBAL_FLOATS:
            meta = AttrMeta(attr_code=attr_code, attr_name=attr_name, attr_type=attr_type_raw or "float")
            
            # Решение проблемы №2: если prom_param_name пуст, ключом становится имя или код
            aliases = prom_param_aliases if prom_param_aliases else ([option_name_uk] if option_name_uk else [attr_name])
            for alias in aliases:
                numeric_map.setdefault(alias, []).append(meta)
            
            if option_name_uk and attr_code not in float_defaults:
                float_defaults[attr_code] = option_name_uk

        # --- КЕЙС 2: ГЛОБАЛЬНЫЕ СЕЛЕКТЫ (measure, country_of_origin, brand) ---
        elif attr_code in GLOBAL_SELECTS:
            if option_code:
                opt = AttrOption(attr_code=attr_code, attr_name=attr_name, option_code=option_code, option_name=option_name_uk)
                key_index[(attr_code, option_code)] = opt

                # Решение проблемы №5: строгая совместная проверка алиасов параметров и опций
                if prom_param_aliases and prom_option_aliases:
                    for p_alias in prom_param_aliases:
                        for o_alias in prom_option_aliases:
                            option_map.setdefault(p_alias, {}).setdefault(o_alias, []).append(opt)

            if default_code:
                pending_global_defaults.append((attr_code, default_code, row_idx))

        # --- КЕЙС 3: ЛОКАЛЬНЫЕ КАТЕГОРИЙНЫЕ АТРИБУТЫ (строго по set_codes) ---
        else:
            for sc in set_codes:
                # А. Локальные числовые / текстовые (Решение проблемы №1 — убран опасный or)
                if attr_type_raw in NON_OPTION_TYPES:
                    meta = AttrMeta(attr_code=attr_code, attr_name=attr_name, attr_type=attr_type_raw)
                    for alias in prom_param_aliases:
                        set_numeric_map.setdefault(sc, {}).setdefault(alias, []).append(meta)
                    
                    if option_name_uk:
                        numeric_defaults.setdefault(sc, {})[attr_code] = (meta, option_name_uk)

                # Б. Локальные селекты
                else:
                    if option_code:
                        opt = AttrOption(attr_code=attr_code, attr_name=attr_name, option_code=option_code, option_name=option_name_uk)
                        set_key_index[(sc, attr_code, option_code)] = opt

                        if prom_param_aliases and prom_option_aliases:
                            for p_alias in prom_param_aliases:
                                for o_alias in prom_option_aliases:
                                    set_option_map.setdefault(sc, {}).setdefault(p_alias, {}).setdefault(o_alias, []).append(opt)

                    if default_code:
                        pending_set_defaults.append((sc, attr_code, default_code, row_idx))

    # --- РЕЗОЛВ ГЛОБАЛЬНЫХ ДЕФОЛТОВ ---
    for attr_code, def_code, r_idx in pending_global_defaults:
        codes = [c.strip() for c in def_code.replace(";", ",").split(",") if c.strip()]
        found = [key_index[(attr_code, c)] for c in codes if (attr_code, c) in key_index]
        if found:
            attr_defaults[attr_code] = _merge_options(found, attr_code, r_idx)

    # --- РЕЗОЛВ ЛОКАЛЬНЫХ ДЕФОЛТОВ ---
    for sc, attr_code, def_code, r_idx in pending_set_defaults:
        codes = [c.strip() for c in def_code.replace(";", ",").split(",") if c.strip()]
        
        # Решение проблемы №3: Исключительно строгий поиск по (set_code, attr_code, option_code)
        found = [set_key_index[(sc, attr_code, c)] for c in codes if (sc, attr_code, c) in set_key_index]
        if found:
            defaults.setdefault(sc, {})[attr_code] = _merge_options(found, attr_code, r_idx)

    wb.close()
    return AttrIndexes(
        option_map=option_map,
        set_option_map=set_option_map,
        defaults=defaults,
        numeric_map=numeric_map,
        set_numeric_map=set_numeric_map,
        attr_defaults=attr_defaults,
        float_defaults=float_defaults,
        numeric_defaults=numeric_defaults,
    )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

@lru_cache(maxsize=None)
def get_category_attr_rules(set_code: str) -> CategoryAttrRules:
    """Повертає готові правила для одного set_code."""
    indexes = _load_indexes()

    # Решение проблемы №4: Никаких блокировок из старой архитектуры. Прямой update.
    option_rules: OptionMap = dict(indexes.option_map)
    option_rules.update(indexes.set_option_map.get(set_code, {}))

    numeric_rules: NumericMap = dict(indexes.numeric_map)
    numeric_rules.update(indexes.set_numeric_map.get(set_code, {}))

    return CategoryAttrRules(
        set_code=set_code,
        option_map=option_rules,
        numeric_map=numeric_rules,
        select_defaults=indexes.defaults.get(set_code, {}),
        numeric_defaults=indexes.numeric_defaults.get(set_code, {}),
        global_select_defaults=indexes.attr_defaults,
        global_non_option_defaults=indexes.float_defaults,
    )