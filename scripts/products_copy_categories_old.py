"""
products_copy_categories_old.py
-----------------------
Копіює маппінг категорій між xlsx-файлами за числовим ключем.

1. mappings_old.xlsx → mappings.xlsx          (лист «Категорія+»)
   Ключ: «ІD категорії фіду»
   Поля: Приналежність*:6 | Група*:13 | Підгрупа*:14 | Вид*:21 | Сезонність*:5

2. epicenter_mappings_old.xlsx → epicenter_mappings.xlsx  (лист «Маппінг»)
   Ключ: prom_category_id
   Поля: epicenter_category_id | Назва категорії Епіцентру | parentCode

Запуск:
    python scripts/products_copy_categories_old.py
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

# ─── Config ───────────────────────────────────────────────────────────────────

DATA_DIR = Path(__file__).parents[1] / "data" / "markets"


@dataclass(frozen=True)
class MappingTask:
    label: str
    source: Path
    target: Path
    sheet: str
    col_id: int    # 1-based, ключова колонка
    col_first: int # 1-based, перша колонка даних
    col_last: int  # 1-based, остання колонка даних


TASKS: list[MappingTask] = [
    MappingTask(
        label="Kasta (mappings)",
        source=DATA_DIR / "mappings_old.xlsx",
        target=DATA_DIR / "mappings.xlsx",
        sheet="Категорія+",
        col_id=1,
        col_first=3,
        col_last=7,
    ),
    MappingTask(
        label="Epicenter (epicenter_mappings)",
        source=DATA_DIR / "epicenter_mappings_old.xlsx",
        target=DATA_DIR / "epicenter_mappings.xlsx",
        sheet="Маппінг",
        col_id=1,
        col_first=3,
        col_last=5,
    ),
]

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger(__name__)


# ─── Service ──────────────────────────────────────────────────────────────────

def _normalize_id(value: object) -> str | None:
    """Приводить ID до рядка для безпечного порівняння int/str варіантів."""
    if value is None:
        return None
    normalized = str(value).strip()
    # Видаляємо .0 якщо openpyxl зчитав int як float (напр. '513.0' → '513')
    if normalized.endswith(".0") and normalized[:-2].lstrip("-").isdigit():
        normalized = normalized[:-2]
    return normalized or None


def _load_source_mapping(task: MappingTask) -> dict[str, tuple] | None:
    """
    Зчитує SOURCE і повертає {category_id: (col_first..col_last)}.
    Повертає None якщо файл не знайдено.
    """
    if not task.source.exists():
        log.warning("⚠️  [%s] SOURCE не знайдено: %s", task.label, task.source)
        return None

    wb = load_workbook(task.source, read_only=True, data_only=True)
    ws = wb[task.sheet]

    result: dict[str, tuple] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = _normalize_id(row[task.col_id - 1])
        if not key:
            continue
        values = tuple(row[ci - 1] for ci in range(task.col_first, task.col_last + 1))
        if any(v is not None and str(v).strip() for v in values):
            result[key] = values

    wb.close()
    log.info("   📂 Source: %d заповнених маппінгів зчитано", len(result))
    return result


def _apply_mapping(task: MappingTask, mapping: dict[str, tuple]) -> tuple[int, int]:
    """
    Записує значення з mapping у TARGET.
    Повертає (updated, skipped).
    Повертає (-1, -1) якщо TARGET не знайдено.
    """
    if not task.target.exists():
        log.warning("⚠️  [%s] TARGET не знайдено: %s", task.label, task.target)
        return -1, -1

    wb = load_workbook(task.target)
    ws = wb[task.sheet]

    updated = skipped = 0
    for row in ws.iter_rows(min_row=2):
        key = _normalize_id(row[task.col_id - 1].value)
        if not key:
            continue
        values = mapping.get(key)
        if values is None:
            skipped += 1
            continue
        for ci, val in enumerate(values, task.col_first):
            row[ci - 1].value = val
        updated += 1

    wb.save(task.target)
    wb.close()
    return updated, skipped


def _run_task(task: MappingTask) -> None:
    log.info("\n─── %s ─────────────────────────────", task.label)

    mapping = _load_source_mapping(task)
    if mapping is None:
        return

    if not mapping:
        log.warning("⚠️  SOURCE не містить жодного заповненого маппінгу — пропущено.")
        return

    updated, skipped = _apply_mapping(task, mapping)
    if updated == -1:
        return

    log.info("   ✅ TARGET: %s", task.target)
    log.info("   ✔  оновлено рядків   : %d", updated)
    log.info("   ⚠  не знайдено в old : %d", skipped)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    log.info("🚀 kasta_map_categories.py — копіювання маппінгу категорій")
    for task in TASKS:
        _run_task(task)
    log.info("\n🏁 Виконано.")


if __name__ == "__main__":
    main()
