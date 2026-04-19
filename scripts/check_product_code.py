# -*- coding: utf-8 -*-
"""
check_product_code.py
---------------------
Перевіряє стовпець `Код_товару` у файлі export-products.xlsx:
  - Кожен реальний код повинен зустрічатися рівно 2 рази.
  - Усі коди від мін до макс повинні бути присутні (без пропусків).

Sentinel-коди (наприклад 777777) — технічні значення,
вони виключаються з усіх перевірок і виводяться окремо.
"""

from __future__ import annotations

import logging
import sys
from collections import Counter
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Конфігурація
# ---------------------------------------------------------------------------
EXCEL_PATH     = Path(r"C:\FullStack\PriceFeedPipeline\data\export-products.xlsx")
COLUMN_NAME    = "Код_товару"
EXPECTED_COUNT = 2

# Sentinel-значення — технічні коди, не є реальними товарними кодами.
# Виключаються з перевірки послідовності та підрахунку повторів.
SENTINEL_CODES: frozenset[int] = frozenset({777777})

# ---------------------------------------------------------------------------
# Логування
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Допоміжні функції
# ---------------------------------------------------------------------------

def load_column_values(path: Path, column_name: str) -> list[int]:
    """Зчитує всі значення з цільового стовпця; повертає список цілих чисел."""
    if not path.exists():
        log.error("Файл не знайдено: %s", path)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    if column_name not in headers:
        log.error(
            "Стовпець '%s' не знайдено. Доступні стовпці: %s",
            column_name,
            headers,
        )
        wb.close()
        sys.exit(1)

    col_index: int = headers.index(column_name)

    values: list[int] = []
    skipped = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = row[col_index]
        if raw is None:
            skipped += 1
            continue
        try:
            values.append(int(raw))
        except (ValueError, TypeError):
            log.warning("Рядок %d: не вдалося перетворити '%s' на ціле число — пропущено.", row_idx, raw)
            skipped += 1

    wb.close()

    if skipped:
        log.info("Пропущено рядків з відсутніми або некоректними значеннями: %d", skipped)

    return values


def check_duplicates(counts: Counter[int]) -> list[tuple[int, int]]:
    """Повертає реальні коди, кількість входжень яких відрізняється від очікуваної."""
    return [
        (code, cnt)
        for code, cnt in sorted(counts.items())
        if cnt != EXPECTED_COUNT and code not in SENTINEL_CODES
    ]


def check_gaps(codes: set[int]) -> list[int]:
    """Повертає відсутні цілі числа в діапазоні [мін, макс] без sentinel-кодів."""
    real_codes = codes - SENTINEL_CODES
    if not real_codes:
        return []
    full_range = range(min(real_codes), max(real_codes) + 1)
    return [n for n in full_range if n not in real_codes]


# ---------------------------------------------------------------------------
# Головна функція
# ---------------------------------------------------------------------------

def main() -> None:
    log.info("Завантаження файлу: %s", EXCEL_PATH)
    values = load_column_values(EXCEL_PATH, COLUMN_NAME)

    if not values:
        log.error("У стовпці '%s' не знайдено жодного валідного значення. Завершення.", COLUMN_NAME)
        sys.exit(1)

    log.info("Всього завантажено значень: %d", len(values))

    counts: Counter[int] = Counter(values)
    all_codes    = set(counts.keys())
    real_codes   = all_codes - SENTINEL_CODES
    sentinel_found = all_codes & SENTINEL_CODES

    # --- sentinel-коди ---
    if sentinel_found:
        for sc in sorted(sentinel_found):
            log.info(
                "Sentinel-код %d знайдено: %d рядків — виключено з перевірок.",
                sc, counts[sc],
            )

    log.info("Унікальних реальних кодів: %d", len(real_codes))
    log.info("Очікувана кількість входжень на код: %d", EXPECTED_COUNT)

    # ------------------------------------------------------------------ #
    # 1. Перевірка повторів (тільки реальні коди)
    # ------------------------------------------------------------------ #
    violations  = check_duplicates(counts)
    over_limit  = [(code, cnt) for code, cnt in violations if cnt > EXPECTED_COUNT]
    under_limit = [(code, cnt) for code, cnt in violations if cnt < EXPECTED_COUNT]

    if over_limit:
        log.warning("--- КОДИ З КІЛЬКІСТЮ ВХОДЖЕНЬ БІЛЬШЕ %d ---", EXPECTED_COUNT)
        for code, cnt in over_limit:
            log.warning("  Код %s — %d входжень (надлишок: %d)", code, cnt, cnt - EXPECTED_COUNT)
        log.warning("Всього кодів з перевищенням: %d", len(over_limit))
    else:
        log.info("Кодів з кількістю входжень більше %d не виявлено.", EXPECTED_COUNT)

    if under_limit:
        log.warning("--- КОДИ З КІЛЬКІСТЮ ВХОДЖЕНЬ МЕНШЕ %d ---", EXPECTED_COUNT)
        for code, cnt in under_limit:
            log.warning("  Код %s — %d входжень (нестача: %d)", code, cnt, EXPECTED_COUNT - cnt)
        log.warning("Всього кодів з нестачею: %d", len(under_limit))
    else:
        log.info("Кодів з кількістю входжень менше %d не виявлено.", EXPECTED_COUNT)

    # ------------------------------------------------------------------ #
    # 2. Перевірка пропусків у послідовності (тільки реальні коди)
    # ------------------------------------------------------------------ #
    gaps = check_gaps(real_codes)

    if gaps:
        log.warning(
            "--- ВІДСУТНІ КОДИ В ПОСЛІДОВНОСТІ [%d … %d] ---",
            min(real_codes), max(real_codes),
        )
        for missing in gaps:
            log.warning("  Відсутній код: %s", missing)
        log.warning("Всього відсутніх кодів: %d", len(gaps))
    else:
        log.info(
            "Послідовність повна — пропусків у діапазоні [%d … %d] не виявлено.",
            min(real_codes), max(real_codes),
        )

    # ------------------------------------------------------------------ #
    # 3. Підсумок
    # ------------------------------------------------------------------ #
    log.info("=== ПІДСУМОК ===")
    log.info("  Всього оброблено рядків    : %d", len(values))
    log.info("  Реальних унікальних кодів  : %d", len(real_codes))
    log.info("  Sentinel-рядків (%s)   : %d", "/".join(str(s) for s in sorted(sentinel_found)), counts[777777] if 777777 in counts else 0)
    log.info("  Кодів з перевищенням       : %d", len(over_limit))
    log.info("  Кодів з нестачею           : %d", len(under_limit))
    log.info("  Відсутніх у послідовності  : %d", len(gaps))

    if not violations and not gaps:
        log.info("✅ Всі перевірки пройдено успішно.")
    else:
        log.warning("⚠️  Виявлено проблеми — дивіться деталі вище.")


if __name__ == "__main__":
    main()
