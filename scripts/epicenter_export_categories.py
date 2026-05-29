"""
epicenter_export_categories.py
-------------------------------
КРОК 1 з 5 пайплайну маппінгу Prom → Epicenter.

Що робить:
  Створює epicenter_mappings.xlsx (тільки якщо файл ще не існує):
    • Лист «Інструкція»          — покрокова інструкція пайплайну
    • Лист «Маппінг»             — заголовки для маппінгу (дані заповнює prom_export_categories.py)
    • Лист «Категорії Епіцентру» — повний довідник категорій Epicenter з API

Якщо файл вже існує, але лист «Категорії Епіцентру» відсутній —
  завантажує категорії з API та додає лист у існуючий файл.

Якщо файл вже існує і лист є — нічого не робить
  (використай --force щоб примусово перезаписати лист).

Наступний крок після першого запуску:
  Заповни лист «Маппінг» (prom_category_id, Категорія Прому):
    python scripts/prom_export_categories.py
  Потім заповни epicenter_category_id одним зі способів:
    • Автоматично: python scripts/epicenter_map_categories.py
    • Вручну:      колонки C, D, E у листі «Маппінг»
  Потім → python scripts/epicenter_export_attr_sets.py

Запуск:
    python scripts/epicenter_export_categories.py
    python scripts/epicenter_export_categories.py --force   # перезаписати лист
"""

from __future__ import annotations

import sys
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ─── Config ───────────────────────────────────────────────────────────────────

API_TOKEN = "5a6489d1a5c48c9d174bd31f2a0a8fd0"
BASE_URL  = "https://api.epicentrm.com.ua/v2/pim"
HEADERS   = {
    "Authorization": f"Bearer {API_TOKEN}",
    "Content-Type": "application/json",
    "Accept": "application/json",
}

OUTPUT_PATH = Path(__file__).parents[1] / "data" / "markets" / "epicenter_mappings.xlsx"
REQ_TIMEOUT = (10, 30)

CATEGORIES_SHEET = "Категорії Епіцентру"

# Заголовки та ширини колонок листа «Маппінг»
MAPPING_COLUMNS: list[tuple[str, int]] = [
    ("prom_category_id",          22),
    ("Категорія Прому",           55),
    ("epicenter_category_id",     25),
    ("Назва категорії Епіцентру", 45),
    ("parentCode",                20),
    ("Коментар / Примітка",       35),
]

# Заголовки та ширини колонок листа «Категорії Епіцентру»
EPICENTER_COLUMNS: list[tuple[str, int]] = [
    ("code",       30),
    ("name_uk",    50),
    ("parentCode", 30),
    ("hasChild",   12),
    ("deleted",    12),  # E — True = мертва категорія, не використовувати в маппінгу
]


# ─── Styles ───────────────────────────────────────────────────────────────────

_HDR_FILL    = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
_HDR_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
_YELLOW_FILL = PatternFill("solid", start_color="FFFF99", end_color="FFFF99")
_RED_FILL    = PatternFill("solid", start_color="FFB3B3", end_color="FFB3B3")  # deleted=True
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _style_header(cell, fill=_HDR_FILL) -> None:
    cell.font      = _HDR_FONT
    cell.fill      = fill
    cell.alignment = _CENTER
    cell.border    = _THIN_BORDER


def _style_data(cell, fill=None) -> None:
    cell.font      = Font(name="Calibri", size=14)
    cell.alignment = _LEFT
    cell.border    = _THIN_BORDER
    if fill:
        cell.fill = fill


# ─── API ──────────────────────────────────────────────────────────────────────

def _make_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(total=2, backoff_factor=0.5, status_forcelist=(429,), allowed_methods=["GET"])
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://",  adapter)
    session.headers.update(HEADERS)
    return session


def fetch_epicenter_categories() -> list[dict]:
    """Завантажує всі категорії Epicenter з API (з пагінацією)."""
    print("⬇️  Категорії Епіцентру (всі)...")
    session = _make_session()
    items: list[dict] = []
    page = 1

    while True:
        try:
            data = session.get(
                f"{BASE_URL}/categories", params={"page": page}, timeout=REQ_TIMEOUT
            ).json()
        except Exception as e:
            print(f"❌ categories p{page}: {e}")
            break

        batch = data.get("items", [])
        if not batch:
            break

        items.extend(batch)
        total = data.get("pages", 1)
        print(f"   {page}/{total}: {len(batch)} категорій")

        if page >= total:
            break
        page += 1

    deleted_count = sum(1 for c in items if c.get("deleted"))
    print(f"✅ Категорій всього: {len(items)}  (з них deleted=True: {deleted_count})")
    return items


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _get_translation(translations: list[dict], lang: str = "ua") -> str:
    """Повертає переклад за пріоритетом мов."""
    for priority_lang in (lang, "ua", "uk", "ru", "en"):
        for t in translations:
            if t.get("languageCode") == priority_lang:
                val = t.get("value") or t.get("title") or ""
                if str(val).strip():
                    return str(val).strip()
    return ""


# ─── Sheet builders ───────────────────────────────────────────────────────────

def _build_instructions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Інструкція", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 100

    lines: list[tuple[str, str]] = [
        ("title", "📋  ІНСТРУКЦІЯ З МАППІНГУ КАТЕГОРІЙ ПРОМУ → ЕПІЦЕНТР"),
        ("",      ""),
        ("step",  "КРОК 1 — Генерація файлу маппінгу"),
        ("body",  "   Запуск: python scripts/epicenter_export_categories.py"),
        ("body",  "   Результат: створюється файл epicenter_mappings.xlsx з трьома листами."),
        ("body",  "   Лист «Категорії Епіцентру» — заповнюється автоматично з API."),
        ("body",  "   ⚠️  Рядки підсвічені червоним = deleted=True (мертві категорії, не використовувати)."),
        ("body",  "   Лист «Маппінг» — містить тільки заголовки (дані додає наступний крок)."),
        ("body",  "   Запуск: python scripts/prom_export_categories.py"),
        ("body",  "   Результат: лист «Маппінг» заповнюється колонками prom_category_id та Категорія Прому."),
        ("",      ""),
        ("step",  "КРОК 2 — Заповнення маппінгу категорій"),
        ("body",  "   Варіант А (автоматично): python scripts/epicenter_map_categories.py"),
        ("body",  "   Варіант Б (вручну): у «Маппінг» заповни жовті колонки:"),
        ("body",  "     C — epicenter_category_id  (code з листа «Категорії Епіцентру»)"),
        ("body",  "     D — Назва категорії Епіцентру"),
        ("body",  "     E — parentCode"),
        ("body",  "   ⚠️  epicenter_category_id = code = set_code — одне й те саме число!"),
        ("",      ""),
        ("step",  "КРОК 3 — Завантаження сетів атрибутів"),
        ("body",  "   Запуск: python scripts/epicenter_export_attr_sets.py"),
        ("body",  "   Результат: лист «Сети атрибутів» — атрибути тільки для заповнених категорій."),
        ("",      ""),
        ("step",  "КРОК 4 — Заповнення prom_param_name"),
        ("body",  "   Варіант А (автоматично): python scripts/epicenter_map_attributes.py"),
        ("body",  "   Варіант Б (вручну): у «Сети атрибутів» заповни колонку J (prom_param_name)."),
        ("body",  "   🔴 Червоні клітинки — isRequired=True, обов'язково заповнити!"),
        ("",      ""),
        ("step",  "КРОК 5 — Завантаження опцій атрибутів"),
        ("body",  "   Запуск: python scripts/epicenter_export_attr_options.py"),
        ("body",  "   Результат: лист «Опції атрибутів» — тільки для атрибутів з заповненим prom_param_name."),
        ("",      ""),
        ("step",  "КРОК 6 — Заповнення prom_option_name"),
        ("body",  "   Заповни колонку H (prom_option_name) вручну або напиши скрипт маппінгу."),
        ("body",  "   option_code (F) підставляється у XML Епіцентру."),
        ("",      ""),
        ("warn",  f"   API: https://api.epicentrm.com.ua/swagger/ | Токен: {API_TOKEN}"),
    ]

    for ri, (kind, text) in enumerate(lines, 1):
        ws.row_dimensions[ri].height = 18
        cell = ws.cell(row=ri, column=2, value=text)
        if kind == "title":
            cell.font = Font(bold=True, size=14, color="1F4E79", name="Calibri")
        elif kind == "step":
            cell.font = Font(bold=True, size=14, color="2E75B6", name="Calibri")
        elif kind == "warn":
            cell.font = Font(bold=True, size=14, color="C00000", name="Calibri")
        else:
            cell.font = Font(size=10, name="Calibri")
        cell.alignment = _LEFT


def _build_mapping_sheet(wb: Workbook) -> None:
    """Створює лист «Маппінг» тільки із заголовками. Дані заповнює prom_export_categories.py."""
    ws = wb.create_sheet("Маппінг")

    for ci, (header, width) in enumerate(MAPPING_COLUMNS, 1):
        _style_header(ws.cell(row=1, column=ci, value=header))
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 30

    # Підказки у колонці H та I (поза основною таблицею, обидві в row=1)
    hint_col = len(MAPPING_COLUMNS) + 2
    ws.cell(row=1, column=hint_col, value="🟡 C — epicenter_category_id — заповнити (крок 2)").font = (
        Font(bold=True, color="7F6000", name="Calibri", size=9)
    )
    ws.cell(row=1, column=hint_col + 1, value="🟢 A, B — заповнює prom_export_categories.py").font = (
        Font(bold=True, color="375623", name="Calibri", size=9)
    )

    ws.freeze_panes = "A2"


def _build_categories_sheet(wb: Workbook, categories: list[dict]) -> None:
    """Створює лист «Категорії Епіцентру» з даними з API."""
    ws = wb.create_sheet(CATEGORIES_SHEET)

    for ci, (header, width) in enumerate(EPICENTER_COLUMNS, 1):
        _style_header(ws.cell(row=1, column=ci, value=header))
        ws.column_dimensions[get_column_letter(ci)].width = width

    for ri, cat in enumerate(categories, 2):
        deleted = cat.get("deleted", False)
        row_fill = _RED_FILL if deleted else None

        values = [
            cat.get("code", ""),
            _get_translation(cat.get("translations", [])),
            cat.get("parentCode", ""),
            cat.get("hasChild", ""),
            deleted,
        ]
        for ci, val in enumerate(values, 1):
            _style_data(ws.cell(row=ri, column=ci, value=val), fill=row_fill)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EPICENTER_COLUMNS))}{len(categories) + 1}"


def _build_empty_categories_sheet(wb: Workbook) -> None:
    """Запасний варіант: якщо API не відповів — лист з попередженням."""
    ws = wb.create_sheet(CATEGORIES_SHEET)
    ws["A1"] = "⚠️ Не завантажено. Перевір токен API та повтори запуск."
    ws["A1"].font = Font(bold=True, color="C00000", name="Calibri")


def _repopulate_categories_sheet(wb: Workbook, categories: list[dict]) -> None:
    """
    Видаляє існуючий лист «Категорії Епіцентру» (якщо є) та створює заново.
    Використовується при відновленні відсутнього листа або при --force.
    """
    if CATEGORIES_SHEET in wb.sheetnames:
        del wb[CATEGORIES_SHEET]

    if categories:
        _build_categories_sheet(wb, categories)
    else:
        _build_empty_categories_sheet(wb)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    force = "--force" in sys.argv
    print("🚀 epicenter_export_categories.py — КРОК 1\n")

    # ── Сценарій 1: файл не існує → створити з нуля ──────────────────────────
    if not OUTPUT_PATH.exists():
        categories = fetch_epicenter_categories()

        wb = Workbook()
        wb.remove(wb.active)

        _build_instructions_sheet(wb)
        _build_mapping_sheet(wb)

        if categories:
            _build_categories_sheet(wb, categories)
        else:
            _build_empty_categories_sheet(wb)

        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        wb.save(OUTPUT_PATH)

        print(f"\n✅ Збережено: {OUTPUT_PATH}")
        print(f"   Листи: {wb.sheetnames}")
        print(
            "\n📌 Наступні кроки:\n"
            "   1. Заповни «Маппінг» (prom_category_id, Категорія Прому):\n"
            "      python scripts/prom_export_categories.py\n"
            "   2. Заповни epicenter_category_id (крок 2):\n"
            "      • Автоматично: python scripts/epicenter_map_categories.py\n"
            "      • Вручну:      колонки C, D, E у листі «Маппінг»\n"
            "   3. Далі → python scripts/epicenter_export_attr_sets.py"
        )
        return

    # ── Сценарій 2: файл є → перевірити наявність листа ─────────────────────
    wb = load_workbook(OUTPUT_PATH)

    if CATEGORIES_SHEET in wb.sheetnames and not force:
        # Сценарій 2а: лист є, --force не передано → нічого не робити
        print(
            f"ℹ️  Файл вже існує і лист «{CATEGORIES_SHEET}» присутній: {OUTPUT_PATH}\n"
            f"   Щоб перезаписати лист (наприклад, після оновлення схеми),\n"
            f"   запусти з прапором: python scripts/epicenter_export_categories.py --force\n"
            f"\n"
            f"   Подальші кроки:\n"
            f"   1. Заповни «Маппінг»:  python scripts/prom_export_categories.py\n"
            f"   2. Заповни epicenter_category_id: python scripts/epicenter_map_categories.py\n"
            f"   3. Далі → python scripts/epicenter_export_attr_sets.py"
        )
        return

    # Сценарій 2б: лист відсутній АБО --force → докачати і перезаписати
    if force and CATEGORIES_SHEET in wb.sheetnames:
        print(f"🔄 --force: перезаписуємо лист «{CATEGORIES_SHEET}»...\n")
    else:
        print(
            f"⚠️  Файл існує, але лист «{CATEGORIES_SHEET}» відсутній.\n"
            f"   Завантажуємо категорії з API та відновлюємо лист...\n"
        )

    categories = fetch_epicenter_categories()
    _repopulate_categories_sheet(wb, categories)
    wb.save(OUTPUT_PATH)

    print(f"\n✅ Лист «{CATEGORIES_SHEET}» оновлено: {OUTPUT_PATH}")
    print(f"   Листи: {wb.sheetnames}")


if __name__ == "__main__":
    main()
