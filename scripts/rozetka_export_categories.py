"""
rozetka_export_categories.py
----------------------------
Exports active Rozetka marketplace categories to:
  data/markets/rozetka_mappings.xlsx, sheet "Категорії Розетки"

API:
  GET https://api-seller.rozetka.com.ua/market-categories/search

Auth:
  Set one of these environment variables:
    ROZETKA_API_TOKEN
    ROZETKA_TOKEN
    ROZETKA_SELLER_API_TOKEN
  Local runs also load suppliers/.env automatically.

Run:
  python scripts/rozetka_export_categories.py
"""

from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover - optional local convenience
    load_dotenv = None


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT / "data" / "markets" / "rozetka_mappings.xlsx"
TARGET_SHEET = "Категорії Розетки"
MAPPING_SHEET = "Маппінг"

API_BASE_URL = "https://api-seller.rozetka.com.ua"
SEARCH_ENDPOINT = "/market-categories/search"

DEFAULT_LANGUAGE = "uk"
DEFAULT_PAGE_SIZE = 100
REQUEST_TIMEOUT = (10, 60)

TOKEN_ENV_NAMES = (
    "ROZETKA_API_TOKEN",
    "ROZETKA_TOKEN",
    "ROZETKA_SELLER_API_TOKEN",
)

CATEGORY_COLUMNS: list[tuple[str, int]] = [
    ("rozetka_category_id", 22),
    ("Назва категорії Розетки", 55),
    ("parentCode", 18),
    ("Назва батьківської категорії", 45),
    ("level", 10),
    ("is_leaf", 10),
    ("Повний шлях категорії", 90),
    ("Коментар / Примітка", 28),
]

MAPPING_COLUMNS: list[tuple[str, int]] = [
    ("prom_category_id", 20),
    ("Категорія Прому", 55),
    ("rozetka_category_id", 22),
    ("Назва категорії Розетки", 55),
    ("parentCode", 18),
    ("Коментар / Примітка", 28),
]


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(cell) -> None:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def style_data(cell) -> None:
    cell.font = Font(name="Calibri", size=14)
    cell.alignment = LEFT
    cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class RozetkaCategory:
    category_id: str
    name: str
    parent_id: str


# ---------------------------------------------------------------------------
# Env / HTTP
# ---------------------------------------------------------------------------


def load_env_files() -> None:
    if load_dotenv is None:
        return
    for path in (ROOT / ".env", ROOT / "suppliers" / ".env"):
        if path.exists():
            load_dotenv(path, override=False)


def get_api_token(cli_token: str | None = None) -> str:
    if cli_token and cli_token.strip():
        return normalize_token(cli_token)

    load_env_files()
    for name in TOKEN_ENV_NAMES:
        token = os.environ.get(name, "").strip()
        if token:
            return normalize_token(token)

    names = ", ".join(TOKEN_ENV_NAMES)
    raise RuntimeError(f"Rozetka API token not found. Set one of: {names}")


def normalize_token(token: str) -> str:
    token = token.strip()
    if token.lower().startswith("bearer "):
        return token.split(None, 1)[1].strip()
    return token


def make_session(token: str, language: str) -> requests.Session:
    retry = Retry(
        total=3,
        backoff_factor=1.0,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET",),
    )
    adapter = HTTPAdapter(max_retries=retry)

    session = requests.Session()
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update(
        {
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Accept-Validate-Exception": "1",
            "Content-Language": language,
        }
    )
    return session


# ---------------------------------------------------------------------------
# API parsing
# ---------------------------------------------------------------------------


def norm_id(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    if number.is_integer():
        return str(int(number))
    return text


def category_from_payload(payload: dict[str, Any]) -> RozetkaCategory | None:
    category_id = norm_id(payload.get("id") or payload.get("category_id"))
    name = str(payload.get("name") or payload.get("title") or "").strip()
    parent_id = norm_id(
        payload.get("parent_id")
        or payload.get("parentId")
        or payload.get("parentCode")
        or payload.get("parent_code")
    )

    if not category_id:
        return None

    return RozetkaCategory(category_id=category_id, name=name, parent_id=parent_id)


def extract_categories(data: dict[str, Any]) -> list[RozetkaCategory]:
    content = data.get("content")
    if isinstance(content, list):
        raw_items = content
    elif isinstance(content, dict):
        raw_items = (
            content.get("marketCategorys")
            or content.get("marketCategories")
            or content.get("categories")
            or content.get("items")
            or []
        )
    else:
        raw_items = []

    categories: list[RozetkaCategory] = []
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        category = category_from_payload(item)
        if category is not None:
            categories.append(category)
    return categories


def extract_meta(data: dict[str, Any]) -> dict[str, Any]:
    content = data.get("content")
    if isinstance(content, dict):
        meta = content.get("_meta") or content.get("meta") or {}
        if isinstance(meta, dict):
            return meta
    return {}


def fetch_categories(
    session: requests.Session,
    *,
    page_size: int,
    category_id: int | None = None,
    parent_id: int | None = None,
) -> list[RozetkaCategory]:
    url = f"{API_BASE_URL}{SEARCH_ENDPOINT}"
    categories_by_id: dict[str, RozetkaCategory] = {}
    page = 1

    while True:
        params: dict[str, Any] = {
            "page": page,
            "pageSizeLimit": page_size,
        }
        if category_id is not None:
            params["category_id"] = category_id
        if parent_id is not None:
            params["parent_id"] = parent_id

        response = session.get(url, params=params, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        data = response.json()

        if data.get("success") is False:
            errors = data.get("errors") or data.get("error") or data
            raise RuntimeError(f"Rozetka API error on page {page}: {errors}")

        batch = extract_categories(data)
        for category in batch:
            categories_by_id[category.category_id] = category

        meta = extract_meta(data)
        page_count = int(meta.get("pageCount") or page)
        total_count = int(meta.get("totalCount") or len(categories_by_id))
        current_page = int(meta.get("currentPage") or page)

        print(
            f"page {current_page}/{page_count}: "
            f"{len(batch)} categories, total={total_count}"
        )

        if not batch or page >= page_count:
            break
        page += 1

    return sorted(categories_by_id.values(), key=lambda item: int(item.category_id))


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------


def ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)

    workbook = Workbook()
    workbook.remove(workbook.active)
    mapping = workbook.create_sheet(MAPPING_SHEET)
    for column, (header, width) in enumerate(MAPPING_COLUMNS, start=1):
        cell = mapping.cell(row=1, column=column, value=header)
        style_header(cell)
        mapping.column_dimensions[get_column_letter(column)].width = width
    mapping.freeze_panes = "A2"
    return workbook


def replace_sheet(workbook: Workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        index = workbook.sheetnames.index(sheet_name)
        del workbook[sheet_name]
    else:
        index = len(workbook.sheetnames)
    return workbook.create_sheet(sheet_name, index)


def build_category_rows(categories: list[RozetkaCategory]) -> list[list[Any]]:
    by_id = {category.category_id: category for category in categories}
    child_parent_ids = {category.parent_id for category in categories if category.parent_id}

    def build_path(category: RozetkaCategory) -> tuple[int, str]:
        names: list[str] = []
        current = category
        seen: set[str] = set()

        while current and current.category_id not in seen:
            seen.add(current.category_id)
            names.append(current.name or current.category_id)
            current = by_id.get(current.parent_id)

        names.reverse()
        return max(len(names) - 1, 0), " > ".join(names)

    rows: list[list[Any]] = []
    for category in categories:
        parent = by_id.get(category.parent_id)
        level, path = build_path(category)
        rows.append(
            [
                int(category.category_id),
                category.name,
                int(category.parent_id) if category.parent_id else None,
                parent.name if parent else "",
                level,
                category.category_id not in child_parent_ids,
                path,
                "",
            ]
        )

    return rows


def write_categories_sheet(workbook: Workbook, categories: list[RozetkaCategory]) -> None:
    worksheet = replace_sheet(workbook, TARGET_SHEET)
    worksheet.sheet_view.showGridLines = False

    for column, (header, width) in enumerate(CATEGORY_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        style_header(cell)
        worksheet.column_dimensions[get_column_letter(column)].width = width

    rows = build_category_rows(categories)
    for row_index, row in enumerate(rows, start=2):
        for column, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            style_data(cell)

    worksheet.freeze_panes = "A2"
    if rows:
        last_col = get_column_letter(len(CATEGORY_COLUMNS))
        worksheet.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export active Rozetka categories to rozetka_mappings.xlsx"
    )
    parser.add_argument("--token", help="Rozetka Bearer token; env is preferred")
    parser.add_argument(
        "--language",
        default=DEFAULT_LANGUAGE,
        choices=("uk", "ru"),
        help="Content-Language request header",
    )
    parser.add_argument(
        "--page-size",
        type=int,
        default=DEFAULT_PAGE_SIZE,
        help="API pageSizeLimit",
    )
    parser.add_argument("--category-id", type=int, help="Optional category_id filter")
    parser.add_argument("--parent-id", type=int, help="Optional parent_id filter")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Fetch categories but do not write workbook",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv or sys.argv[1:])
    token = get_api_token(args.token)
    session = make_session(token, args.language)

    print("Rozetka categories export")
    print(f"target: {OUTPUT_PATH}")
    print(f"language: {args.language}, page_size: {args.page_size}")

    categories = fetch_categories(
        session,
        page_size=args.page_size,
        category_id=args.category_id,
        parent_id=args.parent_id,
    )

    print(f"loaded categories: {len(categories)}")
    if args.dry_run:
        print("dry-run: workbook not changed")
        return

    workbook = ensure_workbook(OUTPUT_PATH)
    write_categories_sheet(workbook, categories)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    workbook.close()
    print(f"saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
