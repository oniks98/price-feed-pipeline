"""
epicenter_map_attr_options.py
------------------------------
КРОК 6 пайплайну маппінгу Prom → Epicenter.

Заповнює лист «Опції атрибутів» в epicenter_mappings.xlsx.

══════════════════════════════════════════════════════════════════════
4 ПРАВИЛА ОБРОБКИ
══════════════════════════════════════════════════════════════════════

Rule 1 — float/array/text + needs_default=TRUE
   → option_name_uk = hardcoded дефолт (мм / г / шт.)
     Кратність=1, Висота/Ширина/Глибина/Довжина=150, Вага=500, …
     Fallback якщо attr_name не відомий: 150

Rule 2 — float/array/text + needs_default=FALSE
   → option_name_uk = медіана всіх значень із Prom-фіду по prom_param_name
     • "Розміри 47,2 х 49,2 х 45 см":
         Ширина  → 1-ша цифра × 10 (см→мм) = 472
         Висота  → 2-га цифра × 10          = 492
         Глибина/Довжина → 3-тя цифра × 10  = 450
     • "Вага"   → кг × 1000 (кг→г), г залишаємо як є
     • fallback → hardcoded або 150

Rule 3 — select/multiselect (needs_default не впливає на алгоритм)
   → prom_option_name = fuzzy-match Epicenter option_name_uk ↔ Prom feed values
     (для кожного рядка, незалежно від needs_default)
   → default_option_code = евристика (один код на весь attr_code, fallback для
     товарів без параметра у фіді):
     a) Hardcode:
        "Одиниця виміру / Міра виміру" → measure_pcs
        "Колір / Колір виробника / Базовий колір" → option_code з option_name_uk="білий"
     b) Пріоритетні назви: no > ні > немає > універсальний > середній >
        комбінований > стандартний > загальний
     c) Перша опція у листі для цього attr_code
   → Записує ОДНАКОВИЙ default_option_code у ВСІ рядки attr_code

   needs_default=TRUE  → дефолт обовʼязковий (товар може не мати параметра)
   needs_default=FALSE → дефолт як fallback (параметр очікується у більшості товарів)

══════════════════════════════════════════════════════════════════════
Ідемпотентність: вже заповнені клітинки НЕ перезаписуються.

Запуск:
    python scripts/epicenter_map_attr_options.py            # prod
    python scripts/epicenter_map_attr_options.py --dry-run  # preview без збереження
    python scripts/epicenter_map_attr_options.py --verbose  # детальний лог

Попередній крок: КРОК 5 — epicenter_export_attr_options.py
"""

from __future__ import annotations

import argparse
import logging
import re
import statistics
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Final

import openpyxl
import requests

# ---------------------------------------------------------------------------
# Guard: script must run from repo root or scripts/ to resolve imports
# ---------------------------------------------------------------------------
_HERE = Path(__file__).parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

from constants_feed_url import FEED_URL_PROM as _FEED_URL  # noqa: E402

# ═══════════════════════════════════════════════════════════════════════════
# Logging
# ═══════════════════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════
# Paths & constants
# ═══════════════════════════════════════════════════════════════════════════

ROOT      : Final[Path] = Path(__file__).parents[1]
XLSX_PATH : Final[Path] = ROOT / "data" / "markets" / "epicenter_mappings.xlsx"
SHEET     : Final[str]  = "Опції атрибутів"

# ── Column header names (must match HEADERS_OPTIONS in epicenter_export_attr_options.py) ──
_C_ATTR_CODE    : Final[str] = "attr_code"
_C_ATTR_NAME    : Final[str] = "attr_name_uk"
_C_ATTR_TYPE    : Final[str] = "attr_type"
_C_OPTION_CODE  : Final[str] = "option_code"
_C_OPTION_NAME  : Final[str] = "option_name_uk"
_C_PROM_OPTION  : Final[str] = "prom_option_name"
_C_NEEDS_DEF    : Final[str] = "needs_default"
_C_DEFAULT_CODE : Final[str] = "default_option_code"
_C_SET_CODES    : Final[str] = "set_codes"
_C_PROM_PARAM   : Final[str] = "prom_param_name"

# ── Attribute type groups ──
_SELECT_TYPES  : Final[frozenset[str]] = frozenset({"select", "multiselect"})
_NUMERIC_TYPES : Final[frozenset[str]] = frozenset({"float", "int", "text", "string", "array"})

# ── Feed request timeout ──
_FEED_TIMEOUT: Final[tuple[int, int]] = (15, 120)


# ═══════════════════════════════════════════════════════════════════════════
# Rule 1 & 2: hardcoded numeric defaults (target units: мм / г / шт.)
# ═══════════════════════════════════════════════════════════════════════════

_NUMERIC_DEFAULTS: Final[dict[str, str]] = {
    # Кратність / штуки
    "мінімальна кратність товару" : "1",
    "кратність"                   : "1",
    "кількість в упаковці"        : "1",
    # Габарити → мм
    "висота"                      : "150",
    "ширина"                      : "150",
    "глибина"                     : "150",
    "довжина"                     : "150",
    "діаметр"                     : "100",
    "товщина"                     : "20",
    "радіус"                      : "50",
    "висота упаковки"             : "200",
    "ширина упаковки"             : "200",
    "глибина упаковки"            : "200",
    "довжина упаковки"            : "200",
    # Вага → г
    "вага"                        : "500",
    "вага нетто"                  : "500",
    "вага брутто"                 : "600",
    "вага з упаковкою"            : "600",
    # Об'єм / ємність → л
    "об'єм"                       : "5",
    "ємність"                     : "5",
    "об'єм бака"                  : "50",
    # Електрика
    "потужність"                  : "1000",
    "напруга"                     : "220",
    "струм"                       : "10",
    "частота"                     : "50",
    "споживана потужність"        : "1000",
    # Температура / тиск
    "температура"                 : "20",
    "тиск"                        : "1",
    # Швидкість
    "швидкість"                   : "100",
}

_NUMERIC_FALLBACK: Final[str] = "150"


# ═══════════════════════════════════════════════════════════════════════════
# Rule 3 & 4: select hardcodes
# "by": "code" → шукаємо точний option_code
# "by": "name" → шукаємо option_code за normalized option_name_uk
# ═══════════════════════════════════════════════════════════════════════════

_SELECT_HARDCODES: Final[dict[str, dict[str, str]]] = {
    # Одиниці виміру — завжди шт.
    "одиниця виміру та кількість"  : {"by": "code", "value": "measure_pcs"},
    "одиниця виміру"               : {"by": "code", "value": "measure_pcs"},
    "міра виміру"                  : {"by": "code", "value": "measure_pcs"},
    "unit of measure"              : {"by": "code", "value": "measure_pcs"},
    # Кольори — за замовчуванням білий
    "колір"                        : {"by": "name", "value": "білий"},
    "колір виробника"              : {"by": "name", "value": "білий"},
    "базовий колір"                : {"by": "name", "value": "білий"},
    "основний колір"               : {"by": "name", "value": "білий"},
    "колір корпусу"                : {"by": "name", "value": "білий"},
    "колір виробу"                 : {"by": "name", "value": "білий"},
    # ── NEW: Бренд — дефолт Anker (option_code uxcbmw0sgjsa8sbv) ──────────
    "бренд"                        : {"by": "code", "value": "uxcbmw0sgjsa8sbv"},
    # ── NEW: Країна виробника — дефолт Китай (option_code chn) ────────────
    "країна-виробник"              : {"by": "code", "value": "chn"},
    "країна виробник"              : {"by": "code", "value": "chn"},
    "країна виробника"             : {"by": "code", "value": "chn"},
    "країна походження"            : {"by": "code", "value": "chn"},
    "країна"                       : {"by": "code", "value": "chn"},
}

# Пріоритетний порядок option_name_uk при виборі дефолту
# (від найвищого до найнижчого пріоритету)
_PRIORITY_OPTION_NAMES: Final[list[str]] = [
    # Заперечення / відсутність → завжди якщо є пара no/yes
    "no", "ні", "нет", "немає", "немає (no)", "відсутній", "відсутня", "відсутнє",
    # Загальні / нейтральні
    "загальний", "загальна", "загальне",
    "стандартний", "стандартна", "стандартне",
    "звичайний", "звичайна", "звичайне",
    "побутовий", "побутова", "побутове",
    "для дому",
    "класичний", "класична", "класичне",
    "універсальний", "універсальна", "універсальне",
    # Розмір
    "середній", "середня", "середнє",
    # Тип з'єднання / конструкції
    "комбінований", "комбінована", "комбіноване",
    "змішаний", "змішана", "змішане",
    # Матеріал / колір
    "інший", "інша", "інше",
]


# ═══════════════════════════════════════════════════════════════════════════
# Regex
# ═══════════════════════════════════════════════════════════════════════════

_PARAM_RE : Final = re.compile(
    r'<param\b[^>]*\bname="([^"]+)"[^>]*>(.*?)</param>', re.DOTALL
)
_CDATA_RE : Final = re.compile(r'<!\[CDATA\[(.*?)\]\]>', re.DOTALL)

# "47,2 х 49,2 х 45 см"  or  "470 x 492 x 450 мм"
_ROZMIR_RE: Final = re.compile(
    r"(\d+[\.,]\d+|\d+)\s*[хxX×]\s*(\d+[\.,]\d+|\d+)\s*[хxX×]\s*(\d+[\.,]\d+|\d+)",
)
_NUM_RE   : Final = re.compile(r"(\d+[\.,]\d+|\d+)")
_UNIT_RE  : Final = re.compile(r"\b(мм|см|м|кг|г|гр)\b", re.IGNORECASE)


# ═══════════════════════════════════════════════════════════════════════════
# Dimension group mapping: attr_name_lower → regex group (1-based)
# Формат "Розміри": Ш × В × Г (ширина × висота × глибина)
# ═══════════════════════════════════════════════════════════════════════════

_DIM_GROUP: Final[dict[str, int]] = {
    # Ширина = 1-ша цифра
    "ширина"                    : 1,
    "ширина упаковки"           : 1,
    "width"                     : 1,
    # Висота = 2-га цифра
    "висота"                    : 2,
    "висота упаковки"           : 2,
    "height"                    : 2,
    # Глибина / Довжина = 3-тя цифра
    "глибина"                   : 3,
    "глибина упаковки"          : 3,
    "довжина"                   : 3,
    "довжина упаковки"          : 3,
    "depth"                     : 3,
    "length"                    : 3,
}

# ── Unicode dash variants → normalize to plain hyphen before param lookup ──
# Covers: en-dash, em-dash, non-breaking hyphen, minus-sign, etc.
_DASH_RE: Final = re.compile(
    r"[\u2010\u2011\u2012\u2013\u2014\u2015\u2212\uFE58\uFE63\uFF0D]"
)


def _norm_param(s: str) -> str:
    """Normalize prom param name: collapse all unicode dash variants → hyphen, strip, lower."""
    return _DASH_RE.sub("-", s).strip().lower()


# ═══════════════════════════════════════════════════════════════════════════
# Data model
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class OptionRow:
    """In-memory snapshot of one sheet row."""
    row_idx      : int
    attr_code    : str
    attr_name    : str
    attr_type    : str          # select | multiselect | float | int | text | string | array
    option_code  : str          # epicenter option code
    option_name  : str          # option_name_uk — вже заповнено або порожньо
    prom_option  : str          # prom_option_name — вже заповнено або порожньо
    needs_default: bool
    default_code : str          # default_option_code — вже заповнено або порожньо
    prom_params  : list[str]    # розпарсений prom_param_name (кома-розділений)
    set_codes    : str = ""     # set_codes — ключ для grouping дефолтів по набору


@dataclass
class RunStats:
    rule1         : int = 0     # numeric + needs_default=TRUE → option_name_uk (hardcode)
    rule2         : int = 0     # numeric + needs_default=FALSE → option_name_uk (feed median)
    rule3_match   : int = 0     # select → prom_option_name matched (fuzzy)
    rule3_default : int = 0     # select → default_option_code written
    skipped       : int = 0     # already filled → not overwritten
    no_match      : int = 0     # select → fuzzy match not found (prom_option залишено порожнім)

    def report(self) -> str:
        lines = [
            "╔══════════════════════════════════════════════════",
            "║  РЕЗУЛЬТАТ epicenter_map_attr_options.py",
            "╠══════════════════════════════════════════════════",
            f"║  Rule 1  numeric + default (hardcode)  : {self.rule1:>6}",
            f"║  Rule 2  numeric + feed median         : {self.rule2:>6}",
            f"║  Rule 3  select prom_option matched    : {self.rule3_match:>6}",
            f"║  Rule 3  select default_option_code    : {self.rule3_default:>6}",
            f"║  Rule 3  select no match (порожньо)    : {self.no_match:>6}",
            f"║  Пропущено (вже заповнено)             : {self.skipped:>6}",
            "╚══════════════════════════════════════════════════",
        ]
        return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════
# Feed helpers
# ═══════════════════════════════════════════════════════════════════════════

def fetch_feed() -> str:
    """Download Prom XML feed. Raises on HTTP / network error."""
    logger.info("⬇️  Завантажуємо Prom фід…")
    resp = requests.get(_FEED_URL, timeout=_FEED_TIMEOUT)
    resp.raise_for_status()
    # requests авто-детектує Latin-1 коли сервер не повертає charset в Content-Type.
    # XML фід Prom.ua завжди UTF-8 — форсуємо явно.
    xml = resp.content.decode("utf-8")
    logger.info("   Отримано: %d символів", len(xml))
    return xml


def _strip_cdata(raw: str) -> str:
    m = _CDATA_RE.match(raw.strip())
    return m.group(1).strip() if m else raw.strip()


def extract_param_values(xml: str) -> dict[str, list[str]]:
    """
    Parse ALL <param name="...">value</param> from Prom feed.

    Returns {normalized_param_name → unique ordered list of values}.
    Keys are normalized via _norm_param() so that unicode dash variants in the feed XML
    (en-dash, non-breaking hyphen, etc.) are collapsed to a plain hyphen — matching
    the values stored in xlsx prom_param_name column.
    """
    raw: dict[str, dict[str, None]] = {}          # param → {value: None} (ordered dedup)
    for m in _PARAM_RE.finditer(xml):
        name  = _norm_param(m.group(1))   # ← normalize unicode dashes
        value = _strip_cdata(m.group(2))
        if value:
            raw.setdefault(name, {})[value] = None

    return {k: list(v.keys()) for k, v in raw.items()}


# ═══════════════════════════════════════════════════════════════════════════
# Numeric conversion (Rules 1 & 2)
# ═══════════════════════════════════════════════════════════════════════════

def _mm_factor(text: str) -> float:
    """Return multiplier to convert detected unit → mm."""
    m = _UNIT_RE.search(text)
    unit = m.group(1).lower() if m else "см"    # default: assume cm
    return {"мм": 1.0, "см": 10.0, "м": 1000.0}.get(unit, 10.0)


def _g_factor(text: str) -> float:
    """Return multiplier to convert detected unit → grams."""
    m = _UNIT_RE.search(text)
    unit = m.group(1).lower() if m else "г"
    return {"г": 1.0, "гр": 1.0, "кг": 1000.0}.get(unit, 1.0)


def _is_weight_attr(attr_name_lower: str) -> bool:
    return any(w in attr_name_lower for w in ("вага", "маса", "weight", "масса"))


def _extract_from_rozmir(value: str, attr_name_lower: str) -> float | None:
    """
    Extract one dimension from "Розміри" string.

    "47,2 х 49,2 х 45 см" + attr="глибина" → 45 × 10 = 450 мм
    Returns float in mm, or None if can't parse.
    """
    m = _ROZMIR_RE.search(value)
    if not m:
        return None
    grp = _DIM_GROUP.get(attr_name_lower)
    if grp is None:
        return None
    num = float(m.group(grp).replace(",", "."))
    return round(num * _mm_factor(value), 2)


def _extract_single_number(value: str, attr_name_lower: str) -> float | None:
    """
    Extract first numeric value from param string and apply unit conversion.

    Weight attrs → grams. Dimension attrs → mm. Others → raw number.
    """
    m = _NUM_RE.search(value)
    if not m:
        return None
    num = float(m.group(1).replace(",", "."))
    if _is_weight_attr(attr_name_lower):
        return round(num * _g_factor(value), 2)
    if attr_name_lower in _DIM_GROUP:
        return round(num * _mm_factor(value), 2)
    return num


def compute_median_from_feed(
    prom_params : list[str],
    attr_name_lower: str,
    prom_vals   : dict[str, list[str]],
) -> str | None:
    """
    Gather all numeric values for prom_params from feed, compute median.

    Special handling:
      - If prom_param == "Розміри" → use _extract_from_rozmir
      - Otherwise → use _extract_single_number
    Returns rounded int string (mm or g), or None if no data.
    """
    nums: list[float] = []

    for param in prom_params:
        values = prom_vals.get(param, [])
        for raw in values:
            if param == "Розміри":
                n = _extract_from_rozmir(raw, attr_name_lower)
            else:
                n = _extract_single_number(raw, attr_name_lower)
            if n is not None and n > 0:
                nums.append(n)

    if not nums:
        return None

    median = statistics.median(nums)
    result = int(round(median))
    logger.debug(
        "   Median [%s] via %s: %d samples → %d",
        attr_name_lower, prom_params, len(nums), result,
    )
    return str(result)


# ═══════════════════════════════════════════════════════════════════════════
# Select default logic (Rules 3 & 4)
# ═══════════════════════════════════════════════════════════════════════════

def _norm(s: str) -> str:
    """Normalize: lower + collapse whitespace."""
    return " ".join(s.lower().split())


def resolve_default_option_code(
    attr_name   : str,
    option_rows : list[OptionRow],   # all rows with non-empty option_code for this attr
) -> str | None:
    """
    Choose best default_option_code for a select/multiselect attr_code.

    Priority:
      1. Hardcoded attr_name overrides (Колір → білий, Одиниця виміру → measure_pcs)
      2. Priority option names list (no > ні > universal > середній > комбінований …)
      3. First available option_code in sheet order
    """
    attr_lower = _norm(attr_name)

    # ── 1. Hardcoded overrides ────────────────────────────────────────────
    hc = _SELECT_HARDCODES.get(attr_lower)
    if hc:
        if hc["by"] == "code":
            target_code = hc["value"]
            if any(r.option_code == target_code for r in option_rows):
                logger.debug("  [HC-code] %s → %s", attr_name, target_code)
                return target_code
        elif hc["by"] == "name":
            target_name = _norm(hc["value"])
            for r in option_rows:
                if _norm(r.option_name) == target_name:
                    logger.debug("  [HC-name] %s → %s (%s)", attr_name, hc["value"], r.option_code)
                    return r.option_code

    # ── 2. Priority option names ─────────────────────────────────────────
    names_map: dict[str, str] = {
        _norm(r.option_name): r.option_code
        for r in option_rows
        if r.option_code
    }
    for priority in _PRIORITY_OPTION_NAMES:
        if priority in names_map:
            logger.debug("  [priority] %s → '%s' (%s)", attr_name, priority, names_map[priority])
            return names_map[priority]

    # ── 3. First option_code in sheet order ───────────────────────────────
    for r in option_rows:
        if r.option_code:
            logger.debug("  [first] %s → %s", attr_name, r.option_code)
            return r.option_code

    return None


def match_prom_option_name(
    option_name : str,
    prom_values : list[str],
) -> str | None:
    """
    Find matching Prom-feed value for one Epicenter option_name_uk.

    Exact case-insensitive match only — no fuzzy.
    Handles multi-value Prom params (comma-separated: "тварини, коти").
    Returns the matched raw Prom string, or None.
    """
    if not option_name or not prom_values:
        return None

    norm_opt = _norm(option_name)

    for pv in prom_values:
        for part in pv.split(","):
            part = part.strip()
            if part and _norm(part) == norm_opt:
                return part

    return None


# ═══════════════════════════════════════════════════════════════════════════
# XLSX reader
# ═══════════════════════════════════════════════════════════════════════════

def load_sheet_snapshot(ws) -> tuple[dict[str, int], list[OptionRow]]:
    """
    Read header map and all data rows from worksheet into memory.

    Returns:
        hdr  — {column_name: 0-based index}
        rows — list[OptionRow] (sheet order)

    Raises ValueError if required columns are missing.
    """
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    hdr: dict[str, int] = {
        str(v).strip(): i
        for i, v in enumerate(first_row)
        if v is not None
    }

    required_cols = {
        _C_ATTR_CODE, _C_ATTR_NAME, _C_ATTR_TYPE,
        _C_OPTION_CODE, _C_OPTION_NAME, _C_PROM_OPTION,
        _C_NEEDS_DEF, _C_DEFAULT_CODE, _C_PROM_PARAM,
    }
    missing = required_cols - hdr.keys()
    if missing:
        raise ValueError(
            f"Відсутні колонки в «{SHEET}»: {sorted(missing)}\n"
            f"Наявні: {sorted(hdr.keys())}"
        )

    rows: list[OptionRow] = []
    for row_num, vals in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        def _get(col: str) -> str:
            idx = hdr.get(col)
            if idx is None or idx >= len(vals):
                return ""
            v = vals[idx]
            return str(v).strip() if v is not None else ""

        attr_code = _get(_C_ATTR_CODE)
        if not attr_code:
            continue

        prom_param_raw = _get(_C_PROM_PARAM)
        prom_params = [p.strip() for p in prom_param_raw.split(",") if p.strip()]

        needs_raw = _get(_C_NEEDS_DEF).upper()

        rows.append(OptionRow(
            row_idx       = row_num,
            attr_code     = attr_code,
            attr_name     = _get(_C_ATTR_NAME),
            attr_type     = _get(_C_ATTR_TYPE).lower(),
            option_code   = _get(_C_OPTION_CODE),
            option_name   = _get(_C_OPTION_NAME),
            prom_option   = _get(_C_PROM_OPTION),
            needs_default = needs_raw in ("TRUE", "1", "YES"),
            default_code  = _get(_C_DEFAULT_CODE),
            prom_params   = prom_params,
            set_codes     = _get(_C_SET_CODES),
        ))

    return hdr, rows


def _log_sheet_stats(rows: list[OptionRow]) -> None:
    """Print overview of what's in the sheet."""
    by_type: dict[str, int] = defaultdict(int)
    nd_true = nd_false = 0
    for r in rows:
        by_type[r.attr_type] += 1
        if r.needs_default:
            nd_true += 1
        else:
            nd_false += 1

    logger.info(
        "   Типи атрибутів: %s",
        " | ".join(f"{t}={n}" for t, n in sorted(by_type.items())),
    )
    logger.info(
        "   needs_default: TRUE=%d / FALSE=%d", nd_true, nd_false,
    )


# ═══════════════════════════════════════════════════════════════════════════
# Pre-computation helpers (grouping / indexing)
# ═══════════════════════════════════════════════════════════════════════════

def _group_by_attr_code(rows: list[OptionRow]) -> dict[str, list[OptionRow]]:
    """Group rows by attr_code (preserves sheet order within group)."""
    groups: dict[str, list[OptionRow]] = defaultdict(list)
    for r in rows:
        groups[r.attr_code].append(r)
    return groups


def _precompute_defaults(
    groups: dict[str, list[OptionRow]],
) -> dict[tuple[str, str], str | None]:
    """
    For every select/multiselect (attr_code, set_codes) pair, pre-resolve
    default_option_code using only the option rows that belong to that pair.

    Different set_codes within the same attr_code get independent defaults,
    resolved against their own subset of options — so e.g. attr_code=8142
    with set_codes=2569 and set_codes=8241 each pick their own first/priority
    option rather than sharing one global default.

    Returns {(attr_code, set_codes): code_or_None}.
    """
    result: dict[tuple[str, str], str | None] = {}
    for ac, rows in groups.items():
        if rows[0].attr_type not in _SELECT_TYPES:
            continue
        # Group rows within attr_code by set_codes value
        by_set: dict[str, list[OptionRow]] = defaultdict(list)
        for r in rows:
            by_set[r.set_codes].append(r)
        for sc, sc_rows in by_set.items():
            opt_rows = [r for r in sc_rows if r.option_code]
            if not opt_rows:
                result[(ac, sc)] = None
                continue
            result[(ac, sc)] = resolve_default_option_code(sc_rows[0].attr_name, opt_rows)
    return result


# ═══════════════════════════════════════════════════════════════════════════
# Core processor
# ═══════════════════════════════════════════════════════════════════════════

def process(
    ws        : openpyxl.worksheet.worksheet.Worksheet,
    hdr       : dict[str, int],
    rows      : list[OptionRow],
    prom_vals : dict[str, list[str]],
    dry_run   : bool = False,
) -> RunStats:
    """
    Apply 4 rules to every row. Write to worksheet (unless dry_run=True).

    Uses snapshot values for "already filled" checks → fully idempotent.
    """

    # ── Setup ──────────────────────────────────────────────────────────────
    stats  = RunStats()
    groups = _group_by_attr_code(rows)
    defaults = _precompute_defaults(groups)   # attr_code → default_option_code

    # Track which (attr_code, set_codes) pairs have already had default written
    written_defaults: set[tuple[str, str]] = set()

    def _write(row_idx: int, col: str, value: object) -> None:
        if dry_run:
            logger.info("    [DRY] row=%d col=%s value=%r", row_idx, col, value)
            return
        ws.cell(row=row_idx, column=hdr[col] + 1).value = value

    def _write_default_for_attr(ac: str, sc: str, source_rule: int) -> bool:
        """
        Write default_option_code to all rows of (attr_code, set_codes) that
        don't have it yet.  Each (attr_code, set_codes) pair gets its own
        independently resolved default — rows belonging to different set_codes
        within the same attr_code are never mixed.
        Returns True if anything was written.
        """
        key = (ac, sc)
        if key in written_defaults:
            return False

        written_defaults.add(key)
        code = defaults.get(key)

        if not code:
            logger.warning(
                "   ⚠️  Не знайдено default_option_code для attr_code=%s set_codes=%s (%s)",
                ac, sc, groups[ac][0].attr_name if groups.get(ac) else "?",
            )
            return False

        count = 0
        for r2 in groups[ac]:
            if r2.set_codes == sc and not r2.default_code:
                _write(r2.row_idx, _C_DEFAULT_CODE, code)
                count += 1

        if count:
            logger.info(
                "   R%d default_option_code=%s → %d rows [attr=%s / %s / set=%s]",
                source_rule, code, count, ac, groups[ac][0].attr_name, sc,
            )
        return count > 0

    # ── Main loop ──────────────────────────────────────────────────────────
    for row in rows:
        at  = row.attr_type
        nd  = row.needs_default
        ac  = row.attr_code
        ri  = row.row_idx
        an  = row.attr_name
        anl = _norm(an)

        # ════════════════════════════════════════════════════════════════
        # Rule 1 — numeric + needs_default=TRUE → option_name_uk (hardcode)
        # ════════════════════════════════════════════════════════════════
        if at in _NUMERIC_TYPES and nd:
            if row.option_name:
                stats.skipped += 1
                continue
            val = _NUMERIC_DEFAULTS.get(anl, _NUMERIC_FALLBACK)
            _write(ri, _C_OPTION_NAME, val)
            logger.info(
                "R1 [row %4d] %-42s → option_name_uk=%-8s  (hardcode)",
                ri, an, val,
            )
            stats.rule1 += 1

        # ════════════════════════════════════════════════════════════════
        # Rule 2 — numeric + needs_default=FALSE → option_name_uk (feed)
        # ════════════════════════════════════════════════════════════════
        elif at in _NUMERIC_TYPES and not nd:
            if row.option_name:
                stats.skipped += 1
                continue

            val: str
            if row.prom_params:
                computed = compute_median_from_feed(row.prom_params, anl, prom_vals)
                if computed:
                    val = computed
                    logger.info(
                        "R2 [row %4d] %-42s → option_name_uk=%-8s  (feed median, param=%s)",
                        ri, an, val, row.prom_params,
                    )
                else:
                    val = _NUMERIC_DEFAULTS.get(anl, _NUMERIC_FALLBACK)
                    logger.warning(
                        "R2 [row %4d] %-42s → option_name_uk=%-8s  (fallback, no feed data for %s)",
                        ri, an, val, row.prom_params,
                    )
            else:
                val = _NUMERIC_DEFAULTS.get(anl, _NUMERIC_FALLBACK)
                logger.warning(
                    "R2 [row %4d] %-42s → option_name_uk=%-8s  (fallback, no prom_param)",
                    ri, an, val,
                )

            _write(ri, _C_OPTION_NAME, val)
            stats.rule2 += 1

        # ════════════════════════════════════════════════════════════════
        # Rule 3 — select/multiselect (needs_default не впливає на алгоритм)
        #   3a. prom_option_name = exact match option_name_uk ↔ feed
        #   3b. default_option_code = евристика (один на весь attr_code)
        # ════════════════════════════════════════════════════════════════
        elif at in _SELECT_TYPES:

            # 3a. prom_option_name — per-row fuzzy-match із фіду
            if not row.prom_option and row.option_name and row.prom_params:
                candidates: list[str] = []
                for p in row.prom_params:
                    norm_p = _norm_param(p)
                    found  = prom_vals.get(norm_p, [])
                    if not found:
                        logger.debug(
                            "R3 [row %4d] prom_param '%s' (norm='%s') not in feed "
                            "(%d params total). Sample: %s",
                            ri, p, norm_p, len(prom_vals),
                            list(prom_vals.keys())[:8],
                        )
                    candidates.extend(found)

                matched = match_prom_option_name(row.option_name, candidates)
                if matched:
                    _write(ri, _C_PROM_OPTION, matched)
                    logger.info(
                        "R3 [row %4d] %-40s → prom_option_name='%s'",
                        ri, row.option_name, matched,
                    )
                    stats.rule3_match += 1
                else:
                    logger.debug(
                        "R3 [row %4d] %-40s → no match (%d candidates)",
                        ri, row.option_name, len(candidates),
                    )
                    stats.no_match += 1

            # 3b. default_option_code — один для кожної (attr_code, set_codes) пари
            if _write_default_for_attr(ac, row.set_codes, source_rule=3):
                stats.rule3_default += 1

        else:
            # Невідомий тип атрибута — пропускаємо
            logger.debug("   skip [row %d] unknown attr_type=%r", ri, at)
            stats.skipped += 1

    return stats


# ═══════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════

def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Заповнює «Опції атрибутів» в epicenter_mappings.xlsx",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Показати зміни без запису у файл",
    )
    p.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Детальний лог (DEBUG рівень)",
    )
    p.add_argument(
        "--no-feed",
        action="store_true",
        help=(
            "Не завантажувати Prom фід (тільки Rules 1 & 3).\n"
            "Корисно для первинного заповнення дефолтів."
        ),
    )
    return p.parse_args()


# ═══════════════════════════════════════════════════════════════════════════
# Entry point
# ═══════════════════════════════════════════════════════════════════════════

def main() -> None:
    args = _parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    if args.dry_run:
        logger.info("🔍 DRY-RUN режим — файл не буде збережено")

    # ── 1. Validate xlsx ────────────────────────────────────────────────────
    if not XLSX_PATH.exists():
        logger.error(
            "❌ Файл не знайдено: %s\n"
            "   Виконай КРОК 5: python scripts/epicenter_export_attr_options.py",
            XLSX_PATH,
        )
        sys.exit(1)

    logger.info("📂 Відкриваємо: %s", XLSX_PATH)
    wb = openpyxl.load_workbook(XLSX_PATH)

    if SHEET not in wb.sheetnames:
        logger.error(
            "❌ Аркуш «%s» не знайдено. Наявні: %s\n"
            "   Виконай КРОК 5: python scripts/epicenter_export_attr_options.py",
            SHEET, wb.sheetnames,
        )
        sys.exit(1)

    ws = wb[SHEET]
    logger.info("   Рядків (з заголовком): %d", ws.max_row)

    # ── 2. Load snapshot ────────────────────────────────────────────────────
    hdr, rows = load_sheet_snapshot(ws)
    logger.info("   Рядків даних: %d", len(rows))

    if not rows:
        logger.warning("⚠️  Аркуш «%s» порожній — нічого робити.", SHEET)
        return

    _log_sheet_stats(rows)

    # ── 3. Fetch Prom feed ──────────────────────────────────────────────────
    prom_vals: dict[str, list[str]] = {}
    if not args.no_feed:
        try:
            feed_xml  = fetch_feed()
            prom_vals = extract_param_values(feed_xml)
            logger.info("   Унікальних prom_param_name у фіді: %d", len(prom_vals))
        except requests.RequestException as e:
            logger.warning(
                "⚠️  Не вдалося завантажити фід (%s). "
                "Rules 2 & 4 (feed) будуть пропущені, "
                "спрацює fallback до hardcoded дефолтів.",
                e,
            )
    else:
        logger.info("⏭️  --no-feed: пропускаємо завантаження фіду")

    # ── 4. Process ──────────────────────────────────────────────────────────
    logger.info("\n🔧 Обробляємо рядки…\n")
    stats = process(ws, hdr, rows, prom_vals, dry_run=args.dry_run)

    # ── 5. Save ─────────────────────────────────────────────────────────────
    if not args.dry_run:
        wb.save(XLSX_PATH)
        logger.info("💾 Збережено: %s", XLSX_PATH)
    else:
        logger.info("🔍 DRY-RUN: файл не збережено")

    wb.close()

    # ── 6. Report ────────────────────────────────────────────────────────────
    print("\n" + stats.report())


if __name__ == "__main__":
    main()