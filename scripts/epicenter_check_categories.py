"""
epicenter_check_categories.py
------------------------------
Перевіряє лист «Маппінг» у epicenter_mappings.xlsx:

  1. Збирає всі коди з deleted=True з листа «Категорії Епіцентру» (колонка E)
  2. Перевіряє колонку epicenter_category_id у листі «Маппінг»
  3. Підсвічує рядки з мертвими категоріями червоним
  4. Знімає підсвічування з рядків де категорія жива (idempotent, safe to re-run)

Запуск:
    python scripts/epicenter_check_categories.py
    python scripts/epicenter_check_categories.py --dry-run   # тільки звіт, без збереження
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ─── Config ───────────────────────────────────────────────────────────────────

XLSX_PATH        = Path(__file__).parents[1] / "data" / "markets" / "epicenter_mappings.xlsx"
CATEGORIES_SHEET = "Категорії Епіцентру"
MAPPING_SHEET    = "Маппінг"
EPI_ID_HEADER    = "epicenter_category_id"

# ─── Styles ───────────────────────────────────────────────────────────────────

_RED_FILL     = PatternFill("solid", fgColor="FFB3B3", bgColor="FFB3B3")
_NO_FILL      = PatternFill(fill_type=None)
_RED_RGB      = "FFB3B3"


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _find_col(ws, header: str) -> int | None:
    """1-based індекс колонки за заголовком у рядку 1. None якщо не знайдено."""
    for cell in ws[1]:
        if str(cell.value or "").strip() == header:
            return cell.column
    return None


def _is_our_red(cell) -> bool:
    """True якщо клітинка зафарбована нашим _RED_FILL."""
    try:
        return (
            cell.fill.patternType == "solid"
            and cell.fill.fgColor.rgb[-6:].upper() == _RED_RGB.upper()
        )
    except Exception:
        return False


def _max_header_col(ws) -> int:
    """Кількість колонок з заголовками у рядку 1 (для ширини рядка при фарбуванні)."""
    return max(
        (cell.column for cell in ws[1] if cell.value is not None),
        default=1,
    )


# ─── Step 1: збираємо мертві коди ─────────────────────────────────────────────

def load_deleted_codes(wb) -> frozenset[str]:
    """
    Читає «Категорії Епіцентру», повертає frozenset кодів де deleted=True.
    Падає з ValueError якщо лист або потрібні колонки відсутні.
    """
    if CATEGORIES_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Лист «{CATEGORIES_SHEET}» не знайдено. "
            f"Запусти: python scripts/epicenter_export_categories.py"
        )

    ws        = wb[CATEGORIES_SHEET]
    code_col  = _find_col(ws, "code")
    del_col   = _find_col(ws, "deleted")

    if code_col is None or del_col is None:
        raise ValueError(
            f"Не знайдено колонки в «{CATEGORIES_SHEET}»: "
            f"code={code_col}, deleted={del_col}.\n"
            f"   Оновити лист: python scripts/epicenter_export_categories.py --force"
        )

    deleted: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=False):
        code    = str(row[code_col - 1].value or "").strip()
        raw_del = row[del_col - 1].value
        if code and raw_del is True:
            deleted.add(code)

    print(f"📋 «{CATEGORIES_SHEET}»: {len(deleted)} мертвих категорій (deleted=True)")
    return frozenset(deleted)


# ─── Step 2: перевіряємо і фарбуємо «Маппінг» ────────────────────────────────

def check_mapping(wb, deleted_codes: frozenset[str], *, dry_run: bool) -> None:
    """
    Проходить «Маппінг», фарбує рядки з мертвими epicenter_category_id.
    Знімає червоний з рядків де категорія жива.
    """
    if MAPPING_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Лист «{MAPPING_SHEET}» не знайдено. "
            f"Запусти: python scripts/prom_export_categories.py"
        )

    ws      = wb[MAPPING_SHEET]
    epi_col = _find_col(ws, EPI_ID_HEADER)

    if epi_col is None:
        raise ValueError(
            f"Колонку '{EPI_ID_HEADER}' не знайдено у «{MAPPING_SHEET}». "
            f"Запусти: python scripts/prom_export_categories.py"
        )

    max_col     = _max_header_col(ws)
    highlighted = 0   # рядки з мертвою категорією
    cleared     = 0   # рядки де зняли старий червоний
    skipped     = 0   # рядки без epicenter_category_id

    dead_rows: list[tuple[int, str]] = []  # (row_num, code) для звіту

    for row in ws.iter_rows(min_row=2, max_col=max_col):
        epi_val = str(row[epi_col - 1].value or "").strip()

        # Рядок ще не заповнений маппінгом — не чіпаємо
        if not epi_val:
            skipped += 1
            continue

        is_dead  = epi_val in deleted_codes
        was_red  = _is_our_red(row[epi_col - 1])

        if is_dead:
            highlighted += 1
            dead_rows.append((row[0].row, epi_val))
            if not dry_run:
                for cell in row:
                    cell.fill = _RED_FILL
        else:
            if was_red:
                cleared += 1
            if not dry_run:
                for cell in row:
                    if _is_our_red(cell):
                        cell.fill = _NO_FILL

    # ── Звіт ──────────────────────────────────────────────────────────────────
    tag = "[dry-run] " if dry_run else ""

    print(f"\n{tag}📊 Результати «{MAPPING_SHEET}»:")
    print(f"   🔴 Мертвих рядків (підсвічено):   {highlighted}")
    print(f"   🟢 Знято підсвічування (ожили):   {cleared}")
    print(f"   ⬜ Пропущено (немає epi_id):       {skipped}")

    if dead_rows:
        print(f"\n{tag}🔴 Мертві категорії у маппінгу:")
        for row_num, code in dead_rows:
            print(f"   рядок {row_num:>4}: {code}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    dry_run = "--dry-run" in sys.argv
    label   = "  [dry-run — зміни НЕ зберігаються]" if dry_run else ""
    print(f"🔍 epicenter_check_categories.py{label}\n")

    if not XLSX_PATH.exists():
        print(f"❌ Файл не знайдено: {XLSX_PATH}")
        print("   Запусти спочатку: python scripts/epicenter_export_categories.py")
        sys.exit(1)

    wb = load_workbook(XLSX_PATH)

    try:
        deleted_codes = load_deleted_codes(wb)
    except ValueError as exc:
        print(f"❌ {exc}")
        sys.exit(1)

    if not deleted_codes:
        print("✅ Мертвих категорій не знайдено — нічого підсвічувати.")
        return

    try:
        check_mapping(wb, deleted_codes, dry_run=dry_run)
    except ValueError as exc:
        print(f"❌ {exc}")
        sys.exit(1)

    if not dry_run:
        wb.save(XLSX_PATH)
        print(f"\n✅ Збережено: {XLSX_PATH}")
    else:
        print(f"\nℹ️  Dry-run: файл не змінено.")
        print(f"   Щоб застосувати зміни: python scripts/epicenter_check_categories.py")


if __name__ == "__main__":
    main()
