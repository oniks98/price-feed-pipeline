"""
Скрипт  : python scripts/epicenter_export_royalty.py
Джерела :
    - Комісії    : https://admin.epicentrm.com.ua/public/commissions  (Angular SPA)
    - ID категорій: Google Sheets (стовпець A = ID, стовпець G = назва категорії останнього рівня)

Результат : C:\\FullStack\\PriceFeedPipeline\\data\\markets\\epicenter_royalty.xlsx
Стовпці   : ID категорії | Відкрита категорія | Відсоток роялті | parentCode

Визначення ID:
    Назви категорій, отримані зі SPA, зіставляються зі стовпцем G Google-таблиці.
    Відповідне значення стовпця A стає category_id.
    У файл потрапляють лише категорії останнього рівня (замаплені).
    Кореневі, проміжні та закриті категорії ігноруються.

Запуск:
    python scripts/epicenter_export_royalty.py               # отримати дані та зберегти в XLSX
"""

from __future__ import annotations

import argparse
import csv
import io
import logging
import re
import sys
import unicodedata
from dataclasses import dataclass, fields
from pathlib import Path
from typing import Iterator

import requests

# ---------------------------------------------------------------------------
# Конфігурація
# ---------------------------------------------------------------------------

TARGET_URL  = "https://admin.epicentrm.com.ua/public/commissions"
OUTPUT_PATH = Path(r"C:\FullStack\PriceFeedPipeline\data\markets\epicenter_royalty.xlsx")

# Google-таблиця з канонічними ID категорій
# Лист 1: «Відкриті категорії позначені зеленим»
#   Стовпець A = ID категорії  |  Стовпець G = Категорія останнього рівня
SHEET_ID  = "1Zzt9KHX5E5RPforoM924fDfdB3rx6O2TgxOKZy7t-fw"
SHEET_GID = "631872394"

# Лист 2: «Тількі відкриті категорії» — whitelist відкритих категорій
# Стовпець F (індекс 5) — назва відкритої категорії
OPEN_SHEET_GID  = "917096831"   
OPEN_COL_NAME   = 5                    # стовпець F (0-based)

# Індекси стовпців у таблиці (0-based після csv.reader)
COL_ID   = 0   # A
COL_NAME = 6   # G

PAGE_TIMEOUT_MS   = 30_000
EXPAND_TIMEOUT_MS = 5_000
SETTLE_MS         = 600
MAX_EXPAND_PASSES = 20

DIAGNOSE_SAMPLE = 30

# ---------------------------------------------------------------------------
# Схема (єдине джерело правди)
# ---------------------------------------------------------------------------

@dataclass(slots=True)
class CategoryRow:
    category_id   : str   # ID категорії  (з Google Sheets)
    category_name : str   # Відкрита категорія        (зі SPA)
    commission_pct: str   # Відсоток роялті
    parent_code   : str   # category_id безпосереднього батька


FIELDNAMES: list[str] = [f.name for f in fields(CategoryRow)]

HEADERS: dict[str, str] = {
    "category_id"   : "ID категорії",
    "category_name" : "Відкрита категорія",
    "commission_pct": "Відсоток роялті",
    "parent_code"   : "parentCode",
}

# ---------------------------------------------------------------------------
# Логування
# ---------------------------------------------------------------------------

logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    level=logging.INFO,
    stream=sys.stdout,
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Нормалізація (єдина функція для обох джерел)
# ---------------------------------------------------------------------------

def _normalize(name: str) -> str:
    """
    Агресивна нормалізація для зіставлення назв із двох різних джерел:
      1. NFC — усуває розбіжності NFD/NFC (різні байти для одного символу)
      2. Видалення невидимих Unicode-символів (zero-width space, м'який перенос тощо)
      3. Заміна всіх видів пробілів (\xa0, \u2009 і под.) на звичайний пробіл
      4. Нижній регістр + collapse пробілів
    """
    name = unicodedata.normalize("NFC", name)
    name = "".join(
        ch for ch in name
        if unicodedata.category(ch) not in ("Cc", "Cf")
    )
    name = re.sub(r"\s+", " ", name)
    return name.strip().lower()


# ---------------------------------------------------------------------------
# Google Sheets — завантаження сирого CSV
# ---------------------------------------------------------------------------

def _sheet_csv_url(sheet_id: str, gid: str) -> str:
    return (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}"
        f"/export?format=csv&gid={gid}"
    )


def _fetch_sheet_csv(sheet_id: str = SHEET_ID, gid: str = SHEET_GID) -> list[list[str]]:
    """Завантажує CSV і повертає всі рядки (включно з заголовком)."""
    url = _sheet_csv_url(sheet_id, gid)
    log.info("Завантаження Google Sheet: %s", url)
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        resp.encoding = "utf-8"  # Google Sheets CSV завжди UTF-8
    except requests.RequestException as exc:
        log.error("Не вдалося завантажити Google Sheet: %s", exc)
        sys.exit(1)
    return list(csv.reader(io.StringIO(resp.text)))


def fetch_id_map(sheet_id: str = SHEET_ID, gid: str = SHEET_GID) -> dict[str, str]:
    """
    Завантажує Google Sheet як CSV та повертає {normalized_name: category_id}.

    Враховуються лише рядки, де заповнені обидва поля: ID (стовп. A) і назва (стовп. G).
    """
    all_rows = _fetch_sheet_csv(sheet_id, gid)
    header   = all_rows[0] if all_rows else []
    log.info(
        "Заголовок таблиці (%d стовпців): %s",
        len(header),
        " | ".join(f"[{i}]{v}" for i, v in enumerate(header)),
    )

    id_map: dict[str, str] = {}
    skipped = 0

    for row in all_rows[1:]:
        if len(row) <= COL_NAME:
            skipped += 1
            continue

        raw_id   = row[COL_ID].strip()
        raw_name = row[COL_NAME].strip()

        if not raw_id or not raw_name:
            skipped += 1
            continue

        key = _normalize(raw_name)
        if key in id_map and id_map[key] != raw_id:
            log.warning(
                "Дублікат назви %r у таблиці — залишено ID=%s, ігноровано ID=%s",
                raw_name, id_map[key], raw_id,
            )
            continue

        id_map[key] = raw_id

    log.info("Мапа ID завантажена: %d записів | пропущено рядків: %d", len(id_map), skipped)
    return id_map


def fetch_open_names(
    sheet_id: str = SHEET_ID,
    gid: str = OPEN_SHEET_GID,
    col: int = OPEN_COL_NAME,
) -> frozenset[str]:
    """
    Завантажує лист «Тількі відкриті категорії» і повертає frozenset
    нормалізованих назв відкритих категорій (стовп. F).

    Повертає frozenset() якщо GID не налаштовано або лист недоступний.
    """
    if gid == "ВСТАВТЕ_GID_ТУТ":
        log.warning(
            "⚠️  OPEN_SHEET_GID не налаштовано — фільтр відкритих категорій вимкнено. "
            "Вставте GID листа 'Тількі відкриті категорії' у конфіг скрипта."
        )
        return frozenset()

    try:
        all_rows = _fetch_sheet_csv(sheet_id, gid)
    except SystemExit:
        log.error("Не вдалося завантажити лист відкритих категорій (gid=%s) — фільтр вимкнено.", gid)
        return frozenset()

    names: set[str] = set()
    for row in all_rows[1:]:
        if len(row) <= col:
            continue
        raw = row[col].strip()
        if raw:
            names.add(_normalize(raw))

    log.info(
        "Відкриті категорії (лист 2, стовп. F): %d назв завантажено",
        len(names),
    )
    return frozenset(names)


def filter_open_categories(
    id_map: dict[str, str],
    open_names: frozenset[str],
) -> dict[str, str]:
    """
    Залишає в id_map лише категорії, що є у whitelist open_names.
    Якщо open_names порожній — повертає id_map без змін (фільтр вимкнено).
    """
    if not open_names:
        log.info("Фільтр відкритих категорій: вимкнено (whitelist порожній).")
        return id_map

    filtered = {k: v for k, v in id_map.items() if k in open_names}
    closed   = len(id_map) - len(filtered)
    log.info(
        "Фільтр відкритих категорій: залишено %d | відкинуто закритих: %d",
        len(filtered),
        closed,
    )
    return filtered


# ---------------------------------------------------------------------------
# Діагностика: порівняння назв зі SPA і Google Sheet
# ---------------------------------------------------------------------------

def cmd_sheet_only(n: int = DIAGNOSE_SAMPLE) -> None:
    """Виводить перші N рядків Google Sheet (стовпці A і G) і завершує роботу."""
    all_rows = _fetch_sheet_csv()
    print(f"\n{'='*60}")
    print(f"Google Sheet — перші {n} рядків (стовп. A = ID, стовп. G = назва)")
    print(f"{'='*60}")
    header = all_rows[0] if all_rows else []
    print(f"  Заголовок: {header}")
    print()
    for i, row in enumerate(all_rows[1 : n + 1], 1):
        col_a = row[COL_ID]   if len(row) > COL_ID   else "(немає)"
        col_g = row[COL_NAME] if len(row) > COL_NAME else "(немає)"
        print(f"  {i:>3}. A={col_a!r:>8}  G={col_g!r}  norm={_normalize(col_g)!r}")
    print(f"{'='*60}\n")


def cmd_diagnose(html: str, id_map: dict[str, str], n: int = DIAGNOSE_SAMPLE) -> None:
    """
    Порівнює назви зі SPA і Google Sheet.
    Виводить repr() кожної назви — видно невидимі символи.
    """
    spa_names = [node.name for node in _iter_parsed_nodes(html)]
    sheet_keys = list(id_map.keys())

    print(f"\n{'='*60}")
    print(f"ДІАГНОСТИКА: перші {n} назв зі SPA")
    print(f"{'='*60}")
    for i, name in enumerate(spa_names[:n], 1):
        norm    = _normalize(name)
        matched = "✓" if norm in id_map else "✗"
        print(f"  {matched} {i:>3}. raw={name!r}")
        print(f"           norm={norm!r}")

    print(f"\n{'='*60}")
    print(f"ДІАГНОСТИКА: перші {n} ключів із Google Sheet")
    print(f"{'='*60}")
    for i, key in enumerate(sheet_keys[:n], 1):
        print(f"  {i:>3}. {key!r}")

    print(f"\n{'='*60}")
    print("ЧАСТКОВІ ЗБІГИ (перші 5 SPA-назв vs усі ключі sheet)")
    print(f"{'='*60}")
    for spa_name in spa_names[:5]:
        norm = _normalize(spa_name)
        candidates = [k for k in sheet_keys if norm in k or k in norm]
        print(f"  SPA raw:  {spa_name!r}")
        print(f"  SPA norm: {norm!r}")
        if candidates:
            for c in candidates[:3]:
                print(f"    ~sheet: {c!r}  ->  ID={id_map[c]}")
        else:
            print("    (жодного часткового збігу)")
        print()

    print(f"{'='*60}")
    print(f"Підсумок: SPA={len(spa_names)} назв | Sheet={len(sheet_keys)} ключів")
    print(f"{'='*60}\n")


# ---------------------------------------------------------------------------
# Playwright — завантаження та повне розгортання дерева SPA
# ---------------------------------------------------------------------------

COLLAPSED_SEL = "span.toggle-children-collapsed"
TREE_NODE_SEL = "tree-node"


def _expand_all(page) -> None:
    for pass_num in range(1, MAX_EXPAND_PASSES + 1):
        collapsed = page.query_selector_all(COLLAPSED_SEL)
        if not collapsed:
            log.info("Прохід %d: згорнутих вузлів немає -> завершено", pass_num)
            break
        log.info("Прохід %d: клік по %d стрілках...", pass_num, len(collapsed))
        for btn in collapsed:
            try:
                btn.scroll_into_view_if_needed()
                btn.click(timeout=EXPAND_TIMEOUT_MS)
            except Exception as exc:
                log.debug("Клік пропущено: %s", exc)
        page.wait_for_timeout(SETTLE_MS)
    else:
        log.warning(
            "Досягнуто MAX_EXPAND_PASSES=%d — дерево може бути неповним",
            MAX_EXPAND_PASSES,
        )


def fetch_page(url: str, show_api: bool = False) -> str:
    """Запускає headless Chromium, розгортає всі вузли дерева, повертає HTML сторінки."""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log.error(
            "playwright не встановлено. Виконайте:\n"
            "  pip install playwright\n"
            "  playwright install chromium"
        )
        sys.exit(1)

    log.info("Запуск Chromium -> %s", url)
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        page    = browser.new_page()

        if show_api:
            intercepted: list[str] = []

            def _on_response(response) -> None:
                if "json" in response.headers.get("content-type", ""):
                    intercepted.append(response.url)

            page.on("response", _on_response)

        page.goto(url, timeout=PAGE_TIMEOUT_MS, wait_until="networkidle")
        page.wait_for_selector(TREE_NODE_SEL, timeout=PAGE_TIMEOUT_MS)
        page.wait_for_timeout(1_000)

        if show_api:
            log.info("Перехоплені API-запити (%d):", len(intercepted))
            for u in intercepted:
                log.info("  %s", u)

        _expand_all(page)
        page.wait_for_load_state("networkidle", timeout=PAGE_TIMEOUT_MS)
        page.wait_for_timeout(1_000)

        html = page.content()
        browser.close()

    log.info("HTML отримано: %d байт", len(html))
    return html


# ---------------------------------------------------------------------------
# Допоміжні функції для роботи з DOM
# ---------------------------------------------------------------------------

_RE_PERCENT = re.compile(r"[\d.,]+")


def _title_text(node_div) -> str:
    span = node_div.find("span", class_="title")
    return span.get_text(strip=True) if span else ""


def _commission_text(node_div) -> str:
    ctrl = node_div.find(class_="node-controls-wrapper")
    if not ctrl:
        return ""
    raw = ctrl.get_text(strip=True)
    m = _RE_PERCENT.search(raw)
    return m.group(0).replace(",", ".") if m else ""


def _level(node_div) -> int:
    for cls in (node_div.get("class") or []):
        m = re.search(r"tree-node-level-(\d+)", cls)
        if m:
            return int(m.group(1))
    return 0


# ---------------------------------------------------------------------------
# Проміжний результат парсингу (до визначення ID)
# ---------------------------------------------------------------------------

@dataclass(slots=True)
class _ParsedNode:
    name  : str
    pct   : str
    level : int


def _iter_parsed_nodes(html: str) -> Iterator[_ParsedNode]:
    """
    Повертає вузли з повністю розгорнутого HTML SPA.
    Вузли без заголовку пропускаються.
    """
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(html, "html.parser")

    all_divs = soup.find_all(
        lambda tag: tag.name == "div"
        and any("tree-node-level-" in c for c in (tag.get("class") or []))
    )
    log.info("DOM: знайдено %d div-елементів дерева", len(all_divs))

    skipped = 0
    for div in all_divs:
        name = _title_text(div)
        if not name:
            skipped += 1
            continue
        yield _ParsedNode(name=name, pct=_commission_text(div), level=_level(div))

    if skipped:
        log.debug("DOM: пропущено %d div без заголовку", skipped)


# ---------------------------------------------------------------------------
# Побудова фінальних рядків — тільки листові (замаплені) категорії
# ---------------------------------------------------------------------------

def build_rows(html: str, id_map: dict[str, str]) -> list[CategoryRow]:
    """
    Об'єднує вузли зі SPA з ID із Google Sheets.

    Правила:
    - Стек батьків оновлюється для ВСІХ вузлів — для коректного розрахунку parentCode.
    - До результату потрапляють ТІЛЬКИ вузли, знайдені в id_map (листові категорії).
    - Кореневі та проміжні категорії (не в Google Sheet) відкидаються.
    """
    rows: list[CategoryRow] = []
    parent_stack: list[tuple[int, str]] = []   # (рівень, cat_id або "")
    skipped_root = 0

    for node in _iter_parsed_nodes(html):
        key    = _normalize(node.name)
        cat_id = id_map.get(key, "")

        # Оновлення стеку батьків для всіх вузлів (включно з незамапленими)
        while parent_stack and parent_stack[-1][0] >= node.level:
            parent_stack.pop()

        parent_code = parent_stack[-1][1] if parent_stack else ""
        parent_stack.append((node.level, cat_id))

        # Кореневі/проміжні вузли — не пишемо у файл
        if not cat_id:
            skipped_root += 1
            continue

        rows.append(CategoryRow(
            category_id   = cat_id,
            category_name = node.name,
            commission_pct= node.pct,
            parent_code   = parent_code,
        ))

    log.info(
        "Побудовано %d рядків | відкинуто кореневих/проміжних: %d",
        len(rows),
        skipped_root,
    )
    return rows


# ---------------------------------------------------------------------------
# Експорт -> XLSX  (без кольору, жирного, заливки)
# ---------------------------------------------------------------------------

def export_xlsx(rows: list[CategoryRow], path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl не встановлено. Виконайте: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Epicentr Royalty"

    for col, field in enumerate(FIELDNAMES, 1):
        ws.cell(row=1, column=col, value=HEADERS[field])

    for r_idx, row in enumerate(rows, 2):
        for c_idx, field in enumerate(FIELDNAMES, 1):
            ws.cell(row=r_idx, column=c_idx, value=getattr(row, field))

    for col, width in {1: 14, 2: 48, 3: 18, 4: 14}.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    log.info("Збережено %d рядків -> %s", len(rows), path)


# ---------------------------------------------------------------------------
# Точка входу
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Epicentr commissions -> XLSX")
    parser.add_argument("--dry-run",         action="store_true", help="Вивести результат у консоль замість XLSX")
    parser.add_argument("--show-api",         action="store_true", help="Логувати перехоплені API-запити (діагностика)")
    parser.add_argument("--diagnose",         action="store_true", help="Порівняти назви зі SPA і Google Sheet (показує repr)")
    parser.add_argument("--sheet-only",       action="store_true", help="Показати перші рядки Google Sheet і завершити")
    parser.add_argument("--no-open-filter",   action="store_true", help="Вимкнути фільтр відкритих категорій (debug)")
    args = parser.parse_args()

    if args.sheet_only:
        cmd_sheet_only()
        return

    id_map     = fetch_id_map()
    open_names = frozenset() if args.no_open_filter else fetch_open_names()
    id_map     = filter_open_categories(id_map, open_names)
    html       = fetch_page(TARGET_URL, show_api=args.show_api)

    if args.diagnose:
        cmd_diagnose(html, id_map)
        return

    rows = build_rows(html, id_map)

    if not rows:
        log.error("Рядків не знайдено. Завершення.")
        sys.exit(1)

    if args.dry_run:
        print("\t".join(HEADERS[f] for f in FIELDNAMES))
        for r in rows:
            print("\t".join(str(getattr(r, f)) for f in FIELDNAMES))
        log.info("Dry-run: %d рядків", len(rows))
    else:
        export_xlsx(rows, OUTPUT_PATH)


if __name__ == "__main__":
    main()
