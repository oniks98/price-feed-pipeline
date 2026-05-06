"""
epicenter_export_attr_sets.py
------------------------------
КРОК 3 з 5 пайплайну маппінгу Prom → Epicenter.

Що робить:
  Завантажує з API сети атрибутів Epicenter і записує в лист «Сети атрибутів»
  ТІЛЬКИ для тих set_code, які є в колонці epicenter_category_id листа «Маппінг».

Передумова:
  Лист «Маппінг» в epicenter_mappings.xlsx повинен мати заповнені
  epicenter_category_id (крок 2 пайплайну).

Інкрементальна логіка:
  Якщо лист «Сети атрибутів» вже існує — дописує тільки нові set_codes,
  існуючі рядки не чіпає.

Наступний крок:
  Заповни prom_param_name у «Сети атрибутів» (колонка J):
    • Автоматично: python scripts/epicenter_map_attributes.py
    • Вручну:      жовта колонка J
  Потім → python scripts/epicenter_export_attr_options.py

Запуск:
    python scripts/epicenter_export_attr_sets.py
"""

from __future__ import annotations

from pathlib import Path

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_TOKEN = "5a6489d1a5c48c9d174bd31f2a0a8fd0"
BASE_URL  = "https://api.epicentrm.com.ua/v2/pim"
HEADERS   = {
    "Authorization": f"Bearer {API_TOKEN}",
    "Content-Type": "application/json",
    "Accept": "application/json",
}

ROOT        = Path(__file__).parents[1]
OUTPUT_PATH = ROOT / "data" / "markets" / "epicenter_mappings.xlsx"

REQ_TIMEOUT = (10, 30)


# ─── Session ──────────────────────────────────────────────────────────────────

def _make_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(total=2, backoff_factor=0.5, status_forcelist=(429,), allowed_methods=["GET"])
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update(HEADERS)
    return session


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _norm_id(val: object) -> str:
    """Нормалізує ідентифікатор: '123.0' → '123'."""
    if val is None:
        return ""
    s = str(val).strip()
    if "." in s:
        try:
            f = float(s)
            if f == int(f):
                return str(int(f))
        except (ValueError, OverflowError):
            pass
    return s


def _get_translation(translations: list[dict], lang: str = "ua") -> str:
    for priority_lang in (lang, "ua", "uk", "ru", "en"):
        for t in translations:
            if t.get("languageCode") == priority_lang:
                val = t.get("value") or t.get("title") or ""
                if str(val).strip():
                    return str(val).strip()
    return ""


# ─── Styles ───────────────────────────────────────────────────────────────────

HDR_FILL    = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
YELLOW_FILL = PatternFill("solid", start_color="FFFF99", end_color="FFFF99")
GRAY_FILL   = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _hdr(cell, fill=HDR_FILL) -> None:
    cell.font = HDR_FONT
    cell.fill = fill
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def _data(cell, fill=None) -> None:
    cell.font = Font(name="Arial", size=9)
    cell.alignment = LEFT
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill


# ─── Readers ──────────────────────────────────────────────────────────────────

def load_mapped_set_codes() -> set[str]:
    """Повертає epicenter_category_id з листа «Маппінг» (тільки заповнені)."""
    if not OUTPUT_PATH.exists():
        return set()
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(OUTPUT_PATH, read_only=True, data_only=True)
        if "Маппінг" not in wb.sheetnames:
            wb.close()
            return set()
        rows = list(wb["Маппінг"].iter_rows(values_only=True))
        wb.close()
        if not rows:
            return set()
        headers = [str(c).strip() if c else "" for c in rows[0]]
        try:
            col = headers.index("epicenter_category_id")
        except ValueError:
            return set()
        codes = {
            _norm_id(row[col])
            for row in rows[1:]
            if len(row) > col and row[col]
        }
        print(f"   Знайдено {len(codes)} заповнених epicenter_category_id у «Маппінг».")
        return codes
    except Exception as e:
        print(f"⚠️  Не вдалося прочитати «Маппінг»: {e}")
        return set()


def load_existing_set_codes() -> set[str]:
    """Повертає set_codes, які вже є в листі «Сети атрибутів»."""
    if not OUTPUT_PATH.exists():
        return set()
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(OUTPUT_PATH, read_only=True, data_only=True)
        if "Сети атрибутів" not in wb.sheetnames:
            wb.close()
            return set()
        rows = list(wb["Сети атрибутів"].iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2:
            return set()
        headers = [str(c).strip() if c else "" for c in rows[0]]
        try:
            sc_col = headers.index("set_code")
        except ValueError:
            return set()
        codes = {
            _norm_id(row[sc_col])
            for row in rows[1:]
            if len(row) > sc_col and row[sc_col]
        }
        print(f"   set_codes вже в «Сети атрибутів»: {len(codes)} шт.")
        return codes
    except Exception as e:
        print(f"⚠️  Не вдалося прочитати «Сети атрибутів»: {e}")
        return set()


# ─── API ──────────────────────────────────────────────────────────────────────

def fetch_attribute_sets() -> list[dict]:
    print("⬇️  Сети атрибутів з API...")
    session = _make_session()
    sets: list[dict] = []
    page = 1
    while True:
        try:
            resp = session.get(
                f"{BASE_URL}/attribute-sets", params={"page": page}, timeout=REQ_TIMEOUT
            )
            print(f"   сторінка {page} → HTTP {resp.status_code}")
            if resp.status_code == 403:
                print(f"❌ 403 Forbidden: {resp.text[:200]}")
                break
            data = resp.json()
        except Exception as e:
            print(f"❌ attribute-sets p{page}: {e}")
            break
        batch = data.get("items", [])
        if not batch:
            break
        sets.extend(batch)
        total = data.get("pages", 1)
        print(f"   {page}/{total}: {len(batch)} сетів")
        if page >= total:
            break
        page += 1
    print(f"✅ Сетів отримано: {len(sets)}")
    return sets


# ─── Writers ──────────────────────────────────────────────────────────────────

def _attr_set_to_rows(aset: dict) -> list[list]:
    """Перетворює один сет атрибутів у список рядків для запису в xlsx."""
    sc    = _norm_id(aset.get("code", ""))
    sn    = _get_translation(aset.get("translations", []))
    attrs = aset.get("attributes", [])
    rows = [
        [
            sc, sn,
            a.get("code", ""),
            _get_translation(a.get("translations", [])),
            a.get("type", ""),
            a.get("isRequired", False),
            a.get("isFilter", False),
            a.get("isSystem", False),
            a.get("isModel", False),
            "",  # prom_param_name — заповнить користувач або epicenter_map_attributes.py
        ]
        for a in attrs
    ]
    return rows or [[sc, sn, "", "", "", "", "", "", "", ""]]


def build_attr_sets_sheet(wb: Workbook, attr_sets: list[dict], filter_set_codes: set[str]) -> None:
    """Створює лист «Сети атрибутів» тільки для set_codes з filter_set_codes."""
    ws = wb.create_sheet("Сети атрибутів")
    headers    = ["set_code", "set_name_uk", "attr_code", "attr_name_uk", "attr_type",
                  "isRequired", "isFilter", "isSystem", "isModel", "prom_param_name"]
    col_widths = [30, 40, 30, 42, 16, 12, 12, 12, 12, 35]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        _hdr(ws.cell(row=1, column=ci, value=h))
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.cell(
        row=1, column=12,
        value="🟡 J (prom_param_name) — заповни вручну або запусти epicenter_map_attributes.py | 🔴 = обов'язково",
    ).font = Font(bold=True, color="7F6000", name="Arial", size=9)

    written = 0
    row_idx = 2
    for aset in attr_sets:
        if _norm_id(aset.get("code", "")) not in filter_set_codes:
            continue
        for row in _attr_set_to_rows(aset):
            for ci, val in enumerate(row, 1):
                _data(ws.cell(row=row_idx, column=ci, value=val),
                      fill=GRAY_FILL if ci <= 2 else None)
            row_idx += 1
            written += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    print(f"   ✅ Записано {written} рядків у «Сети атрибутів».")


def append_new_attr_sets(attr_sets: list[dict], new_set_codes: set[str]) -> int:
    """Дописує в існуючий лист «Сети атрибутів» рядки тільки для нових set_codes."""
    import openpyxl as _xl
    wb = _xl.load_workbook(OUTPUT_PATH)

    if "Сети атрибутів" not in wb.sheetnames:
        build_attr_sets_sheet(wb, attr_sets, new_set_codes)
        wb.save(OUTPUT_PATH)
        wb.close()
        return sum(len(_attr_set_to_rows(s)) for s in attr_sets
                   if _norm_id(s.get("code", "")) in new_set_codes)

    ws = wb["Сети атрибутів"]
    added = 0
    for aset in attr_sets:
        if _norm_id(aset.get("code", "")) not in new_set_codes:
            continue
        next_row = ws.max_row + 1
        for row in _attr_set_to_rows(aset):
            for ci, val in enumerate(row, 1):
                _data(ws.cell(row=next_row, column=ci, value=val),
                      fill=GRAY_FILL if ci <= 2 else None)
            next_row += 1
            added += 1

    wb.save(OUTPUT_PATH)
    wb.close()
    print(f"   ✅ Дописано {added} рядків для {len(new_set_codes)} нових set_codes.")
    return added


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    print("🚀 epicenter_export_attr_sets.py — КРОК 3\n")

    if not OUTPUT_PATH.exists():
        print(
            "❌ epicenter_mappings.xlsx не знайдено.\n"
            "   Спочатку виконай КРОК 1: python scripts/epicenter_export_categories.py"
        )
        return

    mapped_set_codes = load_mapped_set_codes()
    if not mapped_set_codes:
        print(
            "\n⏭️  Немає заповнених epicenter_category_id у «Маппінг».\n"
            "   Виконай КРОК 2:\n"
            "   • Автоматично: python scripts/epicenter_map_categories.py\n"
            "   • Вручну:      колонки C, D, E у листі «Маппінг»"
        )
        return

    existing_set_codes = load_existing_set_codes()
    new_set_codes = mapped_set_codes - existing_set_codes

    if not new_set_codes:
        print(
            "   ✅ Сети атрибутів вже завантажено для всіх категорій.\n"
            "   Наступний крок → python scripts/epicenter_export_attr_options.py"
        )
        return

    print(f"   Нових set_codes для завантаження: {len(new_set_codes)} шт.")
    attr_sets = fetch_attribute_sets()

    if not attr_sets:
        print("❌ Сети атрибутів не отримано з API. Перевір токен.")
        return

    added = append_new_attr_sets(attr_sets, new_set_codes)

    if added > 0:
        print(
            f"\n📌 Записано атрибути для {len(new_set_codes)} категорій.\n"
            "   КРОК 4: Заповни prom_param_name (колонка J) у «Сети атрибутів»:\n"
            "   • Автоматично: python scripts/epicenter_map_attributes.py\n"
            "   • Вручну:      жовта колонка J\n"
            "   Потім → python scripts/epicenter_export_attr_options.py"
        )
    else:
        print(
            "\n⚠️  Жоден сет не записано — можливо set_codes з маппінгу не знайдено в API.\n"
            f"   Очікувані set_codes: {sorted(new_set_codes)}"
        )


if __name__ == "__main__":
    main()
