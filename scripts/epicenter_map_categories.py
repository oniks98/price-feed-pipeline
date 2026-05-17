"""
epicenter_map_categories.py

Запуск:
    python scripts/epicenter_map_categories.py

Логіка:
  1. Завантажує відкриті категорії Епіцентру з royalty_epicenter.xlsx
     (лист ROYALTY_SHEET, колонка ROYALTY_COL_ID).
  2. Фільтрує лист "Категорії Епіцентру" з epicenter_mappings.xlsx —
     залишає лише ті рядки, чий code є серед відкритих.
  3. Зіставляє категорії Прому з відфільтрованим набором (fuzzy, ≥80%).
  4. Записує назад: epicenter_category_id, Назва категорії Епіцентру, parentCode.
     parentCode береться з листа "Категорії Епіцентру" (там він точний).
"""

import re
import logging
from pathlib import Path

import openpyxl
from rapidfuzz import fuzz

# ──────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s | %(message)s",
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────
ROOT         = Path(__file__).parents[1]
MAPPINGS_PATH = ROOT / "data" / "markets" / "epicenter_mappings.xlsx"
ROYALTY_PATH  = ROOT / "data" / "markets" / "royalty_epicenter.xlsx"

# Sheets
SHEET_MAPPING   = "Маппінг"
SHEET_EPICENTER = "Категорії Епіцентру"
ROYALTY_SHEET   = "Epicentr Royalty"

# Columns — royalty file
ROYALTY_COL_ID   = "ID категорії"       # open epicenter category id
ROYALTY_COL_NAME = "Відкрита категорія" # category name (informational)

# Columns — "Маппінг" sheet
COL_PROMO    = "Категорія Прому"
COL_EPI_ID   = "epicenter_category_id"
COL_EPI_NAME = "Назва категорії Епіцентру"
COL_PARENT   = "parentCode"

# Columns — "Категорії Епіцентру" sheet
EPI_COL_CODE   = "code"
EPI_COL_NAME   = "name_uk"
EPI_COL_PARENT = "parentCode"

MATCH_THRESHOLD = 80  # percent


# ──────────────────────────────────────────────
# Text helpers
# ──────────────────────────────────────────────

def extract_last_segment(text: str) -> str:
    """Return the part after the last '>', or the full string if none."""
    if not text:
        return ""
    return text.split(">")[-1].strip()


def normalize(text: str) -> str:
    """Lowercase, remove punctuation, collapse spaces."""
    text = text.lower()
    text = re.sub(r"[^\w\s]", " ", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


def stem_uk(word: str) -> str:
    """Lightweight Ukrainian/Russian suffix stripper for category matching."""
    suffixes = [
        "ання", "ення", "іння", "яння",
        "ація", "яція", "ування", "ювання",
        "ський", "зький", "цький",
        "ний", "ній", "ова", "ові", "ого",
        "ів", "ій", "их", "ах", "ам", "ом", "ем", "им",
        "ание", "ение", "ование", "ации",
        "ский", "зкий", "ный", "ной", "ных", "ным",
        "ый", "ые", "ов", "ев",
        "ій", "ая", "ое",
    ]
    for sfx in sorted(suffixes, key=len, reverse=True):
        if word.endswith(sfx) and len(word) - len(sfx) >= 3:
            return word[: -len(sfx)]
    return word


def tokenize(text: str) -> list[str]:
    """Normalize → split → stem; drop tokens shorter than 3 chars."""
    return [stem_uk(t) for t in normalize(text).split() if len(t) > 2]


def token_overlap_score(query_tokens: list[str], target_tokens: list[str]) -> float:
    """Percentage of query tokens with a fuzzy match (≥80) in target. Returns 0–100."""
    if not query_tokens or not target_tokens:
        return 0.0
    matched = sum(
        1
        for qt in query_tokens
        if any(fuzz.ratio(qt, tt) >= 80 for tt in target_tokens)
    )
    return (matched / len(query_tokens)) * 100


# ──────────────────────────────────────────────
# Data loaders
# ──────────────────────────────────────────────

def load_open_category_ids(royalty_path: Path) -> frozenset[str]:
    """
    Read royalty_epicenter.xlsx and return a frozenset of open category IDs
    (column ROYALTY_COL_ID, all non-empty rows).
    """
    if not royalty_path.exists():
        raise FileNotFoundError(f"Royalty file not found: {royalty_path}")

    wb = openpyxl.load_workbook(royalty_path, read_only=True, data_only=True)

    if ROYALTY_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{ROYALTY_SHEET}' not found in {royalty_path.name}. "
            f"Available: {wb.sheetnames}"
        )

    ws = wb[ROYALTY_SHEET]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {str(v).strip(): i for i, v in enumerate(header_row) if v is not None}

    if ROYALTY_COL_ID not in headers:
        raise ValueError(
            f"Column '{ROYALTY_COL_ID}' not found in '{ROYALTY_SHEET}'. "
            f"Found: {list(headers)}"
        )

    id_col = headers[ROYALTY_COL_ID]
    open_ids: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = row[id_col]
        if raw is not None and str(raw).strip():
            open_ids.add(str(raw).strip())

    wb.close()
    log.info("Loaded %d open category IDs from '%s'.", len(open_ids), royalty_path.name)
    return frozenset(open_ids)


def load_epicenter_categories(
    ws_epi,
    open_ids: frozenset[str],
) -> list[dict]:
    """
    Read 'Категорії Епіцентру' sheet and return only rows whose `code`
    is in open_ids. Each row is a dict keyed by column header.
    """
    header_row = next(ws_epi.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(v).strip() if v is not None else "" for v in header_row]

    if EPI_COL_CODE not in headers:
        raise ValueError(
            f"Column '{EPI_COL_CODE}' not found in '{SHEET_EPICENTER}'. "
            f"Found: {headers}"
        )

    rows: list[dict] = []
    for row in ws_epi.iter_rows(min_row=2, values_only=True):
        record = {headers[i]: (v if v is not None else "") for i, v in enumerate(row)}
        code = str(record.get(EPI_COL_CODE, "")).strip()
        if code in open_ids:
            rows.append(record)

    log.info(
        "Filtered Epicenter categories: %d open out of sheet total.",
        len(rows),
    )
    return rows


def read_sheet_headers(ws) -> dict[str, int]:
    """Return {column_name: 1-based column index} from the first row."""
    return {
        str(cell.value).strip(): j + 1
        for j, cell in enumerate(ws[1])
        if cell.value is not None
    }


# ──────────────────────────────────────────────
# Matching
# ──────────────────────────────────────────────

def best_match(
    query: str,
    epi_rows: list[dict],
) -> dict | None:
    """
    Find the best matching Epicenter row for a query (Prom category path).
    Compares only against the last path segment.
    Returns the row dict or None if best score is below MATCH_THRESHOLD.
    """
    seg = extract_last_segment(query)
    q_tokens = tokenize(seg)
    seg_norm = normalize(seg)

    best_score = 0.0
    best_row: dict | None = None

    for row in epi_rows:
        name = str(row.get(EPI_COL_NAME) or "")
        t_tokens = tokenize(name)

        score = max(
            token_overlap_score(q_tokens, t_tokens) if q_tokens else 0.0,
            fuzz.partial_ratio(seg_norm, normalize(name)),
        )

        if score > best_score:
            best_score = score
            best_row = row

    if best_score >= MATCH_THRESHOLD and best_row is not None:
        log.info(
            "  ✓ [%3.0f%%] '%s'  →  '%s'",
            best_score,
            seg,
            best_row.get(EPI_COL_NAME, ""),
        )
        return best_row

    log.info("  ✗ [%3.0f%%] '%s'  —  no match", best_score, seg)
    return None


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

def main() -> None:
    # ── 1. Load open category IDs from royalty file ──────────────────────────
    open_ids = load_open_category_ids(ROYALTY_PATH)

    # ── 2. Open mappings workbook ─────────────────────────────────────────────
    if not MAPPINGS_PATH.exists():
        raise FileNotFoundError(f"Mappings file not found: {MAPPINGS_PATH}")

    wb = openpyxl.load_workbook(MAPPINGS_PATH)

    for sheet in (SHEET_MAPPING, SHEET_EPICENTER):
        if sheet not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet}' not found. Available: {wb.sheetnames}"
            )

    ws_epi = wb[SHEET_EPICENTER]
    ws_map = wb[SHEET_MAPPING]

    # ── 3. Load only open Epicenter categories ────────────────────────────────
    epi_rows = load_epicenter_categories(ws_epi, open_ids)

    # ── 4. Resolve / create output columns in Mapping sheet ──────────────────
    map_headers = read_sheet_headers(ws_map)

    def ensure_col(name: str) -> int:
        if name not in map_headers:
            next_col = max(map_headers.values()) + 1
            ws_map.cell(1, next_col, name)
            map_headers[name] = next_col
            log.info("Added column '%s' at position %d.", name, next_col)
        return map_headers[name]

    col_epi_id   = ensure_col(COL_EPI_ID)
    col_epi_name = ensure_col(COL_EPI_NAME)
    col_parent   = ensure_col(COL_PARENT)

    if COL_PROMO not in map_headers:
        raise ValueError(
            f"Column '{COL_PROMO}' not found in '{SHEET_MAPPING}'. "
            f"Found: {list(map_headers)}"
        )
    col_promo = map_headers[COL_PROMO]

    # Fast lookup: epicenter code → full row (used for parentCode back-fill)
    epi_by_code: dict[str, dict] = {
        str(r.get(EPI_COL_CODE, "")).strip(): r
        for r in epi_rows
        if r.get(EPI_COL_CODE)
    }

    # ── 5. Iterate Mapping rows and fill matches ──────────────────────────────
    updated         = 0
    skipped         = 0
    already_matched = 0
    parent_fixed    = 0

    for row_idx in range(2, ws_map.max_row + 1):
        promo_val = str(ws_map.cell(row_idx, col_promo).value or "").strip()
        if not promo_val:
            continue

        existing_epi_id   = str(ws_map.cell(row_idx, col_epi_id).value or "").strip()
        existing_epi_name = str(ws_map.cell(row_idx, col_epi_name).value or "").strip()
        existing_parent   = str(ws_map.cell(row_idx, col_parent).value or "").strip()

        # Case 1: all three columns filled — nothing to do
        if existing_epi_id and existing_epi_name and existing_parent:
            already_matched += 1
            continue

        # Case 2: epi_id + name filled, but parentCode missing → back-fill from lookup
        if existing_epi_id and existing_epi_name and not existing_parent:
            epi_row = epi_by_code.get(existing_epi_id)
            if epi_row:
                ws_map.cell(row_idx, col_parent).value = epi_row.get(EPI_COL_PARENT, "")
                log.info(
                    "Row %d: back-filled parentCode '%s' for epi_id '%s'.",
                    row_idx,
                    epi_row.get(EPI_COL_PARENT, ""),
                    existing_epi_id,
                )
                parent_fixed += 1
            else:
                log.warning(
                    "Row %d: epi_id '%s' not in open categories — parentCode left empty.",
                    row_idx,
                    existing_epi_id,
                )
            continue

        # Case 3: not matched yet → run fuzzy match
        log.info("Row %d: '%s'", row_idx, promo_val)
        match = best_match(promo_val, epi_rows)

        if match:
            # parentCode from "Категорії Епіцентру" — authoritative source
            ws_map.cell(row_idx, col_epi_id).value   = match.get(EPI_COL_CODE, "")
            ws_map.cell(row_idx, col_epi_name).value = match.get(EPI_COL_NAME, "")
            ws_map.cell(row_idx, col_parent).value   = match.get(EPI_COL_PARENT, "")
            updated += 1
        else:
            skipped += 1

    # ── 6. Save ───────────────────────────────────────────────────────────────
    wb.save(MAPPINGS_PATH)
    log.info(
        "Done. Updated: %d | parentCode fixed: %d | No match: %d | Already complete (skipped): %d",
        updated,
        parent_fixed,
        skipped,
        already_matched,
    )
    log.info("Saved → %s", MAPPINGS_PATH)


if __name__ == "__main__":
    main()