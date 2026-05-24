"""
prom_export_categories.py
-------------------------
Синхронізує категорії з фіду PROM з локальними файлами маркетплейсів.
Експортує ТІЛЬКИ ті категорії, під якими є реальні товари у фіді.
Додає НОВІ категорії (яких ще немає по ID) до:
  - data/markets/mappings.xlsx (лист 'Категорія+')    — маппінг категорій Kasta
  - data/markets/epicenter_mappings.xlsx (лист 'Маппінг') — маппінг категорій Epicenter
  - data/markets/rozetka_mappings.xlsx (лист 'Маппінг')   — маппінг категорій Rozetka

Запуск:
    python scripts/prom_export_categories.py
"""

import csv
import re
from dataclasses import dataclass
from pathlib import Path

import requests
from openpyxl import load_workbook

from constants_feed_url import FEED_URL_PROM as FEED_URL

# ─── Config: feed ─────────────────────────────────────────────────────────────

_ROOT = Path(__file__).parents[1]

# ─── Config: mappings.xlsx ────────────────────────────────────────────────────

MAPPINGS_XLSX     = _ROOT / "data" / "markets" / "mappings.xlsx"
MAPPINGS_SHEET    = "Категорія+"
MAPPINGS_ID_COL   = "ІD категорії фіду"
MAPPINGS_NAME_COL = "Категорії фіду"

# ─── Config: epicenter_mappings.xlsx ─────────────────────────────────────────

EPICENTER_MAPPINGS_XLSX     = _ROOT / "data" / "markets" / "epicenter_mappings.xlsx"
EPICENTER_MAPPINGS_SHEET    = "Маппінг"
EPICENTER_MAPPINGS_ID_COL   = "prom_category_id"
EPICENTER_MAPPINGS_NAME_COL = "Категорія Прому"

# ─── Config: rozetka_mappings.xlsx ───────────────────────────────────────────

ROZETKA_MAPPINGS_XLSX     = _ROOT / "data" / "markets" / "rozetka_mappings.xlsx"
ROZETKA_MAPPINGS_SHEET    = "Маппінг"
ROZETKA_MAPPINGS_ID_COL   = "prom_category_id"
ROZETKA_MAPPINGS_NAME_COL = "Категорія Прому"

# ─── Config: prom_coefficients.csv ──────────────────────────────────────────
# Колонки: prom_category_id | prom_category_name | coef | coef_uncategorized | coef_no_base
# Рядок 1 — заголовки, рядок 2 — дефолтні коефіцієнти (sentinel, без ID), рядки 3+ — дані.

PROM_COEF_CSV = _ROOT / "data" / "markets" / "prom_coefficients.csv"

# ─── Excel target descriptor ──────────────────────────────────────────────────

@dataclass(frozen=True)
class ExcelTarget:
    """Описує один Excel-файл + лист для інкрементального оновлення категорій."""
    path: Path
    sheet: str
    id_col: str    # заголовок колонки з ID категорії
    name_col: str  # заголовок колонки з назвою категорії


# Єдине місце для реєстрації всіх Excel-файлів, що оновлюються.
EXCEL_TARGETS: tuple[ExcelTarget, ...] = (
    ExcelTarget(
        path=MAPPINGS_XLSX,
        sheet=MAPPINGS_SHEET,
        id_col=MAPPINGS_ID_COL,
        name_col=MAPPINGS_NAME_COL,
    ),
    ExcelTarget(
        path=EPICENTER_MAPPINGS_XLSX,
        sheet=EPICENTER_MAPPINGS_SHEET,
        id_col=EPICENTER_MAPPINGS_ID_COL,
        name_col=EPICENTER_MAPPINGS_NAME_COL,
    ),
    ExcelTarget(
        path=ROZETKA_MAPPINGS_XLSX,
        sheet=ROZETKA_MAPPINGS_SHEET,
        id_col=ROZETKA_MAPPINGS_ID_COL,
        name_col=ROZETKA_MAPPINGS_NAME_COL,
    ),
)


# ─── Feed fetching ────────────────────────────────────────────────────────────

def fetch_xml(url: str) -> str:
    response = requests.get(url, timeout=120)
    response.raise_for_status()
    raw = response.content
    match = re.search(rb'encoding=["\']([^"\']+)["\']', raw[:200])
    encoding = match.group(1).decode("ascii") if match else (response.encoding or "utf-8")
    print(f"🔍 Кодування: {encoding}")
    return raw.decode(encoding)


# ─── Feed parsing ─────────────────────────────────────────────────────────────

def parse_categories(xml: str) -> dict[str, dict]:
    """Повертає {id: {name, parentId}} для всіх категорій у фіді."""
    pattern = r'<category\s+id="(\d+)"(?:\s+parentId="(\d+)")?[^>]*>(.*?)</category>'
    return {
        cat_id: {"name": name.strip(), "parentId": parent_id or None}
        for cat_id, parent_id, name in re.findall(pattern, xml)
    }


def parse_used_category_ids(xml: str) -> set[str]:
    """Повертає set id категорій, які реально використовуються в офферах."""
    return set(re.findall(r"<categoryId>(\d+)</categoryId>", xml))


# ─── Category helpers ─────────────────────────────────────────────────────────

def build_display_name(cat_id: str, categories: dict[str, dict]) -> str:
    """
    Будує повну назву категорії з батьківською: 'Батьківська > Дочірня'.
    Якщо батька немає — просто назва.
    """
    cat = categories.get(cat_id)
    if not cat:
        return cat_id

    name = cat["name"]
    parent_id = cat["parentId"]

    if parent_id and parent_id in categories:
        return f"{categories[parent_id]['name']} > {name}"

    return name


def filter_active_categories(
    all_categories: dict[str, dict],
    used_ids: set[str],
) -> dict[str, dict]:
    """Залишає тільки категорії, під якими є реальні товари."""
    return {cid: cat for cid, cat in all_categories.items() if cid in used_ids}


# ─── Excel: shared incremental writer ────────────────────────────────────────

def _resolve_column_indices(
    header: list,
    id_col: str,
    name_col: str,
    target_label: str,
) -> tuple[int, int] | None:
    """
    Повертає (id_col_idx, name_col_idx) — 1-based.
    Логує та повертає None при помилці (не падає).
    """
    try:
        id_idx   = header.index(id_col) + 1
        name_idx = header.index(name_col) + 1
        return id_idx, name_idx
    except ValueError as e:
        print(f"⚠️  [{target_label}] Колонку не знайдено: {e}")
        return None


def _collect_existing_ids(ws, id_col_idx: int) -> set[str]:
    """Збирає вже існуючі category ID з рядків 2+ (рядок 1 — заголовок)."""
    existing: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[id_col_idx - 1]
        if val is not None:
            existing.add(str(val))
    return existing


def append_new_categories_to_excel(
    target: ExcelTarget,
    active_categories: dict[str, dict],
    all_categories: dict[str, dict],
) -> None:
    """
    Дописує в Excel-файл (target.sheet) тільки НОВІ категорії по ID.
    Існуючі рядки не чіпає. Ідемпотентна операція.
    """
    label = f"{target.path.name} / '{target.sheet}'"

    if not target.path.exists():
        print(f"⚠️  Файл не знайдено: {target.path}")
        return

    wb = load_workbook(target.path)

    if target.sheet not in wb.sheetnames:
        print(f"⚠️  [{label}] Лист не знайдено. Доступні: {wb.sheetnames}")
        return

    ws = wb[target.sheet]
    header = [cell.value for cell in ws[1]]

    col_indices = _resolve_column_indices(header, target.id_col, target.name_col, label)
    if col_indices is None:
        return

    id_col_idx, name_col_idx = col_indices
    existing_ids = _collect_existing_ids(ws, id_col_idx)

    new_categories = {
        cat_id: cat
        for cat_id, cat in active_categories.items()
        if str(cat_id) not in existing_ids
    }

    if not new_categories:
        print(f"✅ [{label}] Нових категорій немає — файл не змінено")
        return

    for cat_id in sorted(new_categories, key=lambda x: int(x)):
        display_name = build_display_name(cat_id, all_categories)
        new_row = [None] * len(header)
        new_row[id_col_idx - 1]   = int(cat_id)
        new_row[name_col_idx - 1] = display_name
        ws.append(new_row)

    wb.save(target.path)
    print(f"✅ [{label}] Додано {len(new_categories)} нових категорій → {target.path}")
    for cat_id in sorted(new_categories, key=lambda x: int(x)):
        print(f"   + [{cat_id}] {build_display_name(cat_id, all_categories)}")


# ─── CSV: prom_coefficients ────────────────────────────────────────────────

def update_prom_coef_csv(
    active_categories: dict[str, dict],
    all_categories: dict[str, dict],
) -> None:
    """Дописує в prom_coefficients.csv тільки НОВІ категорії. Поля coef не заповнюються."""
    if not PROM_COEF_CSV.exists():
        print(f"⚠️  prom_coefficients.csv не знайдено: {PROM_COEF_CSV}")
        return

    with PROM_COEF_CSV.open("r", encoding="utf-8-sig") as f:
        reader     = csv.DictReader(f, delimiter=";")
        fieldnames = list(reader.fieldnames or [])
        rows       = list(reader)

    existing_ids = {row["prom_category_id"] for row in rows if row.get("prom_category_id")}

    new_categories = {
        cat_id: cat
        for cat_id, cat in active_categories.items()
        if cat_id not in existing_ids
    }

    if not new_categories:
        print("✅ Нових категорій немає — prom_coefficients.csv не змінено")
        return

    with PROM_COEF_CSV.open("a", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";", extrasaction="ignore")
        for cat_id in sorted(new_categories, key=lambda x: int(x)):
            writer.writerow({
                "prom_category_id":   cat_id,
                "prom_category_name": build_display_name(cat_id, all_categories),
            })

    print(f"✅ Додано {len(new_categories)} нових категорій → {PROM_COEF_CSV}")
    for cat_id in sorted(new_categories, key=lambda x: int(x)):
        print(f"   + [{cat_id}] {build_display_name(cat_id, all_categories)}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    print("⬇️  Завантаження фіду...")
    xml = fetch_xml(FEED_URL)
    print(f"📄 Отримано {len(xml):,} символів")

    all_categories = parse_categories(xml)
    used_ids       = parse_used_category_ids(xml)

    if not all_categories:
        print("⚠️  Категорії не знайдено — перевір URL фіду")
        return

    if not used_ids:
        print("⚠️  Товари не знайдено — перевір структуру фіду")
        return

    active_categories = filter_active_categories(all_categories, used_ids)

    skipped = len(all_categories) - len(active_categories)
    print(
        f"📦 Всього категорій: {len(all_categories)}, "
        f"з товарами: {len(active_categories)}, "
        f"пропущено порожніх: {skipped}"
    )

    # ── Excel targets (всі зареєстровані файли) ───────────────────────────────
    for target in EXCEL_TARGETS:
        append_new_categories_to_excel(target, active_categories, all_categories)

    # ── CSV ───────────────────────────────────────────────────────────────────
    update_prom_coef_csv(active_categories, all_categories)


if __name__ == "__main__":
    main()
