"""
epicenter_attr_defaults.py
---------------------------
Допоміжний скрипт для вибору дефолтних опцій атрибутів.

Проблема:
  Деякі обов'язкові атрибути Epicenter (isRequired=True, тип select/multiselect)
  не мають відповідника у постачальника. Для таких товарів треба підставити
  дефолтне значення — інакше Epicenter не прийме товар.

Що робить:
  1. Читає лист «Опції атрибутів» з epicenter_mappings.xlsx.
  2. Вибирає рядки де needs_default=True (обов'язкові атрибути без prom_param_name).
  3. Групує по set_code → attr_code → список доступних опцій.
  4. Якщо файл дефолтів вже існує — зберігає вже вибрані значення (не перезаписує).
  5. Генерує / оновлює data/markets/epicenter_attr_defaults.json.

Формат JSON:
  {
    "<set_code>": {
      "<attr_code>": {
        "attr_name":  "Назва атрибута",
        "attr_type":  "select" | "multiselect",
        "default_option_codes": [],   ← ЗАПОВНИТИ: один code для select, кілька для multiselect
        "default_option_names": [],   ← для зручності читання (заповнюється разом з codes)
        "available_options": [
          {"code": "...", "name": "..."},
          ...
        ]
      }
    }
  }

Правило:
  select      → default_option_codes містить рівно 1 елемент
  multiselect → default_option_codes може містити 1+ елементів
  Feed generator завжди ітерує по списку — однакова логіка для обох типів.

Як користуватись:
  1. Запусти скрипт → отримай epicenter_attr_defaults.json
  2. Відкрий JSON, для кожного атрибута вибери опцію зі списку available_options
  3. Заповни default_option_code і default_option_name
  4. generate_epicenter_feed.py читає цей файл як fallback для товарів
     без відповідного атрибута від постачальника

Запуск:
    python scripts/epicenter_attr_defaults.py

  Перегляд статистики (без запису):
    python scripts/epicenter_attr_defaults.py --dry-run
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path

ROOT         = Path(__file__).parents[1]
XLSX_PATH    = ROOT / "data" / "markets" / "epicenter_mappings.xlsx"
DEFAULTS_PATH = ROOT / "data" / "markets" / "epicenter_attr_defaults.json"

SHEET_NAME   = "Опції атрибутів"
REQUIRED_COLS = {"attr_code", "attr_name_uk", "attr_type", "option_code",
                 "option_name_uk", "needs_default", "set_codes"}


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _norm_id(val: object) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if "." in s:
        try:
            f = float(s)
            if f == int(f):
                return str(int(f))
        except (ValueError, OverflowError):
            pass
    return s


def _is_truthy(val: object) -> bool:
    return str(val).strip().upper() in ("TRUE", "1", "YES")


# ─── Reader ───────────────────────────────────────────────────────────────────

def load_options_sheet() -> list[dict]:
    """
    Читає лист «Опції атрибутів» і повертає тільки рядки де needs_default=True.
    Повертає список dicts з полями: attr_code, attr_name, set_codes, option_code, option_name.
    """
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            wb.close()
            print(f"❌ Лист «{SHEET_NAME}» не знайдено в {XLSX_PATH}")
            print("   Спочатку виконай КРОК 5: python scripts/epicenter_export_attr_options.py")
            return []
        rows = list(wb[SHEET_NAME].iter_rows(values_only=True))
        wb.close()
    except FileNotFoundError:
        print(f"❌ Файл не знайдено: {XLSX_PATH}")
        return []
    except Exception as e:
        print(f"❌ Помилка читання xlsx: {e}")
        return []

    if not rows:
        return []

    headers = [str(c).strip() if c else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    missing = REQUIRED_COLS - idx.keys()
    if missing:
        print(f"❌ Відсутні колонки в «{SHEET_NAME}»: {missing}")
        return []

    result: list[dict] = []
    for row in rows[1:]:
        if not _is_truthy(row[idx["needs_default"]] if len(row) > idx["needs_default"] else ""):
            continue
        ac = _norm_id(row[idx["attr_code"]])
        if not ac:
            continue
        result.append({
            "attr_code":   ac,
            "attr_name":   str(row[idx["attr_name_uk"]] or "").strip(),
            "attr_type":   str(row[idx["attr_type"]]    or "").strip().lower(),
            "set_codes":   str(row[idx["set_codes"]]    or "").strip(),
            "option_code": _norm_id(row[idx["option_code"]]),
            "option_name": str(row[idx["option_name_uk"]] or "").strip(),
        })

    total_rows     = len(rows) - 1
    default_rows   = len(result)
    print(f"   Прочитано рядків: {total_rows}, з needs_default=True: {default_rows}")
    return result


# ─── Builder ──────────────────────────────────────────────────────────────────

def build_defaults_structure(
    rows: list[dict],
    existing: dict,
) -> dict:
    """
    Групує рядки по set_code → attr_code.
    Для кожного атрибута:
      - збирає доступні опції
      - зберігає вже вибране значення з existing (якщо є)
    """
    # set_code → attr_code → {attr_name, options: set of (code, name)}
    structure: dict[str, dict[str, dict]] = defaultdict(dict)

    for row in rows:
        ac        = row["attr_code"]
        name      = row["attr_name"]
        atype     = row["attr_type"]
        oc        = row["option_code"]
        on_       = row["option_name"]

        for sc in [s.strip() for s in row["set_codes"].split(",") if s.strip()]:
            if ac not in structure[sc]:
                structure[sc][ac] = {
                    "attr_name":         name,
                    "attr_type":         atype,
                    "available_options": [],
                    "_seen_codes":       set(),  # тимчасово для дедублікації
                }
            entry = structure[sc][ac]
            if oc and oc not in entry["_seen_codes"]:
                entry["_seen_codes"].add(oc)
                entry["available_options"].append({"code": oc, "name": on_})

    # Збираємо фінальний dict, підставляємо збережені значення з existing
    result: dict = {}
    for sc in sorted(structure, key=lambda x: (len(x), x)):
        result[sc] = {}
        existing_sc = existing.get(sc, {})
        for ac in sorted(structure[sc]):
            entry       = structure[sc][ac]
            existing_ac = existing_sc.get(ac, {})
            atype       = entry["attr_type"]

            # Зберігаємо вже вибрані дефолти (завжди список)
            saved_codes = existing_ac.get("default_option_codes", [])
            saved_names = existing_ac.get("default_option_names", [])

            # Міграція старого формату (рядок → список)
            if isinstance(saved_codes, str):
                saved_codes = [c.strip() for c in saved_codes.split(",") if c.strip()]
            if isinstance(saved_names, str):
                saved_names = [saved_names] if saved_names else []

            # Валідація: перевіряємо що збережені codes досі є в доступних опціях
            available_codes = {o["code"] for o in entry["available_options"]}
            invalid = [c for c in saved_codes if c not in available_codes]
            if invalid:
                print(f"   ⚠️  set={sc} attr={ac}: codes {invalid} не знайдено в опціях — скидаємо")
                saved_codes = []
                saved_names = []

            # Валідація: select повинен мати не більше 1 значення
            if atype == "select" and len(saved_codes) > 1:
                print(f"   ⚠️  set={sc} attr={ac}: select не може мати >1 дефолту, "
                      f"залишаємо перший: {saved_codes[0]}")
                saved_codes = [saved_codes[0]]
                saved_names = [saved_names[0]] if saved_names else []

            result[sc][ac] = {
                "attr_name":           entry["attr_name"],
                "attr_type":           atype,
                "default_option_codes": saved_codes,  # ← ЗАПОВНИТИ: 1 для select, 1+ для multiselect
                "default_option_names": saved_names,  # ← для зручності читання
                "available_options":    entry["available_options"],
            }

    return result


# ─── Stats ────────────────────────────────────────────────────────────────────

def print_stats(defaults: dict) -> None:
    total_sets  = len(defaults)
    total_attrs = sum(len(attrs) for attrs in defaults.values())
    select_total      = sum(1 for a in defaults.values() for e in a.values() if e["attr_type"] == "select")
    multiselect_total = sum(1 for a in defaults.values() for e in a.values() if e["attr_type"] == "multiselect")
    filled            = sum(1 for a in defaults.values() for e in a.values() if e["default_option_codes"])
    empty             = total_attrs - filled

    print(f"\n📊 Статистика дефолтів:")
    print(f"   set_codes:              {total_sets}")
    print(f"   attr (пар):             {total_attrs}")
    print(f"   • select:               {select_total}")
    print(f"   • multiselect:          {multiselect_total}")
    print(f"   ✅ заповнено:           {filled}")
    print(f"   🔴 порожньо:            {empty}")

    if empty:
        print(f"\n   Перші 10 незаповнених:")
        shown = 0
        for sc, attrs in defaults.items():
            for ac, entry in attrs.items():
                if not entry["default_option_codes"]:
                    opts_preview = ", ".join(
                        f"{o['code']}={o['name']}" for o in entry["available_options"][:3]
                    )
                    atype_marker = "[multi]" if entry["attr_type"] == "multiselect" else "[sel]  "
                    print(f"     {atype_marker} set={sc:>8}  attr={ac:<12}  "
                          f"{entry['attr_name']:<28}  options: [{opts_preview}]")
                    shown += 1
                    if shown >= 10:
                        break
            if shown >= 10:
                break


# ─── Main ─────────────────────────────────────────────────────────────────────

def main(dry_run: bool = False) -> None:
    print("🚀 epicenter_attr_defaults.py\n")

    if not XLSX_PATH.exists():
        print(f"❌ {XLSX_PATH} не знайдено.")
        print("   Виконай КРОКИ 1–5 пайплайну маппінгу.")
        return

    # Зчитуємо наявний файл дефолтів (щоб не затерти вже вибрані значення)
    existing: dict = {}
    if DEFAULTS_PATH.exists():
        try:
            existing = json.loads(DEFAULTS_PATH.read_text(encoding="utf-8"))
            filled_count = sum(
                1 for attrs in existing.values()
                for e in attrs.values()
                if e.get("default_option_codes")
            )
            print(f"   Існуючий файл: {DEFAULTS_PATH.name} "
                  f"({filled_count} вже заповнених дефолтів — збережемо)")
        except Exception as e:
            print(f"   ⚠️  Не вдалося прочитати існуючий файл: {e} — починаємо з нуля")

    rows = load_options_sheet()
    if not rows:
        return

    defaults = build_defaults_structure(rows, existing)
    print_stats(defaults)

    if dry_run:
        print("\n⏭️  dry-run: файл не записано.")
        return

    DEFAULTS_PATH.parent.mkdir(parents=True, exist_ok=True)
    DEFAULTS_PATH.write_text(
        json.dumps(defaults, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\n✅ Збережено: {DEFAULTS_PATH}")
    print(
        "\n📌 Наступний крок:\n"
        "   Відкрий epicenter_attr_defaults.json\n"
        "   Для кожного атрибута вибери опцію зі списку available_options\n"
        "   і заповни default_option_code + default_option_name.\n"
        "   Потім generate_epicenter_feed.py буде використовувати ці дефолти."
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Генерує заготовку дефолтних опцій атрибутів Epicenter")
    parser.add_argument("--dry-run", action="store_true", help="Показати статистику без запису файлу")
    args = parser.parse_args()
    main(dry_run=args.dry_run)
