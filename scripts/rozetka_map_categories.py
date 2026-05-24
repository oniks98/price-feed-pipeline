"""
rozetka_map_categories.py
-------------------------
Маппінг категорій Prom з аркуша "Маппінг" файлу rozetka_mappings.xlsx
на категорії Rozetka з аркуша "Категорії Розетки".

Консервативний режим за замовчуванням: вже заповнені рядки не перезаписуються.
Використовуй --overwrite для примусового перерахунку існуючих маппінгів.

Запуск:
    python scripts/rozetka_map_categories.py
    python scripts/rozetka_map_categories.py --overwrite
    python scripts/rozetka_map_categories.py --overwrite --dry-run

Оптимізації відносно v1:
    1. Інвертований токен-індекс — попередня фільтрація кандидатів до скорингу (O(R) → O(k)).
    2. Розбиття складових назв Prom по "," — кожна частина скориться окремо.
    3. Швидкий set-lookup у token_overlap перед fuzz.ratio.
    4. М'якший extra_penalty (множник 16 → 10) — менш агресивний на часткових збігах.
    5. Fuzzy-розширення індексу — токени Prom порівнюються з ключами індексу через fuzz.ratio,
       щоб близькі за значенням стеми потрапляли до кандидатів.
    6. Дедуплікація суфіксів у stem() + безпечний мінімальний залишок кореня.
    7. AUTO_THRESHOLD знижено з 88 → 85 (перевірено на граничних REVIEW-кейсах).
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from rapidfuzz import fuzz


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parents[1]
MAPPINGS_PATH = ROOT / "data" / "markets" / "rozetka_mappings.xlsx"

SHEET_MAPPING = "Маппінг"
SHEET_ROZETKA = "Категорії Розетки"

COL_PROM_ID = "prom_category_id"
COL_PROM_NAME = "Категорія Прому"
COL_ROZ_ID = "rozetka_category_id"
COL_ROZ_NAME = "Назва категорії Розетки"
COL_PARENT = "parentCode"

ROZ_COL_ID = "rozetka_category_id"
ROZ_COL_NAME = "Назва категорії Розетки"
ROZ_COL_PARENT = "parentCode"
ROZ_COL_PARENT_NAME = "Назва батьківської категорії"
ROZ_COL_LEVEL = "level"
ROZ_COL_LEAF = "is_leaf"
ROZ_COL_PATH = "Повний шлях категорії"

COL_MATCH_SCORE = "match_score"
COL_MATCH_STATUS = "match_status"
COL_MATCH_NOTE = "match_note"
COL_CANDIDATE_1 = "candidate_1"
COL_CANDIDATE_2 = "candidate_2"
COL_CANDIDATE_3 = "candidate_3"

# Lowered from 88 → 85: "Ручний обтискний інструмент" at 84.9 is a clear match.
AUTO_THRESHOLD = 85.0
REVIEW_THRESHOLD = 72.0

# How many token-index candidates to pre-select before full scoring.
# If fewer than MIN_CANDIDATES found → fall back to full category list.
MIN_CANDIDATES = 30

# Fuzz threshold for fuzzy token-index key expansion (stem similarity).
INDEX_FUZZY_THRESHOLD = 80

# High-confidence domain overrides for known generic Prom categories where
# fuzzy matching tends to choose too narrow Rozetka child categories.
MANUAL_OVERRIDES: dict[str, tuple[str, str]] = {
    "518": ("80100", "manual: broad acoustic systems"),
    "608": ("80158", "manual: broad vacuum cleaners"),
    "623": ("80070", "manual: TV remote controls"),
    "706": ("80089", "manual: broad monitors"),
    "1507": ("654239", "manual: batteries"),
    "3029": ("259632", "manual: broad intercoms"),
    "53004": ("4675025", "manual: electrical sockets"),
    "63022": ("80151", "manual: steam cleaners"),
    "50505": ("100957", "manual: metal detectors"),
    "70501": ("80045", "manual: USB flash drives"),
    "71109": ("387969", "manual: powerbanks"),
    "71902": ("80193", "manual: routers / network equipment"),
    "410402": ("152564", "manual: generators"),
    "141702": ("3510520", "manual: screwdriver bits"),
    "141703": ("154195", "manual: routers / milling tools"),
    "141714": ("4669237", "manual: impact wrenches"),
    "141717": ("152503", "manual: grinders"),
    "141718": ("156959", "manual: wall chasers"),
    "141726": ("200468", "manual: glue guns"),
    "14191106": ("80108", "manual: UPS"),
    "1250321": ("155089", "manual: trimmers and brushcutters"),
}

ACCESSORY_TOKENS = {
    "аксесуар",
    "аксесуари",
    "аксесуара",
    "аксесуарів",
    "запчастина",
    "запчастини",
    "комплектуючі",
    "комплектуючих",
    "кріплення",
    "рамки",
    "насадки",
    "чохли",
    "сумки",
}

NARROWING_TOKENS = {
    "промислові",
    "активні",
    "пасивні",
    "студійні",
    "автомобільні",
    "дитячі",
    "дівчаток",
    "хлопчиків",
    "військові",
    "манікюру",
    "педикюру",
    "косметики",
}

STOPWORDS = {
    "для",
    "та",
    "і",
    "й",
    "або",
    "до",
    "з",
    "зі",
    "із",
    "у",
    "в",
    "на",
    "по",
    "від",
    "при",
}

SYNONYM_REPLACEMENTS = {
    "дистанційного керування": "дк",
    "ду": "дк",
    "пз": "програмне забезпечення",
    "wi fi": "wifi",
    "usb накопичувачі": "флешки usb накопичувачі",
    "жорсткі диски": "hdd ssd жорсткі диски",
    # NEW: frequent mismatches between Prom and Rozetka terminology
    "полотна для лобзиків": "пилки для лобзиків",
    "пиляльні полотна": "пилки",
    "шліфувальні круги": "шліфувальні диски",
    "відрізні круги": "відрізні диски",
    "зачисні круги": "зачисні диски",
}

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class RozetkaCategory:
    category_id: str
    name: str
    parent_id: str
    parent_name: str
    level: int
    is_leaf: bool
    path: str
    name_norm: str
    path_norm: str
    name_tokens: tuple[str, ...]
    path_tokens: tuple[str, ...]


@dataclass(frozen=True)
class Candidate:
    category: RozetkaCategory
    score: float
    name_score: float
    context_score: float
    penalty: float
    reason: str


@dataclass(frozen=True)
class MatchDecision:
    status: str
    selected: Candidate | None
    candidates: tuple[Candidate, ...]
    note: str


# ---------------------------------------------------------------------------
# Text helpers
# ---------------------------------------------------------------------------


def norm_id(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    if number.is_integer():
        return str(int(number))
    return text


def as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"true", "1", "yes", "y", "так"}


def split_path(text: str) -> list[str]:
    return [part.strip() for part in str(text or "").split(">") if part and part.strip()]


def last_segment(text: str) -> str:
    parts = split_path(text)
    return parts[-1] if parts else str(text or "").strip()


def normalize(text: str) -> str:
    value = str(text or "").lower().replace("\u2019", " ").replace("'", " ")
    for source, target in SYNONYM_REPLACEMENTS.items():
        value = re.sub(rf"\b{re.escape(source)}\b", target, value, flags=re.IGNORECASE)
    value = re.sub(r"[^\w\s]", " ", value, flags=re.UNICODE)
    return re.sub(r"\s+", " ", value).strip()


# Deduplicated, ordered by length descending (critical for greedy matching).
_STEM_SUFFIXES: tuple[str, ...] = tuple(
    sorted(
        {
            "ування", "ювання", "ання", "ення", "іння",
            "ський", "цький", "зький",
            "ними", "ному", "ного", "ної", "ній",
            "ові", "ева", "ями", "ами",
            "ого", "ому", "ими", "их",
            "ою", "ею", "ів", "ов", "ев",
            "ий", "ій",
            "а", "и", "і", "у", "я", "ю", "е",
        },
        key=len,
        reverse=True,
    )
)


def stem(word: str) -> str:
    """Light Ukrainian stemmer — strips inflectional suffixes."""
    for suffix in _STEM_SUFFIXES:
        if word.endswith(suffix) and len(word) - len(suffix) >= 4:
            return word[: -len(suffix)]
    return word


def tokens(text: str) -> tuple[str, ...]:
    result: list[str] = []
    for token in normalize(text).split():
        if token in STOPWORDS or len(token) < 2:
            continue
        result.append(stem(token))
    return tuple(result)


def token_overlap(
    query_tokens: tuple[str, ...],
    target_tokens: tuple[str, ...],
) -> float:
    """
    Fraction of query_tokens covered by target_tokens.
    Fast path: exact stem match via set lookup.
    Slow path: fuzz.ratio only for unmatched tokens.
    """
    if not query_tokens or not target_tokens:
        return 0.0

    target_set = set(target_tokens)
    matched = 0

    for query in query_tokens:
        if query in target_set:
            # O(1) exact match — no fuzz needed.
            matched += 1
        elif any(fuzz.ratio(query, t) >= INDEX_FUZZY_THRESHOLD for t in target_tokens):
            matched += 1

    return matched / len(query_tokens) * 100.0


def format_candidate(candidate: Candidate | None) -> str:
    if candidate is None:
        return ""
    cat = candidate.category
    return (
        f"{cat.category_id} | {cat.name} | score={candidate.score:.1f} | "
        f"parent={cat.parent_id} | path={cat.path}"
    )


# ---------------------------------------------------------------------------
# Inverted token index
# ---------------------------------------------------------------------------


def build_token_index(
    categories: list[RozetkaCategory],
) -> dict[str, list[str]]:
    """
    Maps every stem token (from name + path) → list of category_ids.
    Used to pre-filter candidates before expensive full scoring.
    """
    index: defaultdict[str, list[str]] = defaultdict(list)
    for cat in categories:
        seen: set[str] = set()
        for token in cat.name_tokens + cat.path_tokens:
            if token not in seen:
                index[token].append(cat.category_id)
                seen.add(token)
    log.debug("Token index built: %d unique stems.", len(index))
    return dict(index)


def get_candidates(
    prom_tokens: tuple[str, ...],
    index: dict[str, list[str]],
    by_id: dict[str, RozetkaCategory],
    all_categories: list[RozetkaCategory],
) -> list[RozetkaCategory]:
    """
    Fast candidate pre-selection:
      1. Exact stem lookup in index.
      2. If still < MIN_CANDIDATES, fuzzy-expand against all index keys.
      3. If still < MIN_CANDIDATES, fall back to full category list.
    """
    candidate_ids: set[str] = set()

    # Step 1: exact stem hits.
    for token in prom_tokens:
        for cid in index.get(token, []):
            candidate_ids.add(cid)

    # Step 2: fuzzy expansion against index keys (handles synonym stems).
    if len(candidate_ids) < MIN_CANDIDATES:
        index_keys = list(index.keys())
        for token in prom_tokens:
            for key in index_keys:
                if key not in candidate_ids and fuzz.ratio(token, key) >= INDEX_FUZZY_THRESHOLD:
                    for cid in index[key]:
                        candidate_ids.add(cid)

    # Step 3: hard fallback.
    if len(candidate_ids) < MIN_CANDIDATES:
        log.debug(
            "Candidate index too small (%d), falling back to full list.", len(candidate_ids)
        )
        return all_categories

    return [by_id[cid] for cid in candidate_ids if cid in by_id]


# ---------------------------------------------------------------------------
# Composite name splitting
# ---------------------------------------------------------------------------


def split_composite_name(name: str) -> list[str]:
    """
    "Відрізні, зачисні, шліфувальні, пиляльні круги" →
    ["Відрізні, зачисні, шліфувальні, пиляльні круги",
     "Відрізні", "зачисні", "шліфувальні", "пиляльні круги"]

    The full name is always tried first.
    Individual comma-parts are added only if they are meaningful (> 3 chars).
    This lets us find narrow rozetka sub-categories even from long compound prom names.
    """
    parts = [p.strip() for p in name.split(",")]
    if len(parts) == 1:
        return [name]
    # Full name + each non-trivial part.
    meaningful = [p for p in parts if len(p) > 3]
    return [name] + meaningful


# ---------------------------------------------------------------------------
# Workbook loading
# ---------------------------------------------------------------------------


def headers(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): index
        for index, cell in enumerate(ws[1], start=1)
        if cell.value is not None and str(cell.value).strip()
    }


def require_columns(
    found: dict[str, int], required: tuple[str, ...], sheet_name: str
) -> None:
    missing = [column for column in required if column not in found]
    if missing:
        raise ValueError(
            f"Missing columns in '{sheet_name}': {missing}. Found: {list(found)}"
        )


def ensure_column(ws, found: dict[str, int], name: str) -> int:
    if name in found:
        return found[name]
    next_col = max(found.values(), default=0) + 1
    ws.cell(row=1, column=next_col, value=name)
    found[name] = next_col
    return next_col


def load_rozetka_categories(ws) -> list[RozetkaCategory]:
    found = headers(ws)
    require_columns(
        found,
        (
            ROZ_COL_ID, ROZ_COL_NAME, ROZ_COL_PARENT,
            ROZ_COL_PARENT_NAME, ROZ_COL_LEVEL, ROZ_COL_LEAF, ROZ_COL_PATH,
        ),
        SHEET_ROZETKA,
    )

    categories: list[RozetkaCategory] = []
    for row in ws.iter_rows(min_row=2, values_only=True):

        def value(column: str) -> Any:
            index = found[column] - 1
            return row[index] if index < len(row) else None

        category_id = norm_id(value(ROZ_COL_ID))
        name = str(value(ROZ_COL_NAME) or "").strip()
        if not category_id or not name:
            continue

        parent_id = norm_id(value(ROZ_COL_PARENT))
        path = str(value(ROZ_COL_PATH) or name).strip()
        try:
            level = int(value(ROZ_COL_LEVEL) or 0)
        except (TypeError, ValueError):
            level = 0

        categories.append(
            RozetkaCategory(
                category_id=category_id,
                name=name,
                parent_id=parent_id,
                parent_name=str(value(ROZ_COL_PARENT_NAME) or "").strip(),
                level=level,
                is_leaf=as_bool(value(ROZ_COL_LEAF)),
                path=path,
                name_norm=normalize(name),
                path_norm=normalize(path),
                name_tokens=tokens(name),
                path_tokens=tokens(path),
            )
        )

    log.info("Loaded %d Rozetka categories.", len(categories))
    return categories


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------


def score_name(
    prom_name: str,
    prom_tokens: tuple[str, ...],
    category: RozetkaCategory,
) -> float:
    prom_norm = normalize(prom_name)
    if prom_norm and prom_norm == category.name_norm:
        return 100.0

    forward = token_overlap(prom_tokens, category.name_tokens)
    reverse = token_overlap(category.name_tokens, prom_tokens)
    ratio = float(fuzz.ratio(prom_norm, category.name_norm))
    token_sort = float(fuzz.token_sort_ratio(prom_norm, category.name_norm))

    coverage = 0.58 * forward + 0.42 * reverse
    return max(ratio, token_sort, coverage)


def score_context(
    prom_path: str,
    prom_context_tokens: tuple[str, ...],
    category: RozetkaCategory,
) -> float:
    if not prom_context_tokens:
        return 0.0
    overlap = token_overlap(prom_context_tokens, category.path_tokens)
    path_ratio = float(fuzz.partial_ratio(normalize(prom_path), category.path_norm))
    return max(overlap, path_ratio * 0.75)


def score_penalty(
    prom_tokens: tuple[str, ...],
    category: RozetkaCategory,
    *,
    exact_name: bool,
) -> tuple[float, list[str]]:
    penalty = 0.0
    reasons: list[str] = []

    extra_tokens = set(category.name_tokens) - set(prom_tokens)
    if extra_tokens and not exact_name:
        extra_ratio = len(extra_tokens) / max(len(category.name_tokens), 1)
        # v1 used 16.0 — too aggressive. 10.0 allows partial matches to surface.
        extra_penalty = min(15.0, extra_ratio * 10.0)
        penalty += extra_penalty
        reasons.append(f"extra-name-tokens={len(extra_tokens)}")

    if extra_tokens & NARROWING_TOKENS and not exact_name:
        penalty += 12.0
        reasons.append("narrow-child")

    prom_has_accessory = bool(set(prom_tokens) & ACCESSORY_TOKENS)
    target_has_accessory = bool(set(category.name_tokens) & ACCESSORY_TOKENS)
    if target_has_accessory and not prom_has_accessory:
        penalty += 18.0
        reasons.append("accessory-target")

    return penalty, reasons


def score_candidate(
    prom_path: str,
    prom_name_variant: str,
    category: RozetkaCategory,
) -> Candidate:
    """Score a single (prom_name_variant, rozetka_category) pair."""
    prom_tokens_var = tokens(prom_name_variant)
    context_text = " > ".join(split_path(prom_path)[:-1])
    context_tokens = tokens(context_text)

    exact_name = normalize(prom_name_variant) == category.name_norm
    name_s = score_name(prom_name_variant, prom_tokens_var, category)
    context_s = score_context(prom_path, context_tokens, category)
    penalty, penalty_reasons = score_penalty(prom_tokens_var, category, exact_name=exact_name)

    score = (0.78 * name_s) + (0.22 * context_s) - penalty
    if exact_name:
        score = max(score, 96.0 + min(context_s, 50.0) * 0.08)

    score = max(0.0, min(100.0, score))
    reason_parts = [
        "exact" if exact_name else "fuzzy",
        f"name={name_s:.1f}",
        f"ctx={context_s:.1f}",
    ]
    if penalty:
        reason_parts.append(f"pen={penalty:.1f}")
        reason_parts.extend(penalty_reasons)

    return Candidate(
        category=category,
        score=score,
        name_score=name_s,
        context_score=context_s,
        penalty=penalty,
        reason=", ".join(reason_parts),
    )


def rank_candidates(
    prom_path: str,
    all_categories: list[RozetkaCategory],
    index: dict[str, list[str]],
    by_id: dict[str, RozetkaCategory],
) -> tuple[Candidate, ...]:
    """
    1. Split composite prom name into variants.
    2. Pre-filter candidates via inverted token index per variant.
    3. Score each (variant, candidate) pair.
    4. Keep the best score per category across all variants.
    5. Return top-5 sorted by score.
    """
    prom_name = last_segment(prom_path)
    name_variants = split_composite_name(prom_name)

    # Best score per rozetka category_id across all name variants.
    best_per_category: dict[str, Candidate] = {}

    for variant in name_variants:
        variant_tokens = tokens(variant)
        candidates = get_candidates(variant_tokens, index, by_id, all_categories)

        for category in candidates:
            candidate = score_candidate(prom_path, variant, category)
            existing = best_per_category.get(category.category_id)
            if existing is None or candidate.score > existing.score:
                best_per_category[category.category_id] = candidate

    ranked = sorted(
        best_per_category.values(),
        key=lambda c: (c.score, c.name_score, -c.penalty, -c.category.level),
        reverse=True,
    )
    return tuple(ranked[:5])


def select_match(
    prom_id: str,
    prom_path: str,
    all_categories: list[RozetkaCategory],
    by_id: dict[str, RozetkaCategory],
    index: dict[str, list[str]],
) -> MatchDecision:
    if prom_id in MANUAL_OVERRIDES:
        rozetka_id, note = MANUAL_OVERRIDES[prom_id]
        category = by_id.get(rozetka_id)
        if category is not None:
            candidate = Candidate(
                category=category,
                score=100.0,
                name_score=100.0,
                context_score=100.0,
                penalty=0.0,
                reason=note,
            )
            return MatchDecision("OVERRIDE", candidate, (candidate,), note)

    candidates = rank_candidates(prom_path, all_categories, index, by_id)
    best = candidates[0] if candidates else None
    if best is None:
        return MatchDecision("NO_MATCH", None, (), "no candidates")

    if best.score >= AUTO_THRESHOLD:
        return MatchDecision("AUTO", best, candidates, best.reason)

    if best.score >= REVIEW_THRESHOLD:
        return MatchDecision("REVIEW", None, candidates, f"review: {best.reason}")

    return MatchDecision("NO_MATCH", None, candidates, f"low score: {best.reason}")


# ---------------------------------------------------------------------------
# Mapping update
# ---------------------------------------------------------------------------


def clear_mapping_cells(ws, row: int, columns: tuple[int, int, int]) -> None:
    for column in columns:
        ws.cell(row=row, column=column).value = None


def write_decision(
    ws,
    row: int,
    decision: MatchDecision,
    columns: dict[str, int],
) -> None:
    selected = decision.selected

    if selected is not None:
        category = selected.category
        ws.cell(row=row, column=columns[COL_ROZ_ID]).value = int(category.category_id)
        ws.cell(row=row, column=columns[COL_ROZ_NAME]).value = category.name
        ws.cell(row=row, column=columns[COL_PARENT]).value = (
            int(category.parent_id) if category.parent_id else None
        )
        score = selected.score
    else:
        clear_mapping_cells(
            ws,
            row,
            (columns[COL_ROZ_ID], columns[COL_ROZ_NAME], columns[COL_PARENT]),
        )
        score = decision.candidates[0].score if decision.candidates else 0.0

    ws.cell(row=row, column=columns[COL_MATCH_SCORE]).value = round(score, 1)
    ws.cell(row=row, column=columns[COL_MATCH_STATUS]).value = decision.status
    ws.cell(row=row, column=columns[COL_MATCH_NOTE]).value = decision.note

    for index_i, column_name in enumerate(
        (COL_CANDIDATE_1, COL_CANDIDATE_2, COL_CANDIDATE_3),
        start=0,
    ):
        candidate = decision.candidates[index_i] if index_i < len(decision.candidates) else None
        ws.cell(row=row, column=columns[column_name]).value = format_candidate(candidate)


def map_categories(*, overwrite: bool, dry_run: bool) -> dict[str, int]:
    if not MAPPINGS_PATH.exists():
        raise FileNotFoundError(f"Mappings file not found: {MAPPINGS_PATH}")

    workbook = openpyxl.load_workbook(MAPPINGS_PATH)
    try:
        for sheet in (SHEET_MAPPING, SHEET_ROZETKA):
            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found. Available: {workbook.sheetnames}")

        ws_map = workbook[SHEET_MAPPING]
        ws_roz = workbook[SHEET_ROZETKA]

        # Load once — build index once.
        all_categories = load_rozetka_categories(ws_roz)
        by_id = {cat.category_id: cat for cat in all_categories}
        index = build_token_index(all_categories)

        map_headers = headers(ws_map)
        require_columns(
            map_headers,
            (COL_PROM_ID, COL_PROM_NAME, COL_ROZ_ID, COL_ROZ_NAME, COL_PARENT),
            SHEET_MAPPING,
        )

        for column in (
            COL_MATCH_SCORE, COL_MATCH_STATUS, COL_MATCH_NOTE,
            COL_CANDIDATE_1, COL_CANDIDATE_2, COL_CANDIDATE_3,
        ):
            ensure_column(ws_map, map_headers, column)

        stats: dict[str, int] = {
            "AUTO": 0, "OVERRIDE": 0, "REVIEW": 0,
            "NO_MATCH": 0, "EXISTING": 0, "EMPTY_PROM": 0,
        }

        for row_index in range(2, ws_map.max_row + 1):
            prom_id = norm_id(ws_map.cell(row=row_index, column=map_headers[COL_PROM_ID]).value)
            prom_path = str(
                ws_map.cell(row=row_index, column=map_headers[COL_PROM_NAME]).value or ""
            ).strip()

            if not prom_path:
                stats["EMPTY_PROM"] += 1
                continue

            existing_id = norm_id(ws_map.cell(row=row_index, column=map_headers[COL_ROZ_ID]).value)
            existing_name = str(
                ws_map.cell(row=row_index, column=map_headers[COL_ROZ_NAME]).value or ""
            ).strip()
            existing_parent = norm_id(
                ws_map.cell(row=row_index, column=map_headers[COL_PARENT]).value
            )

            if existing_id and existing_name and existing_parent and not overwrite:
                stats["EXISTING"] += 1
                continue

            decision = select_match(prom_id, prom_path, all_categories, by_id, index)
            stats[decision.status] += 1
            write_decision(ws_map, row_index, decision, map_headers)

            selected = decision.selected
            selected_label = selected.category.name if selected else "no write"
            log.info(
                "row=%d prom_id=%s status=%s score=%.1f: %s -> %s",
                row_index,
                prom_id,
                decision.status,
                (
                    selected.score
                    if selected
                    else (decision.candidates[0].score if decision.candidates else 0.0)
                ),
                last_segment(prom_path),
                selected_label,
            )

        if dry_run:
            log.info("Dry-run: workbook not saved.")
        else:
            workbook.save(MAPPINGS_PATH)
            log.info("Saved -> %s", MAPPINGS_PATH)

        return stats
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Map Prom categories to Rozetka categories")
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Recompute rows that already have rozetka_category_id/name/parentCode",
    )
    parser.add_argument("--dry-run", action="store_true", help="Do not save workbook")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv or sys.argv[1:])
    stats = map_categories(overwrite=args.overwrite, dry_run=args.dry_run)
    log.info(
        "Done. AUTO=%d | OVERRIDE=%d | REVIEW=%d | NO_MATCH=%d | EXISTING=%d | EMPTY_PROM=%d",
        stats["AUTO"],
        stats["OVERRIDE"],
        stats["REVIEW"],
        stats["NO_MATCH"],
        stats["EXISTING"],
        stats["EMPTY_PROM"],
    )


if __name__ == "__main__":
    main()