"""
change_image.py

Algorithm:
    For every row where Ідентифікатор_товару starts with "prom_",
    find the matching base row (same ID without the "prom_" prefix)
    and copy its Посилання_зображення value into the prom_ row.

Usage:
    python scripts/change_image.py
    python scripts/change_image.py --input data/markets/export-products.xlsx
    python scripts/change_image.py --input data/markets/export-products.xlsx --output data/markets/export-products-fixed.xlsx
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

DEFAULT_INPUT = Path(r"C:\FullStack\PriceFeedPipeline\data\markets\export-products.xlsx")
PROM_PREFIX = "prom_"
COL_ID = "Ідентифікатор_товару"
COL_IMAGE = "Посилання_зображення"

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def get_column_map(ws: Worksheet) -> dict[str, int]:
    """Return {header_name: col_index_1based} from the first row."""
    return {
        cell.value: cell.column
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
        if cell.value is not None
    }


def build_base_image_index(
    ws: Worksheet,
    col_id: int,
    col_image: int,
) -> dict[str, str | None]:
    """
    Scan all rows and collect {product_id: image_value}
    only for rows that do NOT have the prom_ prefix.
    """
    index: dict[str, str | None] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_id = row[col_id - 1]
        if raw_id is None:
            continue
        product_id = str(raw_id).strip()
        if product_id.startswith(PROM_PREFIX):
            continue
        image = row[col_image - 1]
        index[product_id] = str(image).strip() if image is not None else None

    log.info("Base index built: %d unique product IDs", len(index))
    return index


def patch_prom_rows(
    ws: Worksheet,
    col_id: int,
    col_image: int,
    base_index: dict[str, str | None],
) -> tuple[int, int, int]:
    """
    Iterate prom_ rows and overwrite Посилання_зображення from base_index.

    Returns:
        (patched, skipped_no_match, skipped_no_image)
    """
    patched = skipped_no_match = skipped_no_image = 0

    for row_cells in ws.iter_rows(min_row=2):
        id_cell = row_cells[col_id - 1]
        raw_id = id_cell.value
        if raw_id is None:
            continue

        product_id = str(raw_id).strip()
        if not product_id.startswith(PROM_PREFIX):
            continue

        base_id = product_id[len(PROM_PREFIX):]

        if base_id not in base_index:
            log.warning("No base row found for prom_ ID: %s (base_id=%s)", product_id, base_id)
            skipped_no_match += 1
            continue

        image_value = base_index[base_id]
        if not image_value:
            log.warning("Base row for %s has empty image — skipping", base_id)
            skipped_no_image += 1
            continue

        image_cell = row_cells[col_image - 1]
        image_cell.value = image_value
        patched += 1
        log.debug("Patched %s ← %s", product_id, image_value)

    return patched, skipped_no_match, skipped_no_image


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Copy images from base rows to prom_ rows.")
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="Path to the source xlsx file.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Path to save the result (defaults to overwriting input).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path: Path = args.input
    output_path: Path = args.output or input_path

    if not input_path.exists():
        log.error("Input file not found: %s", input_path)
        sys.exit(1)

    log.info("Loading workbook: %s", input_path)
    wb = openpyxl.load_workbook(input_path)
    ws: Worksheet = wb.active  # type: ignore[assignment]

    # --- resolve columns ---
    col_map = get_column_map(ws)

    missing = [c for c in (COL_ID, COL_IMAGE) if c not in col_map]
    if missing:
        log.error("Required columns not found in sheet: %s", missing)
        log.error("Available columns: %s", list(col_map.keys()))
        sys.exit(1)

    col_id = col_map[COL_ID]
    col_image = col_map[COL_IMAGE]
    log.info("Column '%s' → index %d", COL_ID, col_id)
    log.info("Column '%s' → index %d", COL_IMAGE, col_image)

    # --- build lookup ---
    base_index = build_base_image_index(ws, col_id, col_image)

    # --- patch ---
    patched, no_match, no_image = patch_prom_rows(ws, col_id, col_image, base_index)

    # --- save ---
    wb.save(output_path)
    log.info("Saved: %s", output_path)
    log.info(
        "Done — patched: %d | no base row: %d | empty image: %d",
        patched,
        no_match,
        no_image,
    )


if __name__ == "__main__":
    main()
